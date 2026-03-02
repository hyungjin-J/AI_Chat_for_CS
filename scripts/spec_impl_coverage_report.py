#!/usr/bin/env python3
"""Build deterministic Spec -> Implementation coverage artifacts."""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


KOR_API_LIST = "\uc804\uccb4API\ubaa9\ub85d"
KOR_PROGRAM_ID = "\ud504\ub85c\uadf8\ub7a8ID"
KOR_NOTE = "\ube44\uace0"
KOR_IMPORTANCE = "\uc911\uc694\ub3c4"

REQID_SEGMENT_RE = re.compile(r"ReqID\+?\s*:\s*([^\n\r;]+)", re.IGNORECASE)
REQID_TOKEN_RE = re.compile(r"\b[A-Z]{2,5}-\d{3}\b")

CLASS_DECL_RE = re.compile(r"\bclass\s+([A-Za-z_][A-Za-z0-9_]*)\b")
ANNOTATION_NAME_RE = re.compile(r"^\s*@([A-Za-z_][A-Za-z0-9_$.]*)")
REQUEST_METHOD_RE = re.compile(r"RequestMethod\.(GET|POST|PUT|DELETE|PATCH|HEAD|OPTIONS)", re.IGNORECASE)
STRING_LITERAL_RE = re.compile(r"\"((?:[^\"\\]|\\.)*)\"")

HTTP_METHODS = ("GET", "POST", "PUT", "DELETE", "PATCH", "HEAD", "OPTIONS")
HTTP_MAPPING_NAME_TO_METHOD = {
    "getmapping": ("GET",),
    "postmapping": ("POST",),
    "putmapping": ("PUT",),
    "deletemapping": ("DELETE",),
    "patchmapping": ("PATCH",),
    "requestmapping": (),
}

IMPORTANCE_MUST = "Must"
IMPORTANCE_SHOULD = "Should"
IMPORTANCE_UNKNOWN = "Unknown/Noncritical"

DEFAULT_REQUIREMENTS = "docs/references/CS AI Chatbot_Requirements Statement.csv"
DEFAULT_API_WORKBOOK = "docs/references/google_ready_api_spec_v0.3_20260216.xlsx"
DEFAULT_API_SHEET = KOR_API_LIST
DEFAULT_BACKEND_ROOT = "backend/src/main/java"
DEFAULT_BACKEND_TEST_ROOT = "backend/src/test"
DEFAULT_FRONTEND_ROOT = "frontend/src"
DEFAULT_REPORT_TXT = "docs/review/mvp_verification_pack/artifacts/spec_impl_coverage_report.txt"
DEFAULT_REPORT_JSON = "docs/review/mvp_verification_pack/artifacts/spec_impl_coverage_report.json"
DEFAULT_REPORT_MD = "docs/review/mvp_verification_pack/artifacts/spec_impl_coverage_report.md"


@dataclass(frozen=True)
class ApiSpecRow:
    row_index: int
    program_id: str
    method: str
    endpoint: str
    endpoint_normalized: str
    req_ids: list[str]
    importance: str
    req_importance_breakdown: dict[str, str]


@dataclass(frozen=True)
class SourceMatch:
    path: str
    line: int
    snippet: str


@dataclass(frozen=True)
class ControllerMapping:
    file_path: str
    method: str
    endpoint_normalized: str
    line: int
    snippet: str
    controller_class: str


def normalize(path: str) -> str:
    return path.replace("\\", "/").strip()


def to_rel(path: Path, root: Path) -> str:
    try:
        return normalize(path.resolve().relative_to(root.resolve()).as_posix())
    except ValueError:
        return normalize(path.resolve().as_posix())


def run_git(root: Path, args: list[str]) -> str:
    proc = subprocess.run(
        ["git", *args],
        cwd=root,
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    value = (proc.stdout or "").strip()
    if proc.returncode != 0 or not value:
        return "UNKNOWN"
    return value


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build deterministic spec implementation coverage report")
    parser.add_argument("--root", default=".")
    parser.add_argument("--requirements", default=DEFAULT_REQUIREMENTS)
    parser.add_argument("--api-workbook", default=DEFAULT_API_WORKBOOK)
    parser.add_argument("--api-sheet", default=DEFAULT_API_SHEET)
    parser.add_argument("--backend-root", default=DEFAULT_BACKEND_ROOT)
    parser.add_argument("--backend-test-root", default=DEFAULT_BACKEND_TEST_ROOT)
    parser.add_argument("--frontend-root", default=DEFAULT_FRONTEND_ROOT)
    parser.add_argument("--report-txt", default=DEFAULT_REPORT_TXT)
    parser.add_argument("--report-json", default=DEFAULT_REPORT_JSON)
    parser.add_argument("--report-md", default=DEFAULT_REPORT_MD)
    parser.add_argument("--max-matches-per-signal", type=int, default=5)
    return parser.parse_args()


def resolve_path(root: Path, value: str) -> Path:
    candidate = Path(value)
    if candidate.is_absolute():
        return candidate
    return (root / candidate).resolve()


def read_csv_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file_pointer:
        rows = list(csv.reader(file_pointer))
    if not rows:
        return [], []
    return rows[0], rows[1:]


def find_column(headers: list[str], candidates: list[str]) -> int | None:
    normalized = [header.strip().lower() for header in headers]
    normalized_candidates = [candidate.strip().lower() for candidate in candidates]
    for candidate in normalized_candidates:
        if candidate in normalized:
            return normalized.index(candidate)
    for candidate in normalized_candidates:
        for index, header in enumerate(normalized):
            if candidate and candidate in header:
                return index
    return None


def normalize_importance(value: str) -> str:
    raw = value.strip().lower()
    if not raw:
        return IMPORTANCE_UNKNOWN
    if raw.startswith("must") or raw == "\ud544\uc218":
        return IMPORTANCE_MUST
    if raw.startswith("should") or raw == "\uad8c\uc7a5":
        return IMPORTANCE_SHOULD
    return IMPORTANCE_UNKNOWN


def load_requirements_importance(requirements_path: Path) -> dict[str, str]:
    headers, rows = read_csv_rows(requirements_path)
    if not headers:
        raise ValueError(f"requirements CSV is empty: {requirements_path.as_posix()}")

    reqid_col = find_column(headers, ["ReqID", "\uc694\uad6c\uc0ac\ud56dID"])
    importance_col = find_column(headers, [KOR_IMPORTANCE, "importance", "priority"])
    if reqid_col is None or importance_col is None:
        raise ValueError("requirements CSV missing ReqID/importance columns")

    mapping: dict[str, str] = {}
    for row in rows:
        reqid = row[reqid_col].strip() if reqid_col < len(row) and row[reqid_col] else ""
        if not REQID_TOKEN_RE.fullmatch(reqid):
            continue
        raw_importance = row[importance_col] if importance_col < len(row) else ""
        mapping[reqid] = normalize_importance(str(raw_importance))
    return mapping


def parse_reqids_from_note(note: str) -> list[str]:
    if not note:
        return []
    segments = REQID_SEGMENT_RE.findall(note)
    if not segments and "reqid" in note.lower():
        segments = [note]
    tokens: set[str] = set()
    for segment in segments:
        tokens.update(REQID_TOKEN_RE.findall(segment.upper()))
    return sorted(tokens)


def classify_api_importance(req_ids: list[str], req_importance: dict[str, str]) -> tuple[str, dict[str, str]]:
    if not req_ids:
        return IMPORTANCE_UNKNOWN, {}

    breakdown: dict[str, str] = {}
    resolved: list[str] = []
    for req_id in req_ids:
        value = req_importance.get(req_id, IMPORTANCE_UNKNOWN)
        breakdown[req_id] = value
        resolved.append(value)

    if any(item == IMPORTANCE_MUST for item in resolved):
        return IMPORTANCE_MUST, breakdown
    if resolved and all(item == IMPORTANCE_SHOULD for item in resolved):
        return IMPORTANCE_SHOULD, breakdown
    return IMPORTANCE_UNKNOWN, breakdown


def select_api_sheet(workbook_path: Path, preferred_sheet_name: str):
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    preferred_normalized = preferred_sheet_name.strip().lower()

    for worksheet in workbook.worksheets:
        if worksheet.title.strip().lower() == preferred_normalized:
            return workbook, worksheet

    for worksheet in workbook.worksheets:
        lowered = worksheet.title.strip().lower()
        if "api" in lowered and ("list" in lowered or "\ubaa9\ub85d" in worksheet.title):
            return workbook, worksheet

    workbook.close()
    raise ValueError(f"API sheet '{preferred_sheet_name}' not found")


def normalize_endpoint_template(path: str) -> str:
    raw = path.strip()
    if not raw:
        return "/"
    for token in ("?", "#"):
        if token in raw:
            raw = raw.split(token, 1)[0]
    raw = "/" + raw.lstrip("/")
    raw = re.sub(r"/{2,}", "/", raw)
    raw = re.sub(r"\{[^/{}]+\}", "{}", raw)
    if len(raw) > 1 and raw.endswith("/"):
        raw = raw[:-1]
    return raw


def load_api_rows(api_workbook_path: Path, api_sheet: str, req_importance: dict[str, str]) -> list[ApiSpecRow]:
    workbook, worksheet = select_api_sheet(api_workbook_path, api_sheet)
    try:
        header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
        headers = [str(value).strip() if value is not None else "" for value in header_values]

        method_col = find_column(headers, ["Method", "http_method"])
        endpoint_col = find_column(headers, ["Endpoint", "path", "uri"])
        program_col = find_column(headers, [KOR_PROGRAM_ID, "ProgramID", "Program ID", "program_id"])
        note_col = find_column(headers, [KOR_NOTE, "remark", "note"])

        if method_col is None or endpoint_col is None:
            raise ValueError("API sheet missing Method/Endpoint columns")

        rows: list[ApiSpecRow] = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            method = (
                str(values[method_col]).strip().upper()
                if method_col < len(values) and values[method_col] is not None
                else ""
            )
            endpoint = (
                str(values[endpoint_col]).strip()
                if endpoint_col < len(values) and values[endpoint_col] is not None
                else ""
            )
            if not method or not endpoint or method not in HTTP_METHODS:
                continue

            note = (
                str(values[note_col]).strip()
                if note_col is not None and note_col < len(values) and values[note_col] is not None
                else ""
            )
            req_ids = parse_reqids_from_note(note)
            importance, breakdown = classify_api_importance(req_ids, req_importance)

            program_id = ""
            if program_col is not None and program_col < len(values) and values[program_col] is not None:
                program_id = str(values[program_col]).strip()
            if not program_id:
                program_id = f"ROW-{row_number}"

            rows.append(
                ApiSpecRow(
                    row_index=row_number,
                    program_id=program_id,
                    method=method,
                    endpoint=endpoint,
                    endpoint_normalized=normalize_endpoint_template(endpoint),
                    req_ids=req_ids,
                    importance=importance,
                    req_importance_breakdown=breakdown,
                )
            )
        return sorted(rows, key=lambda item: (item.row_index, item.method, item.endpoint))
    finally:
        workbook.close()


def extract_annotation(lines: list[str], start_index: int) -> tuple[str, int]:
    current_index = start_index
    parts = [lines[current_index].rstrip("\n")]
    balance = parts[0].count("(") - parts[0].count(")")
    while balance > 0 and current_index + 1 < len(lines):
        current_index += 1
        line = lines[current_index].rstrip("\n")
        parts.append(line)
        balance += line.count("(") - line.count(")")
    return "\n".join(parts), current_index


def annotation_simple_name(annotation: str) -> str:
    matched = ANNOTATION_NAME_RE.match(annotation)
    if not matched:
        return ""
    return matched.group(1).split(".")[-1]


def annotation_args(annotation: str) -> str:
    open_index = annotation.find("(")
    close_index = annotation.rfind(")")
    if open_index < 0 or close_index <= open_index:
        return ""
    return annotation[open_index + 1 : close_index]


def parse_java_string_literals(text: str) -> list[str]:
    return [value.replace('\\"', '"') for value in STRING_LITERAL_RE.findall(text)]


def dedupe_keep_order(values: Iterable[str]) -> list[str]:
    deduped: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value in seen:
            continue
        deduped.append(value)
        seen.add(value)
    return deduped


def extract_mapping_paths(args: str) -> list[str]:
    stripped = args.strip()
    if not stripped:
        return [""]

    flattened: list[str] = []
    for key in ("path", "value"):
        for match in re.finditer(rf"\b{key}\s*=\s*\{{([^}}]*)\}}", stripped, flags=re.DOTALL):
            flattened.extend(parse_java_string_literals(match.group(1)))
        for match in re.finditer(rf"\b{key}\s*=\s*\"((?:[^\"\\]|\\.)*)\"", stripped, flags=re.DOTALL):
            flattened.append(match.group(1).replace('\\"', '"'))

    flattened = [value.strip() for value in flattened if value.strip()]
    if flattened:
        return dedupe_keep_order(flattened)

    array_match = re.match(r"^\{([^}]*)\}", stripped, flags=re.DOTALL)
    if array_match:
        values = [value.strip() for value in parse_java_string_literals(array_match.group(1)) if value.strip()]
        if values:
            return dedupe_keep_order(values)

    literal_match = re.match(r"^\"((?:[^\"\\]|\\.)*)\"", stripped, flags=re.DOTALL)
    if literal_match:
        value = literal_match.group(1).replace('\\"', '"').strip()
        return [value] if value else [""]

    return [""]


def extract_request_methods(args: str) -> list[str]:
    methods = [token.upper() for token in REQUEST_METHOD_RE.findall(args)]
    return dedupe_keep_order(methods) if methods else ["ALL"]


def normalize_snippet(value: str, max_length: int = 220) -> str:
    compact = re.sub(r"\s+", " ", value).strip()
    return compact if len(compact) <= max_length else compact[: max_length - 3] + "..."


def looks_like_method_declaration(line: str) -> bool:
    stripped = line.strip()
    if not stripped or stripped.startswith("@"):
        return False
    if stripped.startswith("//") or stripped.startswith("/*") or stripped.startswith("*"):
        return False
    if "(" not in stripped or stripped.endswith(";"):
        return False
    if not any(token in stripped for token in ("public ", "protected ", "private ")):
        return False
    if " class " in f" {stripped} ":
        return False
    return True


def join_paths(base_path: str, method_path: str) -> str:
    base = base_path.strip()
    child = method_path.strip()
    if not base and not child:
        return "/"
    if not base:
        return "/" + child.lstrip("/")
    if not child:
        return "/" + base.lstrip("/")
    return "/" + "/".join([base.strip("/"), child.strip("/")])


def parse_class_level_paths(annotations: list[tuple[str, int]]) -> list[str]:
    paths: list[str] = []
    for annotation, _ in annotations:
        if annotation_simple_name(annotation).lower() != "requestmapping":
            continue
        paths.extend(extract_mapping_paths(annotation_args(annotation)))
    return dedupe_keep_order(paths) if paths else [""]


def parse_method_level_mappings(
    annotations: list[tuple[str, int]],
    controller_class: str,
    file_rel: str,
) -> list[ControllerMapping]:
    mappings: list[ControllerMapping] = []
    for annotation, line_number in annotations:
        simple_name = annotation_simple_name(annotation).lower()
        if simple_name not in HTTP_MAPPING_NAME_TO_METHOD:
            continue

        args = annotation_args(annotation)
        paths = extract_mapping_paths(args)
        methods = list(HTTP_MAPPING_NAME_TO_METHOD[simple_name])
        if not methods:
            methods = extract_request_methods(args)

        snippet = normalize_snippet(annotation)
        for method in methods:
            for path_value in paths:
                full_path = normalize_endpoint_template(path_value if path_value else "/")
                mappings.append(
                    ControllerMapping(
                        file_path=file_rel,
                        method=method,
                        endpoint_normalized=full_path,
                        line=line_number,
                        snippet=snippet,
                        controller_class=controller_class,
                    )
                )
    return mappings


def build_backend_mappings(repo_root: Path, backend_root: Path) -> list[ControllerMapping]:
    mappings: list[ControllerMapping] = []
    java_files = sorted(backend_root.rglob("*.java"), key=lambda item: normalize(item.as_posix()).lower())

    for java_file in java_files:
        rel_path = to_rel(java_file, repo_root)
        lines = java_file.read_text(encoding="utf-8", errors="replace").splitlines()

        pending_annotations: list[tuple[str, int]] = []
        class_paths = [""]
        controller_class = java_file.stem

        index = 0
        while index < len(lines):
            line = lines[index]
            stripped = line.strip()

            if stripped.startswith("@"):
                annotation, end_index = extract_annotation(lines, index)
                pending_annotations.append((annotation, index + 1))
                index = end_index + 1
                continue

            class_match = CLASS_DECL_RE.search(line)
            if class_match:
                controller_class = class_match.group(1)
                class_paths = parse_class_level_paths(pending_annotations)
                pending_annotations = []
                index += 1
                continue

            if looks_like_method_declaration(line):
                method_mappings = parse_method_level_mappings(pending_annotations, controller_class, rel_path)
                for mapping in method_mappings:
                    for class_path in class_paths:
                        full = normalize_endpoint_template(join_paths(class_path, mapping.endpoint_normalized))
                        mappings.append(
                            ControllerMapping(
                                file_path=rel_path,
                                method=mapping.method,
                                endpoint_normalized=full,
                                line=mapping.line,
                                snippet=mapping.snippet,
                                controller_class=mapping.controller_class,
                            )
                        )
                pending_annotations = []
                index += 1
                continue

            if stripped and not stripped.startswith("//") and not stripped.startswith("/*") and not stripped.startswith("*"):
                pending_annotations = []

            index += 1

    return sorted(
        mappings,
        key=lambda item: (
            item.method,
            item.endpoint_normalized,
            item.file_path.lower(),
            item.line,
        ),
    )


def split_segments(path: str) -> list[str]:
    return [segment for segment in path.strip("/").split("/") if segment]


def paths_equivalent(lhs: str, rhs: str) -> bool:
    if lhs == rhs:
        return True
    lhs_parts = split_segments(lhs)
    rhs_parts = split_segments(rhs)
    if len(lhs_parts) != len(rhs_parts):
        return False
    for left, right in zip(lhs_parts, rhs_parts):
        if left == "{}" or right == "{}":
            continue
        if left != right:
            return False
    return True


def build_backend_lookup(mappings: list[ControllerMapping]) -> dict[str, list[ControllerMapping]]:
    lookup: dict[str, list[ControllerMapping]] = {}
    for mapping in mappings:
        lookup.setdefault(mapping.method, []).append(mapping)
    lookup.setdefault("ALL", [])
    return lookup


def compile_endpoint_regex(endpoint: str) -> re.Pattern[str]:
    path = endpoint.strip()
    if not path:
        path = "/"
    for token in ("?", "#"):
        if token in path:
            path = path.split(token, 1)[0]
    path = "/" + path.lstrip("/")

    segments = [segment for segment in path.split("/") if segment]
    if not segments:
        core = "/"
    else:
        parts: list[str] = []
        for segment in segments:
            if segment.startswith("{") and segment.endswith("}"):
                parts.append(r"[^/\"'`\s)]+")
            else:
                parts.append(re.escape(segment))
        core = "/" + "/".join(parts)

    pattern = rf"(?<![A-Za-z0-9_]){core}/?(?=[\"'`\s),]|$)"
    return re.compile(pattern)


def first_regex_snippet(text: str, pattern: re.Pattern[str], max_length: int = 220) -> tuple[int, str] | None:
    matched = pattern.search(text)
    if not matched:
        return None
    line = text.count("\n", 0, matched.start()) + 1
    start = max(0, matched.start() - 80)
    end = min(len(text), matched.end() + 80)
    snippet = normalize_snippet(text[start:end], max_length=max_length)
    return line, snippet


def dedupe_matches(matches: list[SourceMatch]) -> list[SourceMatch]:
    dedup: list[SourceMatch] = []
    seen: set[tuple[str, int, str]] = set()
    for item in sorted(matches, key=lambda value: (value.path.lower(), value.line, value.snippet)):
        key = (item.path.lower(), item.line, item.snippet)
        if key in seen:
            continue
        dedup.append(item)
        seen.add(key)
    return dedup


def read_text_files(paths: list[Path], root: Path) -> list[tuple[str, str]]:
    files: list[tuple[str, str]] = []
    for path in sorted(paths, key=lambda item: normalize(item.as_posix()).lower()):
        files.append((to_rel(path, root), path.read_text(encoding="utf-8", errors="replace")))
    return files


def collect_frontend_source_files(frontend_root: Path) -> list[Path]:
    files: list[Path] = []
    for path in frontend_root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in {".ts", ".tsx", ".js", ".jsx", ".mjs"}:
            continue
        lowered = path.name.lower()
        if ".test." in lowered or ".spec." in lowered:
            continue
        files.append(path)
    return files


def collect_test_source_files(backend_test_root: Path, frontend_root: Path) -> list[Path]:
    files: list[Path] = []
    if backend_test_root.exists():
        files.extend(list(backend_test_root.rglob("*.java")))

    for path in frontend_root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in {".ts", ".tsx", ".js", ".jsx", ".mjs"}:
            continue
        lowered = path.name.lower()
        path_parts = [part.lower() for part in path.parts]
        if ".test." in lowered or ".spec." in lowered or "__tests__" in path_parts:
            files.append(path)
    return files


def collect_regex_matches(
    files: list[tuple[str, str]],
    pattern: re.Pattern[str],
    max_matches: int,
) -> list[SourceMatch]:
    matches: list[SourceMatch] = []
    for rel_path, text in files:
        first = first_regex_snippet(text, pattern)
        if first is None:
            continue
        line, snippet = first
        matches.append(SourceMatch(path=rel_path, line=line, snippet=snippet))
        if len(matches) >= max_matches:
            break
    return dedupe_matches(matches)


def collect_controller_reference_matches(
    files: list[tuple[str, str]],
    controller_names: list[str],
    max_matches: int,
) -> list[SourceMatch]:
    if not controller_names:
        return []

    patterns = [re.compile(rf"\b{re.escape(name)}\b") for name in controller_names if name]
    matches: list[SourceMatch] = []
    for rel_path, text in files:
        for pattern in patterns:
            found = pattern.search(text)
            if not found:
                continue
            line = text.count("\n", 0, found.start()) + 1
            snippet = normalize_snippet(text[max(0, found.start() - 80) : min(len(text), found.end() + 80)])
            matches.append(SourceMatch(path=rel_path, line=line, snippet=snippet))
            break
        if len(matches) >= max_matches:
            break
    return dedupe_matches(matches)


def find_backend_matches(
    row: ApiSpecRow,
    backend_lookup: dict[str, list[ControllerMapping]],
    max_matches: int,
) -> list[SourceMatch]:
    candidates = list(backend_lookup.get(row.method, [])) + list(backend_lookup.get("ALL", []))
    matches: list[SourceMatch] = []
    for item in candidates:
        if not paths_equivalent(row.endpoint_normalized, item.endpoint_normalized):
            continue
        matches.append(SourceMatch(path=item.file_path, line=item.line, snippet=item.snippet))
        if len(matches) >= max_matches:
            break
    return dedupe_matches(matches)


def compute_rate(numerator: int, denominator: int) -> float:
    if denominator <= 0:
        return 0.0
    return round((numerator / denominator) * 100.0, 2)


def build_summary(rows: list[dict]) -> dict:
    total = len(rows)
    must_rows = [item for item in rows if item["importance"] == IMPORTANCE_MUST]

    backend_count = sum(1 for item in rows if item["backend_implemented"])
    frontend_count = sum(1 for item in rows if item["frontend_referenced"])
    tests_count = sum(1 for item in rows if item["tests_present"])
    full_count = sum(
        1 for item in rows if item["backend_implemented"] and item["frontend_referenced"] and item["tests_present"]
    )

    must_backend_count = sum(1 for item in must_rows if item["backend_implemented"])
    must_frontend_count = sum(1 for item in must_rows if item["frontend_referenced"])
    must_tests_count = sum(1 for item in must_rows if item["tests_present"])
    must_full_count = sum(
        1
        for item in must_rows
        if item["backend_implemented"] and item["frontend_referenced"] and item["tests_present"]
    )
    must_backend_missing_count = sum(1 for item in must_rows if not item["backend_implemented"])

    return {
        "total_api_rows": total,
        "must_api_rows": len(must_rows),
        "overall": {
            "backend_implemented_count": backend_count,
            "frontend_referenced_count": frontend_count,
            "tests_present_count": tests_count,
            "fully_covered_count": full_count,
            "backend_implemented_rate_pct": compute_rate(backend_count, total),
            "frontend_referenced_rate_pct": compute_rate(frontend_count, total),
            "tests_present_rate_pct": compute_rate(tests_count, total),
            "fully_covered_rate_pct": compute_rate(full_count, total),
        },
        "must_only": {
            "backend_implemented_count": must_backend_count,
            "frontend_referenced_count": must_frontend_count,
            "tests_present_count": must_tests_count,
            "fully_covered_count": must_full_count,
            "backend_implemented_rate_pct": compute_rate(must_backend_count, len(must_rows)),
            "frontend_referenced_rate_pct": compute_rate(must_frontend_count, len(must_rows)),
            "tests_present_rate_pct": compute_rate(must_tests_count, len(must_rows)),
            "fully_covered_rate_pct": compute_rate(must_full_count, len(must_rows)),
            "must_backend_missing_count": must_backend_missing_count,
        },
    }


def render_text(payload: dict) -> str:
    summary = payload["summary"]
    must_only = summary["must_only"]
    overall = summary["overall"]

    lines = [
        "spec_impl_coverage_report",
        f"status={payload['status']}",
        f"generated_at_utc={payload['generated_at_utc']}",
        f"git_head_short={payload['metadata']['git_head_short']}",
        f"git_branch={payload['metadata']['git_branch']}",
        f"total_api_rows={summary['total_api_rows']}",
        f"must_api_rows={summary['must_api_rows']}",
        (
            "overall="
            f"backend:{overall['backend_implemented_count']}/{summary['total_api_rows']} "
            f"({overall['backend_implemented_rate_pct']}%), "
            f"frontend:{overall['frontend_referenced_count']}/{summary['total_api_rows']} "
            f"({overall['frontend_referenced_rate_pct']}%), "
            f"tests:{overall['tests_present_count']}/{summary['total_api_rows']} "
            f"({overall['tests_present_rate_pct']}%)"
        ),
        (
            "must_only="
            f"backend:{must_only['backend_implemented_count']}/{summary['must_api_rows']} "
            f"({must_only['backend_implemented_rate_pct']}%), "
            f"frontend:{must_only['frontend_referenced_count']}/{summary['must_api_rows']} "
            f"({must_only['frontend_referenced_rate_pct']}%), "
            f"tests:{must_only['tests_present_count']}/{summary['must_api_rows']} "
            f"({must_only['tests_present_rate_pct']}%)"
        ),
        f"must_backend_missing_count={must_only['must_backend_missing_count']}",
        "",
        "Missing Must APIs:",
        "ProgramID | Method | Endpoint | Required ReqIDs | Missing signals",
        "--- | --- | --- | --- | ---",
    ]

    missing_must = payload["missing_must_apis"]
    if not missing_must:
        lines.append("(none)")
    else:
        for row in missing_must:
            req_ids = ",".join(row["req_ids"]) if row["req_ids"] else "-"
            missing = ",".join(row["missing_signals"]) if row["missing_signals"] else "-"
            lines.append(f"{row['program_id']} | {row['method']} | {row['endpoint']} | {req_ids} | {missing}")

    lines.extend(
        [
            "",
            "Must Green APIs (all signals true):",
            ",".join(payload["must_green_program_ids"]) if payload["must_green_program_ids"] else "(none)",
            "Must Red APIs (one or more missing signals):",
            ",".join(payload["must_red_program_ids"]) if payload["must_red_program_ids"] else "(none)",
            "",
        ]
    )
    return "\n".join(lines)


def render_markdown(payload: dict) -> str:
    summary = payload["summary"]
    must_only = summary["must_only"]
    overall = summary["overall"]

    lines = [
        "# Spec -> Implementation Coverage",
        "",
        f"- generated_at_utc: `{payload['generated_at_utc']}`",
        f"- git_head_short: `{payload['metadata']['git_head_short']}`",
        f"- git_branch: `{payload['metadata']['git_branch']}`",
        f"- total_api_rows: `{summary['total_api_rows']}`",
        f"- must_api_rows: `{summary['must_api_rows']}`",
        "",
        "## Summary",
        "| Scope | Backend | Frontend | Tests | Fully Covered |",
        "| --- | --- | --- | --- | --- |",
        (
            "| Overall | "
            f"{overall['backend_implemented_count']} ({overall['backend_implemented_rate_pct']}%) | "
            f"{overall['frontend_referenced_count']} ({overall['frontend_referenced_rate_pct']}%) | "
            f"{overall['tests_present_count']} ({overall['tests_present_rate_pct']}%) | "
            f"{overall['fully_covered_count']} ({overall['fully_covered_rate_pct']}%) |"
        ),
        (
            "| Must only | "
            f"{must_only['backend_implemented_count']} ({must_only['backend_implemented_rate_pct']}%) | "
            f"{must_only['frontend_referenced_count']} ({must_only['frontend_referenced_rate_pct']}%) | "
            f"{must_only['tests_present_count']} ({must_only['tests_present_rate_pct']}%) | "
            f"{must_only['fully_covered_count']} ({must_only['fully_covered_rate_pct']}%) |"
        ),
        "",
        "## Missing Must APIs",
        "| ProgramID | Method | Endpoint | Required ReqIDs | Missing signals |",
        "| --- | --- | --- | --- | --- |",
    ]

    if payload["missing_must_apis"]:
        for row in payload["missing_must_apis"]:
            req_ids = ",".join(row["req_ids"]) if row["req_ids"] else "-"
            missing = ",".join(row["missing_signals"]) if row["missing_signals"] else "-"
            lines.append(f"| {row['program_id']} | {row['method']} | {row['endpoint']} | {req_ids} | {missing} |")
    else:
        lines.append("| (none) | - | - | - | - |")

    lines.extend(
        [
            "",
            "## Must Red/Green",
            f"- must_green_program_ids: `{','.join(payload['must_green_program_ids']) if payload['must_green_program_ids'] else '(none)'}`",
            f"- must_red_program_ids: `{','.join(payload['must_red_program_ids']) if payload['must_red_program_ids'] else '(none)'}`",
            "",
        ]
    )
    return "\n".join(lines)


def write_output(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def build_payload(
    *,
    repo_root: Path,
    requirements_path: Path,
    api_workbook_path: Path,
    api_sheet: str,
    backend_root: Path,
    backend_test_root: Path,
    frontend_root: Path,
    max_matches_per_signal: int,
) -> dict:
    req_importance = load_requirements_importance(requirements_path)
    api_rows = load_api_rows(api_workbook_path, api_sheet, req_importance)

    backend_mappings = build_backend_mappings(repo_root, backend_root)
    backend_lookup = build_backend_lookup(backend_mappings)

    frontend_files = read_text_files(collect_frontend_source_files(frontend_root), repo_root)
    test_files = read_text_files(collect_test_source_files(backend_test_root, frontend_root), repo_root)

    coverage_rows: list[dict] = []
    missing_must: list[dict] = []
    must_green_program_ids: list[str] = []
    must_red_program_ids: list[str] = []

    for row in api_rows:
        endpoint_pattern = compile_endpoint_regex(row.endpoint)

        backend_matches = find_backend_matches(row, backend_lookup, max_matches=max_matches_per_signal)
        backend_implemented = bool(backend_matches)

        frontend_matches = collect_regex_matches(frontend_files, endpoint_pattern, max_matches_per_signal)
        frontend_referenced = bool(frontend_matches)

        test_matches = collect_regex_matches(test_files, endpoint_pattern, max_matches_per_signal)
        if not test_matches and backend_matches:
            controller_names = sorted(
                {
                    mapping.controller_class
                    for mapping in backend_lookup.get(row.method, []) + backend_lookup.get("ALL", [])
                    if paths_equivalent(row.endpoint_normalized, mapping.endpoint_normalized)
                }
            )
            test_matches = collect_controller_reference_matches(test_files, controller_names, max_matches_per_signal)
        tests_present = bool(test_matches)

        missing_signals: list[str] = []
        if not backend_implemented:
            missing_signals.append("backend_implemented")
        if not frontend_referenced:
            missing_signals.append("frontend_referenced")
        if not tests_present:
            missing_signals.append("tests_present")

        coverage_rows.append(
            {
                "row_index": row.row_index,
                "program_id": row.program_id,
                "method": row.method,
                "endpoint": row.endpoint,
                "endpoint_normalized": row.endpoint_normalized,
                "req_ids": row.req_ids,
                "importance": row.importance,
                "req_importance_breakdown": row.req_importance_breakdown,
                "backend_implemented": backend_implemented,
                "frontend_referenced": frontend_referenced,
                "tests_present": tests_present,
                "missing_signals": missing_signals,
                "backend_matches": [asdict(item) for item in backend_matches],
                "frontend_matches": [asdict(item) for item in frontend_matches],
                "test_matches": [asdict(item) for item in test_matches],
            }
        )

        if row.importance == IMPORTANCE_MUST:
            if missing_signals:
                must_red_program_ids.append(row.program_id)
                missing_must.append(
                    {
                        "row_index": row.row_index,
                        "program_id": row.program_id,
                        "method": row.method,
                        "endpoint": row.endpoint,
                        "req_ids": row.req_ids,
                        "missing_signals": missing_signals,
                    }
                )
            else:
                must_green_program_ids.append(row.program_id)

    coverage_rows = sorted(coverage_rows, key=lambda item: (item["row_index"], item["method"], item["endpoint"]))
    missing_must = sorted(missing_must, key=lambda item: (item["row_index"], item["program_id"]))
    must_green_program_ids = sorted(set(must_green_program_ids))
    must_red_program_ids = sorted(set(must_red_program_ids))

    summary = build_summary(coverage_rows)
    status = "PASS" if summary["must_only"]["must_backend_missing_count"] == 0 else "FAIL"
    generated_at_utc = run_git(repo_root, ["show", "-s", "--format=%cI", "HEAD"])
    if generated_at_utc == "UNKNOWN":
        generated_at_utc = "UNKNOWN"

    return {
        "status": status,
        "generated_at_utc": generated_at_utc,
        "metadata": {
            "git_head_short": run_git(repo_root, ["rev-parse", "--short", "HEAD"]),
            "git_branch": run_git(repo_root, ["rev-parse", "--abbrev-ref", "HEAD"]),
            "requirements": to_rel(requirements_path, repo_root),
            "api_workbook": to_rel(api_workbook_path, repo_root),
            "api_sheet": api_sheet,
            "backend_root": to_rel(backend_root, repo_root),
            "backend_test_root": to_rel(backend_test_root, repo_root),
            "frontend_root": to_rel(frontend_root, repo_root),
            "max_matches_per_signal": max_matches_per_signal,
        },
        "summary": summary,
        "must_green_program_ids": must_green_program_ids,
        "must_red_program_ids": must_red_program_ids,
        "missing_must_apis": missing_must,
        "rows": coverage_rows,
    }


def main() -> int:
    args = parse_args()
    repo_root = Path(args.root).resolve()

    if args.max_matches_per_signal < 1:
        sys.stderr.write("--max-matches-per-signal must be >= 1\n")
        return 2

    payload = build_payload(
        repo_root=repo_root,
        requirements_path=resolve_path(repo_root, args.requirements),
        api_workbook_path=resolve_path(repo_root, args.api_workbook),
        api_sheet=args.api_sheet,
        backend_root=resolve_path(repo_root, args.backend_root),
        backend_test_root=resolve_path(repo_root, args.backend_test_root),
        frontend_root=resolve_path(repo_root, args.frontend_root),
        max_matches_per_signal=args.max_matches_per_signal,
    )

    report_txt = resolve_path(repo_root, args.report_txt)
    report_json = resolve_path(repo_root, args.report_json)
    report_md = resolve_path(repo_root, args.report_md)

    text_report = render_text(payload)
    json_report = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    markdown_report = render_markdown(payload)

    write_output(report_txt, text_report)
    write_output(report_json, json_report)
    write_output(report_md, markdown_report)

    sys.stdout.write(text_report)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
