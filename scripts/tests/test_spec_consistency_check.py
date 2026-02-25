from __future__ import annotations

import csv
import subprocess
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
SCRIPT_PATH = REPO_ROOT / "scripts" / "spec_consistency_check.py"

KOR_REQ_ID = "\uc694\uad6c\uc0ac\ud56dID"      # 요구사항ID
KOR_API_LIST = "\uc804\uccb4API\ubaa9\ub85d"   # 전체API목록
KOR_NOTE = "\ube44\uace0"                      # 비고


def write_csv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file_pointer:
        writer = csv.writer(file_pointer)
        writer.writerow(headers)
        writer.writerows(rows)


def write_api_workbook(path: Path, note_rows: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = KOR_API_LIST
    headers = [
        "number",
        "category",
        "program_id",
        "api_name",
        "method",
        "endpoint",
        "auth",
        "request",
        "response",
        "ROLE",
        KOR_NOTE,
    ]
    worksheet.append(headers)

    roles = ["AGENT", "CUSTOMER", "ADMIN", "OPS", "SYSTEM"]
    for index, role in enumerate(roles, start=1):
        note = note_rows[index - 1] if index - 1 < len(note_rows) else ""
        worksheet.append(
            [
                str(index),
                "AUTH",
                f"API-{index:03d}",
                f"Test API {index}",
                "GET",
                f"/v1/test/{index}",
                "Bearer",
                "{}",
                "{}",
                role,
                note,
            ]
        )
    workbook.save(path)


def write_uiux_workbook(
    path: Path,
    token: str,
    mapping_key: str = "AI-001",
    mapping_screen_id: str = "AGT-003",
    mapping_sheet_name: str = "05_AGT003_대화스트리밍",
    pii_message: str = "민감정보(PII)가 감지되어 요청이 차단되었습니다.",
    pii_http: str = "409",
    sys_message: str = "요청한 세션을 찾을 수 없습니다.",
    sys_http: str = "404",
    assume1_b: str = "Resolved (근거 확정)",
    assume1_e: str = "Resolved (SEC-003 requirement aligned)",
    assume2_b: str = "Resolved (근거 확정)",
    assume2_e: str = "Resolved (SYS-001 requirement aligned)",
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    trace_sheet = workbook.active
    trace_sheet.title = "91_traceability"
    trace_sheet["A1"] = token

    error_sheet = workbook.create_sheet("01_에러메시지코드")
    error_sheet["A1"] = "에러/성공 메시지 코드"
    error_sheet["A3"] = "메시지코드"
    error_sheet["B3"] = "메시지 내용"
    error_sheet["C3"] = "HTTP"
    error_sheet["D3"] = "설명"
    error_sheet["A4"] = "SEC-003-409-PII"
    error_sheet["B4"] = pii_message
    error_sheet["C4"] = pii_http
    error_sheet["D4"] = "PII 정책 차단(fail-closed)"
    error_sheet["A5"] = "SYS-001-404"
    error_sheet["B5"] = sys_message
    error_sheet["C5"] = sys_http
    error_sheet["D5"] = "활성 세션 미존재/만료"

    inconsistency_sheet = workbook.create_sheet("90_불일치목록")
    inconsistency_sheet["A1"] = "불일치 목록"
    inconsistency_sheet["A3"] = "ID/항목"
    inconsistency_sheet["B3"] = "충돌 소스 A"
    inconsistency_sheet["C3"] = "충돌 소스 B"
    inconsistency_sheet["D3"] = "충돌 내용"
    inconsistency_sheet["E3"] = "결정(우선순위 기반)"
    inconsistency_sheet["F3"] = "영향 범위"
    inconsistency_sheet["G3"] = "리스크/후속 조치"
    inconsistency_sheet["A4"] = "ASSUME-001"
    inconsistency_sheet["B4"] = assume1_b
    inconsistency_sheet["C4"] = "cross validation (screen -> error catalog)"
    inconsistency_sheet["D4"] = "SEC-003-409-PII finalized"
    inconsistency_sheet["E4"] = assume1_e
    inconsistency_sheet["F4"] = "01_에러메시지코드"
    inconsistency_sheet["G4"] = "no follow-up"
    inconsistency_sheet["A5"] = "ASSUME-002"
    inconsistency_sheet["B5"] = assume2_b
    inconsistency_sheet["C5"] = "cross validation (screen -> error catalog)"
    inconsistency_sheet["D5"] = "SYS-001-404 finalized"
    inconsistency_sheet["E5"] = assume2_e
    inconsistency_sheet["F5"] = "01_에러메시지코드"
    inconsistency_sheet["G5"] = "no follow-up"

    unmapped = workbook.create_sheet("94_미매핑처분")
    unmapped["A1"] = "94_미매핑처분"
    headers = [
        "항목유형 (ReqID/API/DB)",
        "항목키",
        "중요도 (Must/Should/Could/Unknown)",
        "Phase (PHASE1/PHASE2/PHASE3)",
        "처분결정 (MapNow / MapToExisting / CreateScreen / BackendOnly / N-A)",
        "매핑대상 screen id",
        "매핑대상 시트명",
        "근거(스펙 참조: 파일명+시트/행 또는 ReqID 설명)",
        "비고(추가 작업/리스크)",
    ]
    for col_index, header in enumerate(headers, start=1):
        unmapped.cell(row=3, column=col_index, value=header)
    unmapped.append([])
    unmapped.append(
        [
            "ReqID",
            mapping_key,
            "Must",
            "PHASE1",
            "MapToExisting",
            mapping_screen_id,
            mapping_sheet_name,
            f"Requirements CSV ReqID={token}",
            "Mapped to existing screen and trace matrix",
        ]
    )

    workbook.save(path)


def write_db_workbook(path: Path, token: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TABLES"
    worksheet["A1"] = token
    workbook.save(path)


class SpecConsistencyCheckTest(unittest.TestCase):
    def run_script(
        self,
        root: Path,
        requirements: Path,
        summary: Path,
        development: Path,
        api_workbook: Path,
        db_workbook: Path,
        uiux_workbook: Path,
    ) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [
                "python",
                str(SCRIPT_PATH),
                "--root",
                str(root),
                "--requirements",
                str(requirements),
                "--summary",
                str(summary),
                "--development",
                str(development),
                "--api-workbook",
                str(api_workbook),
                "--db-workbook",
                str(db_workbook),
                "--uiux-workbook",
                str(uiux_workbook),
                "--report-json",
                str(root / "report.json"),
            ],
            cwd=REPO_ROOT,
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )

    def create_base_fixture(self, root: Path) -> dict[str, Path]:
        requirements = root / "docs/references/CS AI Chatbot_Requirements Statement.csv"
        summary = root / "docs/references/Summary of key features.csv"
        development = root / "docs/references/Development environment.csv"
        api_workbook = root / "docs/references/google_ready_api_spec_v0.3_20260216.xlsx"
        db_workbook = root / "docs/references/CS_AI_CHATBOT_DB.xlsx"
        uiux_workbook = root / "docs/uiux/CS_RAG_UI_UX_\uc124\uacc4\uc11c.xlsx"

        write_csv(
            requirements,
            ["ReqID", "description"],
            [
                [
                    "AI-001",
                    (
                        "secret_ref token tool citation done error heartbeat safe_response "
                        "error_code message trace_id details tenant_key "
                        "X-Trace-Id X-Tenant-Key Idempotency-Key Last-Event-ID"
                    ),
                ],
                ["SEC-001", "ROLE AGENT CUSTOMER ADMIN OPS SYSTEM"],
            ],
        )
        write_csv(
            summary,
            ["feature", KOR_REQ_ID],
            [
                ["A", "AI-001"],
                ["B", "SEC-001"],
            ],
        )
        write_csv(
            development,
            ["name", "description"],
            [["dev", "ReqID: AI-001, SEC-001"]],
        )
        write_api_workbook(
            api_workbook,
            [
                "ReqID: AI-001",
                "ReqID+: SEC-001",
                "secret_ref",
                "token tool citation done error heartbeat safe_response",
                "ReqID: AI-001",
            ],
        )
        write_uiux_workbook(uiux_workbook, "AI-001")
        write_db_workbook(db_workbook, "SEC-001")

        return {
            "requirements": requirements,
            "summary": summary,
            "development": development,
            "api_workbook": api_workbook,
            "db_workbook": db_workbook,
            "uiux_workbook": uiux_workbook,
        }

    def test_pass_case(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            fixture = self.create_base_fixture(root)
            proc = self.run_script(root=root, **fixture)
            self.assertEqual(proc.returncode, 0, msg=proc.stdout + proc.stderr)
            self.assertIn("PASS spec_consistency_check", proc.stdout)
            self.assertIn("invalid_tokens_count=0", proc.stdout)

    def test_fail_when_missing_reqid_reference(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            fixture = self.create_base_fixture(root)
            write_csv(
                fixture["summary"],
                ["feature", KOR_REQ_ID],
                [["A", "AI-001"], ["B", "OPS-999"]],
            )
            proc = self.run_script(root=root, **fixture)
            self.assertNotEqual(proc.returncode, 0)
            self.assertIn("REQID_SUMMARY_UNKNOWN", proc.stdout)
            self.assertIn("OPS-999", proc.stdout)

    def test_fail_when_typo_reqid_token_exists(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            fixture = self.create_base_fixture(root)
            write_api_workbook(
                fixture["api_workbook"],
                [
                    "ReqID: AI-001",
                    "ReqID: SEC-01X",
                    "secret_ref",
                    "token tool citation done error heartbeat safe_response",
                    "ReqID: AI-001",
                ],
            )
            proc = self.run_script(root=root, **fixture)
            self.assertNotEqual(proc.returncode, 0)
            self.assertIn("REQID_API_MALFORMED", proc.stdout)
            self.assertIn("SEC-01X", proc.stdout)

    def test_fail_when_terminology_variant_exists(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            fixture = self.create_base_fixture(root)
            write_csv(
                fixture["development"],
                ["name", "description"],
                [["dev", "ReqID: AI-001, SEC-001 safeResponse"]],
            )
            proc = self.run_script(root=root, **fixture)
            self.assertNotEqual(proc.returncode, 0)
            self.assertIn("TERMINOLOGY_SSE_EVENT_VARIANT", proc.stdout)
            self.assertIn("safeResponse", proc.stdout)

    def test_fail_when_uiux_placeholder_exists(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            fixture = self.create_base_fixture(root)
            write_uiux_workbook(
                fixture["uiux_workbook"],
                token="AI-001",
                mapping_key="-",
            )
            proc = self.run_script(root=root, **fixture)
            self.assertNotEqual(proc.returncode, 0)
            self.assertIn("PLACEHOLDER_UIUX_MAPPING_VALUE", proc.stdout)
            self.assertIn("token=-", proc.stdout)

    def test_fail_when_uiux_error_message_placeholder_exists(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            fixture = self.create_base_fixture(root)
            write_uiux_workbook(
                fixture["uiux_workbook"],
                token="AI-001",
                pii_message="TBD (근거 부족: 상세 메시지 원문 스펙 미정)",
            )
            proc = self.run_script(root=root, **fixture)
            self.assertNotEqual(proc.returncode, 0)
            self.assertIn("UIUX_ERROR_MESSAGE_PLACEHOLDER", proc.stdout)

    def test_fail_when_uiux_error_http_mismatch(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            fixture = self.create_base_fixture(root)
            write_uiux_workbook(
                fixture["uiux_workbook"],
                token="AI-001",
                sys_http="001",
            )
            proc = self.run_script(root=root, **fixture)
            self.assertNotEqual(proc.returncode, 0)
            self.assertIn("UIUX_ERROR_HTTP_MISMATCH", proc.stdout)

    def test_fail_when_uiux_assume_not_resolved(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            fixture = self.create_base_fixture(root)
            write_uiux_workbook(
                fixture["uiux_workbook"],
                token="AI-001",
                assume1_b="자동 보완(TBD)",
                assume1_e="근거 확보 전까지 TBD 유지",
            )
            proc = self.run_script(root=root, **fixture)
            self.assertNotEqual(proc.returncode, 0)
            self.assertIn("UIUX_ASSUME_UNRESOLVED", proc.stdout)


if __name__ == "__main__":
    unittest.main()
