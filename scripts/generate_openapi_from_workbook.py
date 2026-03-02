#!/usr/bin/env python3
"""Generate OpenAPI skeleton from API workbook '전체API목록' sheet."""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


KOR_API_LIST = "\uc804\uccb4API\ubaa9\ub85d"  # 전체API목록
KOR_NOTE = "\ube44\uace0"  # 비고
KOR_ROLE = "\uad8c\ud55c"  # 권한
KOR_CATEGORY = "\uce74\ud14c\uace0\ub9ac"  # 카테고리
KOR_PROGRAM_ID = "\ud504\ub85c\uadf8\ub7a8ID"  # 프로그램ID
KOR_API_NAME = "API\uba85"  # API명
KOR_DESCRIPTION = "\uc124\uba85"  # 설명

ROLE_TAXONOMY = {"AGENT", "CUSTOMER", "ADMIN", "OPS", "SYSTEM"}
ACCESS_LEVEL_TAXONOMY = {"PUBLIC", "AUTHENTICATED"}
METHOD_ORDER = ("get", "post", "put", "patch", "delete", "options", "head")
METHOD_SET = set(METHOD_ORDER)
STATUS_CODE_RE = re.compile(r"\b([1-5]\d{2})\b")
NOTE_KEY_VALUE_RE = re.compile(r"\b([a-z_][a-z0-9_]*)\s*=\s*([a-z0-9_./:-]+)", re.IGNORECASE)
PATH_PARAM_RE = re.compile(r"\{([a-zA-Z0-9_]+)\}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate OpenAPI skeleton from workbook")
    parser.add_argument("--workbook", default="docs/references/google_ready_api_spec_v0.3_20260216.xlsx")
    parser.add_argument("--sheet", default=KOR_API_LIST)
    parser.add_argument("--out", default="openapi/openapi.yaml")
    parser.add_argument("--title", default="AI_Chatbot API Skeleton")
    parser.add_argument("--server-url", default="http://localhost:8080")
    parser.add_argument("--version", default="0.1.0-skeleton")
    return parser.parse_args()


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def find_column(headers: list[str], candidates: list[str]) -> int | None:
    normalized_headers = [item.strip().lower() for item in headers]
    normalized_candidates = [item.strip().lower() for item in candidates]

    for candidate in normalized_candidates:
        if candidate in normalized_headers:
            return normalized_headers.index(candidate)

    for candidate in normalized_candidates:
        for index, header in enumerate(normalized_headers):
            if candidate in header:
                return index
    return None


def select_api_sheet(workbook_path: Path, preferred_sheet_name: str):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    preferred = preferred_sheet_name.strip().lower()

    for worksheet in workbook.worksheets:
        if worksheet.title.strip().lower() == preferred:
            return workbook, worksheet

    for worksheet in workbook.worksheets:
        lowered = worksheet.title.strip().lower()
        if "api" in lowered and ("list" in lowered or "\ubaa9\ub85d" in worksheet.title):
            return workbook, worksheet

    workbook.close()
    raise ValueError(f"sheet '{preferred_sheet_name}' not found in {workbook_path.as_posix()}")


def parse_note_key_values(note_value: str) -> dict[str, set[str]]:
    pairs: dict[str, set[str]] = {}
    for key_raw, value_raw in NOTE_KEY_VALUE_RE.findall(note_value):
        key = key_raw.strip().lower()
        value = value_raw.strip().upper()
        if not key or not value:
            continue
        values = pairs.setdefault(key, set())
        values.add(value)
    return pairs


def parse_roles(role_text: str) -> list[str]:
    if not role_text:
        return []
    tokens = re.split(r"[,/| ]+", role_text.upper())
    output: list[str] = []
    for token in tokens:
        candidate = token.strip()
        if not candidate:
            continue
        if candidate in ROLE_TAXONOMY and candidate not in output:
            output.append(candidate)
    return output


def parse_status_codes(response_text: str) -> list[str]:
    if not response_text:
        return []
    output: list[str] = []
    for code in STATUS_CODE_RE.findall(response_text):
        if code not in output:
            output.append(code)
    return output


def infer_default_status(method: str) -> str:
    if method == "post":
        return "201"
    if method == "delete":
        return "204"
    return "200"


def build_operation_id(program_id: str, method: str, path: str) -> str:
    raw = program_id if program_id else f"{method}_{path}"
    tokens = re.split(r"[^a-zA-Z0-9]+", raw)
    tokens = [token for token in tokens if token]
    if not tokens:
        return f"{method}Operation"
    first = tokens[0].lower()
    tail = [token[:1].upper() + token[1:].lower() for token in tokens[1:]]
    return first + "".join(tail)


def normalize_endpoint(path: str) -> str:
    value = path.strip()
    if not value:
        return ""
    if not value.startswith("/"):
        value = "/" + value
    return value


def parse_query_params(request_text: str) -> list[dict[str, object]]:
    if not request_text:
        return []
    result: list[dict[str, object]] = []
    lines = request_text.splitlines()
    for line in lines:
        if "query params" not in line.lower():
            continue
        if ":" not in line:
            continue
        right = line.split(":", 1)[1].strip()
        if not right or right in {"-", "none", "N/A"}:
            continue
        parts = [item.strip() for item in right.split(",") if item.strip()]
        for part in parts:
            name = part
            name = re.sub(r"\(.*?\)", "", name).strip()
            name = re.sub(r"[^a-zA-Z0-9_]", "", name)
            if not name:
                continue
            result.append(
                {
                    "name": name,
                    "in": "query",
                    "required": False,
                    "schema": {"type": "string"},
                }
            )
        break
    return result


def should_have_request_body(method: str, request_text: str) -> bool:
    if method not in {"post", "put", "patch"}:
        return False
    if not request_text:
        return False
    lowered = request_text.strip().lower()
    return lowered not in {"-", "none", "n/a"}


def build_openapi_document(
    workbook_path: Path,
    sheet_name: str,
    title: str,
    version: str,
    server_url: str,
) -> tuple[dict[str, object], dict[str, object]]:
    workbook, worksheet = select_api_sheet(workbook_path, sheet_name)
    try:
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
        headers = [normalize_text(value) for value in header_row]

        method_col = find_column(headers, ["Method", "HTTP Method"])
        endpoint_col = find_column(headers, ["Endpoint", "Path", "URI"])
        if method_col is None or endpoint_col is None:
            raise ValueError("Method/Endpoint columns are required")

        note_col = find_column(headers, [KOR_NOTE, "Remark", "Note"])
        role_col = find_column(headers, [KOR_ROLE, "Role"])
        category_col = find_column(headers, [KOR_CATEGORY, "Category"])
        program_id_col = find_column(headers, [KOR_PROGRAM_ID, "ProgramID"])
        name_col = find_column(headers, [KOR_API_NAME, "API Name", "Name"])
        description_col = find_column(headers, [KOR_DESCRIPTION, "Description"])
        request_col = find_column(headers, ["Request"])
        response_col = find_column(headers, ["Response"])

        scan_indices = [method_col, endpoint_col]
        for idx in (
            note_col,
            role_col,
            category_col,
            program_id_col,
            name_col,
            description_col,
            request_col,
            response_col,
        ):
            if idx is not None:
                scan_indices.append(idx)
        max_col = max(scan_indices) + 1

        path_map: dict[str, dict[str, object]] = {}
        duplicate_count = 0
        operation_count = 0
        empty_run = 0
        seen_data = False

        for row_number, row_values in enumerate(
            worksheet.iter_rows(min_row=2, max_col=max_col, values_only=True),
            start=2,
        ):
            values = [normalize_text(item) for item in row_values]
            method = values[method_col].lower() if method_col < len(values) else ""
            endpoint = normalize_endpoint(values[endpoint_col]) if endpoint_col < len(values) else ""

            probe_cells = [method, endpoint]
            if name_col is not None and name_col < len(values):
                probe_cells.append(values[name_col])
            if program_id_col is not None and program_id_col < len(values):
                probe_cells.append(values[program_id_col])
            if note_col is not None and note_col < len(values):
                probe_cells.append(values[note_col])

            if any(cell.strip() for cell in probe_cells):
                empty_run = 0
                seen_data = True
            else:
                empty_run += 1
                if seen_data and empty_run >= 120:
                    break
                continue

            if method not in METHOD_SET or not endpoint:
                continue

            operation_key = (endpoint, method)
            method_map = path_map.setdefault(endpoint, {})
            if method in method_map:
                duplicate_count += 1
                continue

            category = values[category_col] if category_col is not None and category_col < len(values) else ""
            program_id = values[program_id_col] if program_id_col is not None and program_id_col < len(values) else ""
            api_name = values[name_col] if name_col is not None and name_col < len(values) else ""
            description = values[description_col] if description_col is not None and description_col < len(values) else ""
            note_text = values[note_col] if note_col is not None and note_col < len(values) else ""
            role_text = values[role_col] if role_col is not None and role_col < len(values) else ""
            request_text = values[request_col] if request_col is not None and request_col < len(values) else ""
            response_text = values[response_col] if response_col is not None and response_col < len(values) else ""

            roles = parse_roles(role_text)
            note_pairs = parse_note_key_values(note_text)
            access_levels = sorted(note_pairs.get("access_level", set()) & ACCESS_LEVEL_TAXONOMY)
            access_level = access_levels[0] if access_levels else ("AUTHENTICATED" if roles else "")

            status_codes = parse_status_codes(response_text)
            primary_status = status_codes[0] if status_codes else infer_default_status(method)

            is_sse = "text/event-stream" in response_text.lower() or "/stream" in endpoint

            operation: dict[str, object] = OrderedDict()
            operation["operationId"] = build_operation_id(program_id=program_id, method=method, path=endpoint)
            if api_name:
                operation["summary"] = api_name
            if description:
                operation["description"] = description
            if category:
                operation["tags"] = [category]

            parameters: list[dict[str, object]] = []
            for name in PATH_PARAM_RE.findall(endpoint):
                parameters.append(
                    {
                        "name": name,
                        "in": "path",
                        "required": True,
                        "schema": {"type": "string"},
                    }
                )
            parameters.extend(parse_query_params(request_text))
            if parameters:
                operation["parameters"] = parameters

            if should_have_request_body(method=method, request_text=request_text):
                operation["requestBody"] = {
                    "required": True,
                    "content": {
                        "application/json": {
                            "schema": {
                                "type": "object",
                                "additionalProperties": True,
                            }
                        }
                    },
                }

            response_content_type = "text/event-stream" if is_sse else "application/json"
            response_schema = {"type": "string"} if is_sse else {"type": "object", "additionalProperties": True}
            responses: dict[str, object] = OrderedDict()
            responses[primary_status] = {
                "description": f"Workbook skeleton response ({primary_status})",
                "content": {
                    response_content_type: {
                        "schema": response_schema,
                    }
                },
            }
            operation["responses"] = responses

            if access_level != "PUBLIC":
                operation["security"] = [{"bearerAuth": []}]

            operation["x-access-level"] = access_level if access_level else "AUTHENTICATED"
            operation["x-rbac-roles"] = roles
            operation["x-program-id"] = program_id
            operation["x-source-row"] = row_number

            method_map[method] = operation
            operation_count += 1

        ordered_paths: OrderedDict[str, object] = OrderedDict()
        for path_key in sorted(path_map):
            raw_methods = path_map[path_key]
            method_bucket: OrderedDict[str, object] = OrderedDict()
            for method_name in METHOD_ORDER:
                if method_name in raw_methods:
                    method_bucket[method_name] = raw_methods[method_name]
            ordered_paths[path_key] = method_bucket

        document: dict[str, object] = OrderedDict(
            {
                "openapi": "3.0.3",
                "info": {
                    "title": title,
                    "version": version,
                    "description": "Auto-generated skeleton from workbook; schemas are intentionally coarse.",
                },
                "servers": [{"url": server_url}],
                "paths": ordered_paths,
                "components": {
                    "securitySchemes": {
                        "bearerAuth": {
                            "type": "http",
                            "scheme": "bearer",
                            "bearerFormat": "JWT",
                        }
                    }
                },
            }
        )
        stats = {
            "sheet": worksheet.title,
            "operation_count": operation_count,
            "path_count": len(ordered_paths),
            "duplicate_count": duplicate_count,
        }
        return document, stats
    finally:
        workbook.close()


def write_yaml_or_json(path: Path, payload: dict[str, object]) -> str:
    def to_plain(value):
        if isinstance(value, dict):
            return {key: to_plain(item) for key, item in value.items()}
        if isinstance(value, list):
            return [to_plain(item) for item in value]
        return value

    normalized_payload = to_plain(payload)
    try:
        import yaml  # type: ignore

        text = yaml.safe_dump(
            normalized_payload,
            sort_keys=False,
            allow_unicode=True,
            width=120,
            default_flow_style=False,
        )
        mode = "yaml"
    except ModuleNotFoundError:
        text = json.dumps(normalized_payload, ensure_ascii=False, indent=2) + "\n"
        mode = "json-as-yaml"

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")
    return mode


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook)
    out_path = Path(args.out)

    document, stats = build_openapi_document(
        workbook_path=workbook_path,
        sheet_name=args.sheet,
        title=args.title,
        version=args.version,
        server_url=args.server_url,
    )
    write_mode = write_yaml_or_json(out_path, document)

    sys.stdout.write(
        "\n".join(
            [
                "openapi_skeleton_generation",
                f"workbook={workbook_path.as_posix()}",
                f"sheet={stats['sheet']}",
                f"out={out_path.as_posix()}",
                f"write_mode={write_mode}",
                f"path_count={stats['path_count']}",
                f"operation_count={stats['operation_count']}",
                f"duplicate_count={stats['duplicate_count']}",
            ]
        )
        + "\n"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
