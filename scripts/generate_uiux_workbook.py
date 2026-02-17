#!/usr/bin/env python3
"""Generate CS_RAG_UI_UX_?ㅺ퀎??xlsx from template workbook."""

from __future__ import annotations

import csv
import re
import shutil
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


ROOT = Path(__file__).resolve().parents[1]
REF_DIR = ROOT / "docs" / "references"
OUT_DIR = ROOT / "docs" / "uiux"
OUT_PATH = OUT_DIR / "CS_RAG_UI_UX_?ㅺ퀎??xlsx"


def find_template() -> Path:
    files = sorted(ROOT.glob("*UI_UX*?ㅺ퀎??.xlsx"))
    if not files:
        raise FileNotFoundError("?쒗뵆由??묒???李얠쓣 ???놁뒿?덈떎.")
    return files[0]


def is_writable(cell) -> bool:
    return not isinstance(cell, MergedCell)


def setv(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col)
    if is_writable(cell):
        cell.value = value


def set_right(ws, row: int, value) -> bool:
    for col in (2, 3, 4, 5):
        cell = ws.cell(row=row, column=col)
        if is_writable(cell):
            cell.value = value
            return True
    return False


def clear(ws, r1: int, r2: int, c1: int = 1, c2: int = 8) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            if is_writable(cell):
                cell.value = None


def by_prefix(wb, prefix: str):
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return wb[name]
    raise KeyError(prefix)


def num_prefix(name: str) -> int:
    m = re.match(r"^(\d{2})_", name)
    return int(m.group(1)) if m else 999


def uniq_name(wb, base: str) -> str:
    base = base[:31]
    if base not in wb.sheetnames:
        return base
    i = 1
    while True:
        suf = f"_{i}"
        cand = base[: 31 - len(suf)] + suf
        if cand not in wb.sheetnames:
            return cand
        i += 1


def load_spec() -> dict:
    req = []
    with (REF_DIR / "CS AI Chatbot_Requirements Statement.csv").open(
        "r", encoding="utf-8-sig", newline=""
    ) as f:
        rd = csv.reader(f)
        next(rd, None)
        for row in rd:
            if row and row[0].strip():
                req.append(row[0].strip())

    feat = []
    with (REF_DIR / "Summary of key features.csv").open(
        "r", encoding="utf-8-sig", newline=""
    ) as f:
        rd = csv.reader(f)
        next(rd, None)
        for row in rd:
            if len(row) >= 6 and row[5].strip():
                feat.append(row[5].strip())

    api_wb = load_workbook(REF_DIR / "google_ready_api_spec_v0.3_20260216.xlsx", read_only=True, data_only=True)
    api_ws = api_wb[api_wb.sheetnames[1]]
    apis = []
    by_req = defaultdict(list)
    by_pid = {}
    for row in api_ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            continue
        pid, method, ep = row[2], row[4], row[5]
        if not (pid and method and ep):
            continue
        remark = str(row[10]).strip() if len(row) > 10 and row[10] else ""
        rids = []
        m = re.search(r"ReqID:\s*([^\n]+)", remark)
        if m:
            rids = [x.strip() for x in m.group(1).split(",") if x.strip()]
        item = {
            "pid": str(pid).strip(),
            "method": str(method).strip().upper(),
            "ep": str(ep).strip(),
            "role": str(row[9]).strip() if len(row) > 9 and row[9] else "",
            "rids": rids,
        }
        apis.append(item)
        by_pid[item["pid"]] = item
        for rid in rids:
            by_req[rid].append(item)

    db_wb = load_workbook(REF_DIR / "CS_AI_CHATBOT_DB.xlsx", read_only=True, data_only=True)
    tables = [s for s in db_wb.sheetnames if str(s).startswith("TB_")]

    req_set = set(req)
    feat_set = set(feat)
    all_req = sorted(req_set | feat_set)
    return {
        "req_set": req_set,
        "feat_set": feat_set,
        "all_req": all_req,
        "apis": apis,
        "api_by_req": by_req,
        "api_by_pid": by_pid,
        "tables": tables,
    }


SCREENS = [
    {
        "sheet": "03_AGT001_濡쒓렇?몄꽭?섎???,
        "sid": "AGT-001",
        "name": "?곷떞??濡쒓렇???몄뀡 遺?몄뒪?몃옪",
        "phase": "PHASE1",
        "role": "AGENT",
        "pids": ["API-AUTH-LOGIN", "API-AUTH-REFRESH", "API-SESSION-BOOTSTRAP"],
        "focus": "?뚮꼳???쇱슦??+ ?몄쬆 + ?몄뀡 蹂듭썝",
        "req": "SEC-001,SEC-002,SYS-001,SYS-002,SYS-004,UI-001,UI-008",
        "tables": "TB_AUTH_SESSION,TB_USER_ROLE,TB_TENANT_DOMAIN",
    },
    {
        "sheet": "04_AGT002_??,
        "sid": "AGT-002",
        "name": "?곷떞????理쒓렐 ?곷떞/?뚮┝)",
        "phase": "PHASE1",
        "role": "AGENT",
        "pids": ["API-MESSAGE-LIST", "OPS-ADMIN-DASHBOARD-SUMMARY"],
        "focus": "理쒓렐 ?몄뀡 + 怨듭? + ?ъ궗???묐떟",
        "req": "UI-001,UI-005,OPS-001,CCH-001,CCH-002",
        "tables": "TB_CONVERSATION,TB_MESSAGE,TB_OPS_EVENT",
    },
    {
        "sheet": "05_AGT003_??붿뒪?몃━諛?,
        "sid": "AGT-003",
        "name": "????붾㈃(硫?고꽩/泥⑤?/?ㅽ듃由щ컢)",
        "phase": "PHASE1",
        "role": "AGENT",
        "pids": [
            "API-MESSAGE-POST",
            "API-MESSAGE-RETRY",
            "API-MESSAGE-LIST",
            "API-STREAM-SSE",
            "API-STREAM-RESUME",
            "API-QUICKREPLY-POST",
        ],
        "focus": "SSE ?좏겙/洹쇨굅/?꾨즺 + ?ъ뿰寃?+ 以묐났 chunk 臾댁떆",
        "req": "UI-001,UI-002,UI-003,UI-004,UI-005,API-003,API-004,API-007,API-008,PERF-001",
        "tables": "TB_MESSAGE,TB_STREAM_EVENT,TB_MESSAGE_ATTACHMENT",
    },
    {
        "sheet": "06_AGT004_洹쇨굅?⑤꼸",
        "sid": "AGT-004",
        "name": "洹쇨굅 ?⑤꼸(citation/異쒖쿂/踰꾩쟾)",
        "phase": "PHASE1",
        "role": "AGENT",
        "pids": ["API-RAG-CITATIONS-GET", "API-RAG-RETRIEVE", "API-RAG-ANSWER"],
        "focus": "?듬? 臾몄옣?붽렐嫄?留ㅽ븨 + PII 留덉뒪??excerpt + 洹쇨굅遺議?李⑤떒",
        "req": "AI-004,AI-005,AI-009,RAG-001,RAG-002,RAG-003,KB-001,KB-002,KB-003",
        "tables": "TB_RAG_CITATION,TB_RAG_SEARCH_LOG,TB_KB_DOCUMENT_VERSION",
    },
    {
        "sheet": "07_AGT005_?쒗뵆由우텛泥?,
        "sid": "AGT-005",
        "name": "?쒗뵆由?異붿쿇(踰꾪듉 ?몃━嫄??꾩슜)",
        "phase": "PHASE1",
        "role": "AGENT",
        "pids": ["TMP-RECOMMEND-BUTTON", "TMP-TEMPLATE-LIST", "TOOL-CALL-VALIDATE"],
        "focus": "踰꾪듉 ?꾩슜 ?ㅽ뻾 + 荑⑤떎???몄뀡cap/?덉궛",
        "req": "TMP-001,TMP-002,TMP-003,TMP-004,TMP-005,TOOL-001,TOOL-002,TOOL-003",
        "tables": "TB_TEMPLATE_VERSION,TB_TEMPLATE_PLACEHOLDER,TB_TOOL_CALL_LOG",
    },
    {
        "sheet": "08_AGT006_?듬??몄쭛?꾩넚",
        "sid": "AGT-006",
        "name": "?듬? ?몄쭛湲?placeholder/?대젰/?꾩넚)",
        "phase": "PHASE1",
        "role": "AGENT",
        "pids": ["API-MESSAGE-POST", "API-QUICKREPLY-POST", "OPS-ERROR-CATALOG-UPSERT"],
        "focus": "?먮룞 梨꾩? + ?꾨씫 蹂??媛?대뱶 + ?뺤콉 寃利?,
        "req": "TMP-002,TMP-004,UI-005,AI-007,SEC-003",
        "tables": "TB_TEMPLATE_VERSION,TB_POLICY_VERSION,TB_GUARDRAIL_EVENT",
    },
    {
        "sheet": "09_AGT007_李⑤떒?몄씠?꾩쓳??,
        "sid": "AGT-007",
        "name": "?꾩넚 李⑤떒/Fail-Closed/safe_response",
        "phase": "PHASE1",
        "role": "AGENT",
        "pids": ["API-RAG-ANSWER", "OPS-ERROR-CATALOG-UPSERT", "OPS-AUDIT-LOG-QUERY"],
        "focus": "contract ?ㅽ뙣 ??safe_response only",
        "req": "AI-009,RAG-002,UI-006,SEC-003,SYS-003",
        "tables": "TB_GUARDRAIL_EVENT,TB_ERROR_CATALOG,TB_AUDIT_LOG",
    },
    {
        "sheet": "10_CUS001_?꾩젽遺?몄뒪?몃옪",
        "sid": "CUS-001",
        "name": "怨좉컼 ?꾩젽 濡쒕뵫/?몄뀡 蹂듭썝",
        "phase": "PHASE2",
        "role": "CUSTOMER",
        "pids": ["API-SESSION-CREATE", "API-SESSION-GET", "API-SESSION-BOOTSTRAP"],
        "focus": "MVP 踰좎씠??UI + ?몄뀡 蹂듭썝",
        "req": "SYS-001,SYS-002,UI-001,UI-007",
        "tables": "TB_WIDGET_INSTANCE,TB_AUTH_SESSION,TB_CONVERSATION",
    },
    {
        "sheet": "11_CUS002_泥⑤?硫붿떆吏",
        "sid": "CUS-002",
        "name": "怨좉컼 硫붿떆吏/泥⑤? ?낅줈??,
        "phase": "PHASE2",
        "role": "CUSTOMER",
        "pids": ["API-ATTACHMENT-PRESIGN", "API-ATTACHMENT-COMPLETE", "API-MESSAGE-POST"],
        "focus": "presign/complete + 遺遺꾩떎???ъ떆??,
        "req": "UI-003,API-003,API-005,SEC-003",
        "tables": "TB_ATTACHMENT,TB_MESSAGE_ATTACHMENT,TB_MESSAGE",
    },
    {
        "sheet": "12_CUS003_CSAT?몃뱶?ㅽ봽",
        "sid": "CUS-003",
        "name": "CSAT/?닿껐?щ?/?몃뱶?ㅽ봽",
        "phase": "PHASE2",
        "role": "CUSTOMER",
        "pids": ["API-CSAT-POST", "API-HANDOFF-REQUEST", "INT-HANDOFF-SYNC"],
        "focus": "CSAT + handoff CTA + ?붿빟 ?꾨떖",
        "req": "UI-007,API-005,INT-001,INT-002,INT-003",
        "tables": "TB_MESSAGE_FEEDBACK,TB_INTEGRATION_LOG,TB_WEBHOOK_ENDPOINT",
    },
    {
        "sheet": "13_ADM001_?뺤콉?꾨＼?꾪듃?뱀씤",
        "sid": "ADM-001",
        "name": "?뺤콉/?꾨＼?꾪듃 CRUD + ?뱀씤",
        "phase": "PHASE1",
        "role": "ADMIN",
        "pids": ["ADM-POLICY-UPDATE", "OPS-DEPLOY-APPROVAL-CREATE", "OPS-DEPLOY-APPROVAL-ACT"],
        "focus": "draft-review-approved + typed confirm",
        "req": "ADM-001,ADM-003,ADM-006,ADM-007,ADM-101,SEC-002",
        "tables": "TB_POLICY_VERSION,TB_PROMPT_VERSION,TB_AUDIT_LOG",
    },
    {
        "sheet": "14_ADM002_?쒗뵆由용쾭?꾨같??,
        "sid": "ADM-002",
        "name": "?쒗뵆由?踰꾩쟾/?뱀씤/諛고룷/濡ㅻ갚",
        "phase": "PHASE1",
        "role": "ADMIN",
        "pids": ["TMP-TEMPLATE-CREATE", "TMP-TEMPLATE-APPROVE", "TMP-TEMPLATE-DEPLOY", "TMP-TEMPLATE-ROLLBACK"],
        "focus": "?쒗뵆由??뱀씤踰꾩쟾 諛고룷 + 移대굹由?+ 濡ㅻ갚",
        "req": "ADM-002,ADM-007,TMP-001,TMP-003,TMP-004,TMP-005",
        "tables": "TB_TEMPLATE,TB_TEMPLATE_VERSION,TB_TEMPLATE_POLICY_MAP",
    },
    {
        "sheet": "15_ADM003_KB紐⑤뜽?퀾CP",
        "sid": "ADM-003",
        "name": "KB/紐⑤뜽/Tool/MCP ?댁쁺愿由?,
        "phase": "PHASE2",
        "role": "ADMIN",
        "pids": [
            "KB-DOC-UPLOAD",
            "KB-DOC-APPROVE",
            "KB-REINDEX-REQUEST",
            "ADM-MODEL-LIST",
            "ADM-MODEL-ACTIVATE",
            "MCP-SERVER-UPSERT",
            "TOOL-ALLOWLIST-UPDATE",
        ],
        "focus": "KB ?뱀씤/?ъ깋??+ secret_ref + allowlist + MCP ?ъ뒪",
        "req": "KB-001,KB-002,KB-003,ADM-004,ADM-005,ADM-102,TOOL-001,MCP-001,MCP-002,LLM-001,LLM-002,LLM-003",
        "tables": "TB_KB_DOCUMENT,TB_KB_INGEST_JOB,TB_LLM_MODEL,TB_PROVIDER_KEY,TB_TOOL_DEFINITION",
    },
    {
        "sheet": "16_OPS001_??쒕낫?쒕え?덊꽣留?,
        "sid": "OPS-001",
        "name": "?댁쁺 ??쒕낫???몃옒??吏??鍮꾩슜/荑쇳꽣)",
        "phase": "PHASE1",
        "role": "OPS",
        "pids": ["OPS-METRIC-SUMMARY", "OPS-ADMIN-DASHBOARD-SUMMARY", "OPS-TENANT-BILLING-REPORT"],
        "focus": "p50/p95/p99 + quota breach + ?댁긽?먯?",
        "req": "OPS-001,OPS-100,OPS-102,PERF-001,SEC-004",
        "tables": "TB_API_METRIC_HOURLY,TB_TENANT_USAGE_DAILY,TB_TENANT_QUOTA,TB_OPS_EVENT",
    },
    {
        "sheet": "17_OPS002_媛먯궗濡쒓렇利됱떆議곗튂",
        "sid": "OPS-002",
        "name": "媛먯궗濡쒓렇/trace_id ?쒕┫?ㅼ슫/利됱떆議곗튂",
        "phase": "PHASE1",
        "role": "OPS",
        "pids": ["OPS-TRACE-QUERY", "OPS-AUDIT-LOG-QUERY", "OPS-AUDIT-CHANGE-DIFF", "OPS-BLOCK-UPSERT", "OPS-PROVIDER-KILLSWITCH", "OPS-ROLLBACK-TRIGGER"],
        "focus": "trace_id ?먯씤遺꾩꽍 + block/kill-switch/rollback",
        "req": "OPS-003,OPS-103,SEC-002,SYS-004,ADM-100",
        "tables": "TB_AUDIT_LOG,TB_OPS_EVENT,TB_ERROR_CATALOG,TB_TENANT",
    },
]


def api_lines(screen: dict, api_by_pid: dict) -> tuple[str, str]:
    pairs = []
    for pid in screen["pids"]:
        api = api_by_pid.get(pid)
        if api:
            pairs.append(f"{api['method']} {api['ep']}")
    top = "; ".join(pairs[:4]) if pairs else "-"
    all_pairs = "; ".join(pairs) if pairs else "-"
    return top, all_pairs


def fill_screen(ws, screen: dict, api_by_pid: dict) -> None:
    clear(ws, 1, 150, 1, 5)
    top_api, all_api = api_lines(screen, api_by_pid)

    setv(ws, 1, 1, f"{screen['sid']}: {screen['name']}")
    setv(ws, 3, 1, "1. ?꾨줈洹몃옩 ?뺣낫")
    setv(ws, 4, 1, "?붾㈃ ID")
    if not set_right(ws, 4, screen["sid"]):
        setv(ws, 4, 1, f"?붾㈃ ID: {screen['sid']}")
    setv(ws, 5, 1, "?붾㈃紐?)
    if not set_right(ws, 5, screen["name"]):
        setv(ws, 5, 1, f"?붾㈃紐? {screen['name']}")
    setv(ws, 6, 1, "?꾨줈洹몃옩 ID")
    if not set_right(ws, 6, ", ".join(screen["pids"])):
        setv(ws, 6, 1, f"?꾨줈洹몃옩 ID: {', '.join(screen['pids'])}")
    setv(ws, 7, 1, "API")
    if not set_right(ws, 7, top_api):
        setv(ws, 7, 1, f"API: {top_api}")
    setv(ws, 8, 1, "沅뚰븳")
    if not set_right(ws, 8, screen["role"]):
        setv(ws, 8, 1, f"沅뚰븳: {screen['role']}")
    setv(ws, 9, 1, "媛쒕컻??Phase)")
    if not set_right(ws, 9, screen["phase"]):
        setv(ws, 9, 1, f"媛쒕컻??Phase): {screen['phase']}")
    setv(ws, 10, 1, "?ъ슜 ?뚯씠釉?)
    if not set_right(ws, 10, screen["tables"]):
        setv(ws, 10, 1, f"?ъ슜 ?뚯씠釉? {screen['tables']}")

    setv(ws, 12, 1, "1-1. ?붾㈃ 援ъ꽦?붿냼/?덉씠?꾩썐")
    setv(ws, 13, 1, f"???듭떖 ?ъ빱?? {screen['focus']}")
    setv(ws, 14, 1, "???덉씠?꾩썐: Header 64px + Sidebar 280px + Content(24px padding)")
    setv(ws, 15, 1, "???곹깭: loading/streaming/success/error/blocked/safe_response")
    setv(ws, 16, 1, "??UX ?? B2B ?댁쁺??臾멸뎄(異붿쿇 ?듬?/洹쇨굅/寃利??꾩넚/李⑤떒/?뱀씤/媛먯궗濡쒓렇)")
    setv(ws, 17, 1, "???묎렐?? ARIA/?ㅻ낫???ъ빱?? 紐⑤컮??360~1440 ???)
    setv(ws, 18, 1, "??蹂댁븞: UI ?④?? 蹂댁“, ?쒕쾭 403 理쒖쥌 媛뺤젣")
    setv(ws, 20, 1, "1-2. 議고쉶議곌굔")
    setv(ws, 21, 1, "??怨듯넻 ?꾪꽣: tenant_id / session_id / trace_id / 湲곌컙")
    setv(ws, 22, 1, "??PII ?ы븿 議고쉶??留덉뒪???댁떆 湲곕컲留??덉슜")

    setv(ws, 40, 1, "2. ?꾨줈洹몃옩 紐⑹쟻")
    setv(ws, 41, 1, f"??紐⑹쟻: {screen['name']}??瑜? ?곸슜 ?댁쁺 湲곗??쇰줈 ?쒓났")
    setv(ws, 42, 1, f"??踰붿쐞: {screen['focus']}")
    setv(ws, 43, 1, f"??愿??ReqID: {screen['req']}")

    setv(ws, 45, 1, "3. ?낅젰/議고쉶 ?꾨뱶 ?곸꽭")
    setv(ws, 46, 1, "?꾨뱶紐?)
    setv(ws, 46, 2, "而댄룷?뚰듃")
    setv(ws, 46, 3, "?꾩닔")
    setv(ws, 46, 4, "?좏슚??洹쒖튃")
    setv(ws, 46, 5, "?먮윭 肄붾뱶")
    fields = [
        ("tenant_key", "Hidden/Text", "O", "?뚮꼳???쇱슦???쇱튂", "SYS-002-403"),
        ("trace_id", "Header", "O", "UUID ?꾩닔", "SYS-004-409-TRACE"),
        ("session_id", "Hidden", "O", "?쒖꽦 ?몄뀡", "SYS-001-404"),
        ("payload", "Form", "O", "schema + policy + pii 寃利?, "API-003-422"),
    ]
    for i, f in enumerate(fields, start=47):
        for c, v in enumerate(f, start=1):
            setv(ws, i, c, v)

    setv(ws, 51, 1, "4. 踰꾪듉 ?숈옉 ?곸꽭")
    setv(ws, 52, 1, "踰꾪듉紐?)
    setv(ws, 52, 2, "?대깽??)
    setv(ws, 52, 3, "?숈옉")
    setv(ws, 52, 4, "?깃났 ??)
    setv(ws, 52, 5, "?ㅽ뙣 ??)
    buttons = [
        ("議고쉶/?ㅽ뻾", "onClick", "?대떦 API ?몄텧", "?곗씠??媛깆떊", "error_code 移댄깉濡쒓렇 臾멸뎄"),
        ("?ъ떆??, "onClick", "?ъ떆??媛???ㅻ쪟留??덉슜", "蹂듦뎄", "429/荑⑤떎???덈궡"),
        ("李⑤떒 ?댁젣 ?붿껌", "onClick", "愿由ъ옄 ?뱀씤 ?뚮줈??, "?뱀씤?湲?, "403 ?먮뒗 ?뺤콉 李⑤떒"),
    ]
    for i, b in enumerate(buttons, start=54):
        for c, v in enumerate(b, start=1):
            setv(ws, i, c, v)

    setv(ws, 57, 1, "5. DB 留ㅽ븨(?쎄린/?곌린)")
    setv(ws, 58, 1, "?뚯씠釉?)
    setv(ws, 58, 2, "?⑸룄")
    setv(ws, 58, 3, "而щ읆")
    setv(ws, 58, 4, "鍮꾧퀬")
    tbls = [x.strip() for x in screen["tables"].split(",") if x.strip()]
    for i, t in enumerate(tbls[:3], start=59):
        setv(ws, i, 1, t)
        setv(ws, i, 2, "議고쉶/?앹꽦/媛깆떊")
        setv(ws, i, 3, "tenant_id, session_id, message_id, trace_id")
        setv(ws, i, 4, "PII 留덉뒪??+ 媛먯궗濡쒓렇 ?곌퀎")

    setv(ws, 62, 1, "6. API/?ㅻ뜑/SSE 怨꾩빟")
    setv(ws, 63, 1, "?곌껐 API")
    if not set_right(ws, 63, all_api):
        setv(ws, 63, 1, f"?곌껐 API: {all_api}")
    setv(ws, 64, 1, "怨듯넻 ?ㅻ뜑: X-Trace-Id / X-Tenant-Key / Authorization / Idempotency-Key(POST)")
    setv(ws, 65, 1, "SSE: token/tool/citation/done/error/heartbeat/safe_response (?대떦 ?붾㈃留?")
    setv(ws, 66, 1, "SSE UX: first-token 1~2s, done ?댁쟾 醫낅즺 ??3???ъ뿰寃? 以묐났 chunk 臾댁떆")
    setv(ws, 67, 1, "?ㅻ쪟 UX: 401/403/409/422/429/5xx + Retry-After + ?고쉶 諛⑹?")

    setv(ws, 69, 1, "7. ?ъ슜???≪뀡 & ?쒖뒪??諛섏쓳")
    setv(ws, 70, 1, "?쒖꽌")
    setv(ws, 70, 2, "泥섎━ ?댁슜")
    setv(ws, 70, 3, "鍮꾧퀬")
    steps = [
        "?낅젰媛??ъ쟾 寃利?PII/?뺤콉/?ㅽ궎留?",
        "API ?몄텧 ??trace_id/tenant_id ?꾪뙆",
        "?깃났 ???곹깭 媛깆떊 諛?媛먯궗 ?대깽??湲곕줉",
        "?ㅽ뙣 ??error_code 湲곕컲 ?ъ슜??臾멸뎄 ?몄텧",
        "Fail-Closed 議곌굔?대㈃ ?꾩넚 李⑤떒 + safe_response留??몄텧",
    ]
    for i, s in enumerate(steps, start=71):
        setv(ws, i, 1, i - 70)
        setv(ws, i, 2, s)
        setv(ws, i, 3, "?댁쁺濡쒓렇 ?곕룞")

    setv(ws, 82, 1, "8. ?덉쇅?ы빆 泥댄겕")
    setv(ws, 83, 1, "No")
    setv(ws, 83, 2, "?덉쇅 ?곹솴")
    setv(ws, 83, 3, "?먮윭 肄붾뱶")
    setv(ws, 83, 4, "泥섎━")
    setv(ws, 83, 5, "?붾㈃ ?쒖떆")
    ex = [
        ("citation ?꾨씫", "AI-009-409-CITATION", "?꾩넚 李⑤떒", "safe_response留??몄텧"),
        ("?ㅽ궎留??ㅽ뙣", "AI-009-422-SCHEMA", "?꾩넚 李⑤떒", "?섏젙/異붽?吏덈Ц ?좊룄"),
        ("洹쇨굅 ?꾧퀎移?誘몃떖", "AI-009-409-EVIDENCE", "?꾩넚 李⑤떒", "洹쇨굅 ?ъ“???덈궡"),
        ("?뺤콉 ?꾨컲", "RAG-002-422-POLICY", "李⑤떒 ?먮뒗 媛뺤젣臾멸뎄 ?쎌엯", "?뺤콉 ?꾨컲 諛곕꼫"),
        ("PII 寃異?, "SEC-003-409-PII", "留덉뒪???낅줈??嫄곕?", "誘쇨컧?뺣낫 ?쒓굅 ?덈궡"),
        ("沅뚰븳 ?놁쓬", "SEC-002-403", "?붿껌 以묐떒", "沅뚰븳 ?붿껌 CTA"),
        ("?덉궛 珥덇낵", "API-008-429-BUDGET", "荑⑤떎??, "Retry-After 移댁슫?몃떎??),
    ]
    for i, item in enumerate(ex, start=84):
        setv(ws, i, 1, i - 83)
        for c, v in enumerate(item, start=2):
            setv(ws, i, c, v)

    setv(ws, 93, 1, "9. ?⑥쐞?뚯뒪???쒕굹由ъ삤")
    setv(ws, 94, 1, "TC No")
    setv(ws, 94, 2, "?뚯뒪????ぉ")
    setv(ws, 94, 3, "?뚯뒪???곗씠??)
    setv(ws, 94, 4, "?덉긽 寃곌낵")
    setv(ws, 94, 5, "Pass/Fail")
    sid = screen["sid"].replace("-", "")
    tests = [
        (f"TC-{sid}-001", "?뺤긽 ?뚮줈??, "?좏슚 ?낅젰", "?뺤긽 泥섎━"),
        (f"TC-{sid}-002", "PII 留덉뒪??, "email/phone ?ы븿", "?붾㈃/蹂듭궗/濡쒓렇 留덉뒪??),
        (f"TC-{sid}-003", "trace_id ?쒖떆/蹂듭궗", "沅뚰븳蹂?怨꾩젙", "OPS/ADMIN留?蹂듭궗 ?덉슜"),
        (f"TC-{sid}-004", "RBAC 403", "沅뚰븳 ?녿뒗 ?몄텧", "?묎렐 李⑤떒 + ?덈궡"),
        (f"TC-{sid}-005", "429 + Retry-After", "?붿껌 ??＜", "荑⑤떎??+ ?고쉶 諛⑹?"),
        (f"TC-{sid}-006", "SSE ?ъ뿰寃?timeout", "1???⑥젅", "resume ?먮뒗 fallback"),
        (f"TC-{sid}-007", "Fail-Closed", "citation/schema ?ㅽ뙣", "safe_response only"),
    ]
    for i, t in enumerate(tests, start=95):
        for c, v in enumerate(t, start=1):
            setv(ws, i, c, v)

    setv(ws, 108, 1, "10. ?곕━ ?꾨줈?앺듃 ?꾩닔 ?쒖빟 ?뱀뀡")
    setv(ws, 109, 1, "??ぉ")
    setv(ws, 109, 2, "UI?먯꽌???쒗쁽(蹂댁씠??寃??④린??寃?")
    setv(ws, 109, 3, "愿??ReqID")
    setv(ws, 109, 4, "愿??API")
    setv(ws, 109, 5, "愿??DB/Telemetry")
    constraints = [
        ("Fail-Closed", "寃利??ㅽ뙣 ???꾩넚 李⑤떒, safe_response留??몄텧", "AI-009,RAG-002", ",".join(screen["pids"][:2]), "TB_GUARDRAIL_EVENT/fail_closed_count"),
        ("PII", "蹂듭궗/?ㅼ슫濡쒕뱶/濡쒓렇 ?ы븿 ?꾧뎄媛?留덉뒪??, "SEC-003", ",".join(screen["pids"][:2]), "TB_MESSAGE,TB_ATTACHMENT/pii_mask_count"),
        ("trace_id", "?붿껌쨌?ㅽ듃由셋룸줈洹?trace_id ?곴?愿怨?, "SYS-004", ",".join(screen["pids"][:2]), "TB_AUDIT_LOG/trace_coverage_pct"),
        ("RBAC", "UI ?④? + ?쒕쾭 403 理쒖쥌 媛뺤젣", "SEC-002", ",".join(screen["pids"][:2]), "TB_ROLE_PERMISSION/rbac_deny_count"),
        ("Budget", "429/Retry-After/荑⑤떎???고쉶 諛⑹?", "API-007,API-008", ",".join(screen["pids"][:2]), "TB_TENANT_QUOTA/quota_breach_count"),
        ("?뱀씤踰꾩쟾", "approved 踰꾩쟾留??댁쁺 寃쎈줈 ?ъ슜", "ADM-101,TMP-003", ",".join(screen["pids"][:2]), "TB_*_VERSION/version_mismatch_count"),
    ]
    for i, row in enumerate(constraints, start=110):
        for c, v in enumerate(row, start=1):
            setv(ws, i, c, v)


def fill_errors(ws) -> None:
    clear(ws, 1, 600, 1, 6)
    setv(ws, 1, 1, "?먮윭/?깃났 硫붿떆吏 肄붾뱶 (CS RAG ?꾨줈?앺듃)")
    setv(ws, 3, 1, "硫붿떆吏肄붾뱶")
    setv(ws, 3, 2, "硫붿떆吏 ?댁슜")
    setv(ws, 3, 3, "HTTP")
    setv(ws, 3, 4, "?ㅻ챸")
    rows = [
        ("SEC-001-401", "?몄쬆??留뚮즺?섏뿀?듬땲?? ?ㅼ떆 濡쒓렇?명빐 二쇱꽭??", "401", "?몄쬆/?몄뀡"),
        ("SEC-002-403", "?대떦 湲곕뒫???묎렐 沅뚰븳???놁뒿?덈떎.", "403", "RBAC 李⑤떒"),
        ("SYS-002-403", "?덉슜?섏? ?딆? ?뚮꼳???먮뒗 ?꾨찓?몄엯?덈떎.", "403", "tenant ?쇱슦??),
        ("API-003-409", "?숈씪 ?붿껌???대? 泥섎━ 以묒엯?덈떎.", "409", "idempotency 以묐났"),
        ("API-003-422", "?낅젰媛??뺤떇???щ컮瑜댁? ?딆뒿?덈떎.", "422", "?붿껌 ?ㅽ궎留?寃利?),
        ("API-008-429-BUDGET", "?붿껌 ?쒕룄瑜?珥덇낵?덉뒿?덈떎. ?좎떆 ???ъ떆?꾪빐 二쇱꽭??", "429", "budget/rate-limit"),
        ("API-008-429-SSE", "?숈떆 ?ㅽ듃由щ컢 ?곌껐 ?섎? 珥덇낵?덉뒿?덈떎.", "429", "SSE ?숈떆 ?곌껐 ?쒗븳"),
        ("AI-009-422-SCHEMA", "?묐떟 ?뺤떇 寃利앹뿉 ?ㅽ뙣?섏뿬 ?꾩넚?????놁뒿?덈떎.", "422", "Answer Contract"),
        ("AI-009-409-CITATION", "洹쇨굅 ?몄슜???꾨씫?섏뼱 ?꾩넚?????놁뒿?덈떎.", "409", "Fail-Closed"),
        ("AI-009-409-EVIDENCE", "洹쇨굅 ?먯닔媛 湲곗???誘몃떖?섏뿬 ?꾩넚?????놁뒿?덈떎.", "409", "Fail-Closed"),
        ("RAG-002-422-POLICY", "?뺤콉 ?꾨컲?쇰줈 ?꾩넚??李⑤떒?섏뿀?듬땲??", "422", "?뺤콉 李⑤떒"),
        ("AI-009-200-SAFE", "異붽? ?뺤씤???꾩슂?섏뿬 ?덉쟾 ?묐떟?쇰줈 ?꾪솚?덉뒿?덈떎.", "200", "safe_response"),
        ("FILE-001-422", "?뚯씪 ?뺤떇 ?먮뒗 ?⑸웾???덉슜 踰붿쐞瑜?珥덇낵?덉뒿?덈떎.", "422", "presign 寃利?),
        ("FILE-002-409", "泥⑤? ?낅줈???꾨즺 ?뺤씤???꾩슂?⑸땲??", "409", "complete ?꾨씫"),
        ("SSE-002-504", "?ㅽ듃由щ컢 ?곌껐??吏?곕릺??蹂듦뎄瑜??쒕룄?⑸땲??", "504", "SSE timeout"),
        ("SSE-003-409-RESUME", "?댁뼱諛쏄린???ㅽ뙣?섏뿬 ?붿빟 ?ъ슂泥?쑝濡??꾪솚?덉뒿?덈떎.", "409", "SSE resume ?ㅽ뙣"),
        ("TMP-001-429-COOLDOWN", "?쒗뵆由?異붿쿇? ?좎떆 ???ㅼ떆 ?ㅽ뻾?????덉뒿?덈떎.", "429", "異붿쿇 荑⑤떎??),
        ("TOOL-003-403-BLOCK", "?쒗뵆由?異붿쿇? 踰꾪듉?쇰줈留??ㅽ뻾?????덉뒿?덈떎.", "403", "踰꾪듉 ?꾩슜 ?ㅽ뻾"),
        ("ADM-101-409", "?뱀씤?섏? ?딆? 踰꾩쟾? ?댁쁺??諛고룷?????놁뒿?덈떎.", "409", "?뱀씤踰꾩쟾 媛뺤젣"),
        ("ADM-102-422", "Provider ?ㅻ뒗 secret_ref濡쒕쭔 ?깅줉?????덉뒿?덈떎.", "422", "?됰Ц ??湲덉?"),
        ("SYS-004-409-TRACE", "?붿껌 異붿쟻 ?뺣낫媛 ?꾨씫?섏뼱 泥섎━?????놁뒿?덈떎.", "409", "trace_id ?꾨씫"),
        ("OPS-003-409", "利됱떆議곗튂 ?뺤씤 臾멸뎄媛 ?쇱튂?섏? ?딆뒿?덈떎.", "409", "typed confirm"),
        ("SYS-003-500", "?쇱떆?곸씤 ?쒖뒪???ㅻ쪟媛 諛쒖깮?덉뒿?덈떎.", "500", "怨듯넻 ?쒕쾭 ?ㅻ쪟"),
        ("SYS-003-503", "?몃? ?곕룞 ?쒕퉬?ㅺ? ?쇱떆?곸쑝濡?遺덉븞?뺥빀?덈떎.", "503", "?몃? ?곕룞 ?μ븷"),
    ]
    for i, row in enumerate(rows, start=4):
        for c, v in enumerate(row, start=1):
            setv(ws, i, c, v)


def fill_codes(ws, ws2) -> None:
    clear(ws, 1, 700, 1, 6)
    setv(ws, 1, 1, "異붽? 醫낇빀肄붾뱶 (CS RAG ?꾨줈?앺듃)")
    setv(ws, 3, 1, "肄붾뱶洹몃９")
    setv(ws, 3, 2, "肄붾뱶媛?)
    setv(ws, 3, 3, "肄붾뱶紐?)
    setv(ws, 3, 4, "?ㅻ챸")
    rows = [
        ("ROLE", "AGENT", "?곷떞??, "Agent Console ?ъ슜??),
        ("ROLE", "CUSTOMER", "怨좉컼", "Customer Widget ?ъ슜??),
        ("ROLE", "ADMIN", "?쒖뒪??愿由ъ옄", "?뺤콉/諛고룷 ?뱀씤 沅뚰븳"),
        ("ROLE", "OPS", "?댁쁺 愿由ъ옄", "?댁쁺 ??쒕낫??利됱떆議곗튂"),
        ("ROLE", "SYSTEM", "?쒖뒪??, "?대? ?쒕퉬??怨꾩젙"),
        ("APPROVAL_STATUS", "draft", "珥덉븞", "?몄쭛 媛??),
        ("APPROVAL_STATUS", "review", "寃?좎쨷", "?뱀씤 ?湲?),
        ("APPROVAL_STATUS", "approved", "?뱀씤", "?댁쁺 諛섏쁺 媛??),
        ("APPROVAL_STATUS", "retired", "?먭린", "?댁쁺 ?ъ슜 遺덇?"),
        ("SSE_EVENT_TYPE", "token", "?좏겙", "?ㅽ듃由щ컢 ?띿뒪??泥?겕"),
        ("SSE_EVENT_TYPE", "tool", "?꾧뎄", "tool ?몄텧 ?곹깭"),
        ("SSE_EVENT_TYPE", "citation", "洹쇨굅", "?몄슜 洹쇨굅 硫뷀?"),
        ("SSE_EVENT_TYPE", "done", "?꾨즺", "?ㅽ듃由щ컢 醫낅즺"),
        ("SSE_EVENT_TYPE", "error", "?ㅻ쪟", "?쒖? ?먮윭 ?대깽??),
        ("SSE_EVENT_TYPE", "heartbeat", "?섑듃鍮꾪듃", "?곌껐 ?좎?"),
        ("SSE_EVENT_TYPE", "safe_response", "?덉쟾?묐떟", "fail-closed ?꾪솚"),
        ("ANSWER_STATUS", "streaming", "?앹꽦以?, "?묐떟 ?앹꽦 吏꾪뻾"),
        ("ANSWER_STATUS", "completed", "?꾨즺", "寃利??듦낵"),
        ("ANSWER_STATUS", "blocked", "李⑤떒", "?꾩넚 李⑤떒"),
        ("ANSWER_STATUS", "safe_only", "?덉쟾?묐떟", "safe_response留??덉슜"),
        ("TENANT_PLAN", "starter", "Starter", "湲곕낯 ?뚮옖"),
        ("TENANT_PLAN", "pro", "Pro", "以묎컙 ?뚮옖"),
        ("TENANT_PLAN", "enterprise", "Enterprise", "?곸쐞 ?뚮옖"),
        ("QUOTA_STATUS", "normal", "?뺤긽", "?쒕룄 ?ъ쑀"),
        ("QUOTA_STATUS", "warning", "二쇱쓽", "?꾧퀎移?洹쇱젒"),
        ("QUOTA_STATUS", "exceeded", "珥덇낵", "?붿껌 ?쒗븳"),
        ("QUOTA_STATUS", "blocked", "李⑤떒", "?붿껌 李⑤떒"),
        ("KB_INGEST_STATUS", "uploaded", "?낅줈??, "臾몄꽌 ?낅줈???꾨즺"),
        ("KB_INGEST_STATUS", "indexing", "?몃뜳?깆쨷", "寃???몃뜳??鍮뚮뱶"),
        ("KB_INGEST_STATUS", "approved", "?뱀씤", "?댁쁺 寃??諛섏쁺"),
        ("KB_INGEST_STATUS", "failed", "?ㅽ뙣", "?ъ떆???꾩슂"),
        ("TOOL_CALL_STATUS", "success", "?깃났", "?뺤긽 ?꾨즺"),
        ("TOOL_CALL_STATUS", "failed", "?ㅽ뙣", "?ㅻ쪟 醫낅즺"),
        ("TOOL_CALL_STATUS", "blocked_by_allowlist", "李⑤떒", "?덉슜紐⑸줉 誘명룷??),
        ("GUARDRAIL_EVENT_TYPE", "pii_masked", "PII 留덉뒪??, "誘쇨컧?뺣낫 留덉뒪??),
        ("GUARDRAIL_EVENT_TYPE", "policy_block", "?뺤콉 李⑤떒", "?뺤콉 ?꾨컲"),
        ("GUARDRAIL_EVENT_TYPE", "schema_fail", "?ㅽ궎留??ㅽ뙣", "Answer Contract ?ㅽ뙣"),
        ("GUARDRAIL_EVENT_TYPE", "citation_missing", "?몄슜 ?꾨씫", "Fail-Closed"),
        ("GUARDRAIL_EVENT_TYPE", "evidence_low", "洹쇨굅 遺議?, "Fail-Closed"),
        ("MCP_TRANSPORT", "stdio", "STDIO", "媛쒕컻 ?섍꼍"),
        ("MCP_TRANSPORT", "sse", "SSE", "?댁쁺 ?ㅼ떆媛?),
        ("MCP_TRANSPORT", "http", "HTTP", "?댁쁺 API"),
    ]
    for i, row in enumerate(rows, start=4):
        for c, v in enumerate(row, start=1):
            setv(ws, i, c, v)

    clear(ws2, 1, 300, 1, 6)
    setv(ws2, 1, 1, "異붽? 醫낇빀肄붾뱶 (CS RAG ?꾨줈?앺듃 異붽?遺?")
    setv(ws2, 4, 1, "肄붾뱶洹몃９")
    setv(ws2, 4, 2, "肄붾뱶媛?)
    setv(ws2, 4, 3, "肄붾뱶紐?)
    setv(ws2, 4, 4, "?ㅻ챸")
    ext = [
        ("CSAT_SCORE", "1", "留ㅼ슦 遺덈쭔議?, "1??),
        ("CSAT_SCORE", "2", "遺덈쭔議?, "2??),
        ("CSAT_SCORE", "3", "蹂댄넻", "3??),
        ("CSAT_SCORE", "4", "留뚯”", "4??),
        ("CSAT_SCORE", "5", "留ㅼ슦 留뚯”", "5??),
        ("BUDGET_SIGNAL", "near_limit", "?꾧퀎移?洹쇱젒", "寃쎄퀬 ?뚮┝"),
        ("BUDGET_SIGNAL", "blocked", "李⑤떒", "?붿껌 李⑤떒"),
        ("SAFE_RESPONSE_TYPE", "needs_clarification", "異붽??뺤씤 ?꾩슂", "異붽? 吏덈Ц ?좊룄"),
        ("SAFE_RESPONSE_TYPE", "insufficient_evidence", "洹쇨굅 遺議?, "洹쇨굅 ?ъ“???좊룄"),
    ]
    for i, row in enumerate(ext, start=5):
        for c, v in enumerate(row, start=1):
            setv(ws2, i, c, v)


def fill_sp(ws) -> None:
    clear(ws, 1, 600, 1, 8)
    setv(ws, 1, 1, "DB Access / Repository / Query ?꾩껜 紐⑸줉 (A?? SP ?泥?")
    setv(ws, 2, 1, "?댁쁺 諛⑹떇: Stored Procedure ???Repository/Query ?묎렐 紐⑸줉??愿由ы븳??")
    setv(ws, 3, 1, "No")
    setv(ws, 3, 2, "?대쫫")
    setv(ws, 3, 3, "?ㅻ챸")
    setv(ws, 3, 4, "二쇱슂 ?뚮씪誘명꽣")
    setv(ws, 3, 5, "Phase")
    rows = [
        ("AuthSessionRepository.findByTenantAndUser", "?곷떞??濡쒓렇???몄뀡 議고쉶", "tenant_key,user_id", "PHASE1"),
        ("AuthSessionRepository.upsertSession", "?몄뀡 ?앹꽦/媛깆떊", "session_id,expires_at,trace_id", "PHASE1"),
        ("ConversationRepository.listRecentByAgent", "?곷떞??理쒓렐 ?몄뀡 議고쉶", "agent_id,limit", "PHASE1"),
        ("MessageRepository.insertMessage", "硫붿떆吏 ???, "session_id,role,content_masked", "PHASE1"),
        ("MessageRepository.retryMessage", "?ъ쟾??硫붿떆吏 ?앹꽦", "message_id,idempotency_key", "PHASE1"),
        ("StreamEventRepository.appendEvent", "SSE ?대깽?????, "message_id,event_type,event_id", "PHASE1"),
        ("StreamEventRepository.listByMessage", "resume ?대깽??議고쉶", "message_id,last_event_id", "PHASE1"),
        ("CitationRepository.listByAnswer", "洹쇨굅 citation 議고쉶", "answer_id", "PHASE1"),
        ("GuardrailEventRepository.insert", "Fail-Closed/PII ?대깽??湲곕줉", "trace_id,event_type,reason", "PHASE1"),
        ("TemplateRepository.listApproved", "?뱀씤 ?쒗뵆由?議고쉶", "tenant_key,category", "PHASE1"),
        ("AttachmentRepository.completeUpload", "泥⑤? ?낅줈???꾨즺 泥섎━", "attachment_id,etag", "PHASE2"),
        ("FeedbackRepository.upsertCsat", "CSAT ???, "session_id,score,resolved_flag", "PHASE2"),
        ("IntegrationLogRepository.appendHandoff", "handoff ?곕룞 濡쒓렇", "handoff_id,status,trace_id", "PHASE2"),
        ("PolicyVersionRepository.listByStatus", "?뺤콉 踰꾩쟾 議고쉶", "policy_id,status", "PHASE1"),
        ("PromptVersionRepository.listByStatus", "?꾨＼?꾪듃 踰꾩쟾 議고쉶", "prompt_id,status", "PHASE1"),
        ("AuditLogRepository.queryByTraceId", "trace_id 媛먯궗議고쉶", "trace_id,from,to", "PHASE1"),
        ("OpsMetricRepository.getSummary", "?댁쁺 KPI ?붿빟", "tenant_id,time_range", "PHASE1"),
        ("TenantQuotaRepository.getQuotaState", "荑쇳꽣 ?곹깭 議고쉶", "tenant_id,plan_id", "PHASE1"),
        ("KbIngestRepository.createJob", "KB ?ъ깋???묒뾽 ?앹꽦", "kb_id,index_version", "PHASE2"),
        ("ProviderKeyRepository.bindSecretRef", "provider secret_ref 諛붿씤??, "provider_id,secret_ref", "PHASE2"),
    ]
    for i, row in enumerate(rows, start=4):
        setv(ws, i, 1, i - 3)
        for c, v in enumerate(row, start=2):
            setv(ws, i, c, v)


def fill_rbac(ws) -> None:
    clear(ws, 1, 260, 1, 6)
    setv(ws, 1, 1, "沅뚰븳 蹂?UI 留ㅽ븨 (?쒕쾭 RBAC 理쒖쥌 沅뚯쐞)")
    setv(ws, 52, 1, "??UI ?④?? 蹂댁“?대ŉ, ?ㅼ젣 ?묎렐?듭젣???쒕쾭 403 ?묐떟??理쒖쥌 沅뚯쐞")
    setv(ws, 54, 1, "沅뚰븳")
    setv(ws, 54, 2, "沅뚰븳 遺??愿由ъ옄)")
    setv(ws, 54, 3, "沅뚰븳 遺???ъ슜??")
    setv(ws, 54, 4, "硫붾돱")
    setv(ws, 54, 5, "硫붾돱紐?)
    menus = {
        "AGENT": [("M-AGT-01", "AGT001 濡쒓렇???몄뀡"), ("M-AGT-02", "AGT002 ??), ("M-AGT-03", "AGT003 ????ㅽ듃由щ컢"), ("M-AGT-04", "AGT004 洹쇨굅?⑤꼸"), ("M-AGT-05", "AGT005 ?쒗뵆由우텛泥?), ("M-AGT-06", "AGT006 ?듬??몄쭛"), ("M-AGT-07", "AGT007 李⑤떒/safe_response")],
        "CUSTOMER": [("M-CUS-01", "CUS001 ?꾩젽遺?몄뒪?몃옪"), ("M-CUS-02", "CUS002 泥⑤?/硫붿떆吏"), ("M-CUS-03", "CUS003 CSAT/?몃뱶?ㅽ봽")],
        "ADMIN": [("M-ADM-01", "ADM001 ?뺤콉/?꾨＼?꾪듃?뱀씤"), ("M-ADM-02", "ADM002 ?쒗뵆由용쾭?꾨같??), ("M-ADM-03", "ADM003 KB/紐⑤뜽/Tool/MCP"), ("M-OPS-01", "OPS001 ??쒕낫??), ("M-OPS-02", "OPS002 媛먯궗濡쒓렇/利됱떆議곗튂")],
        "OPS": [("M-OPS-01", "OPS001 ??쒕낫??), ("M-OPS-02", "OPS002 媛먯궗濡쒓렇/利됱떆議곗튂"), ("M-ADM-03", "ADM003 議고쉶")],
        "SYSTEM": [("M-SYS-01", "?대? API/???ㅽ뻾(?붾㈃ ?몄텧 ?놁쓬)")],
    }
    r = 55
    for role, items in menus.items():
        for code, name in items:
            setv(ws, r, 1, role)
            setv(ws, r, 2, "Y" if role in {"ADMIN", "OPS", "SYSTEM"} else "")
            setv(ws, r, 3, "Y" if role in {"AGENT", "CUSTOMER"} else "")
            setv(ws, r, 4, code)
            setv(ws, r, 5, name)
            r += 1


def fill_inconsistency(ws, spec: dict) -> None:
    clear(ws, 1, 500, 1, 8)
    setv(ws, 1, 1, "遺덉씪移?紐⑸줉 (Inconsistencies)")
    setv(ws, 3, 1, "ID/??ぉ")
    setv(ws, 3, 2, "異⑸룎 ?뚯뒪 A")
    setv(ws, 3, 3, "異⑸룎 ?뚯뒪 B")
    setv(ws, 3, 4, "異⑸룎 ?댁슜")
    setv(ws, 3, 5, "寃곗젙(?곗꽑?쒖쐞 湲곕컲)")
    setv(ws, 3, 6, "?곹뼢 踰붿쐞(?붾㈃/API/DB)")
    setv(ws, 3, 7, "由ъ뒪???꾩냽 議곗튂")

    f_only = sorted(spec["feat_set"] - spec["req_set"])
    r_only = sorted(spec["req_set"] - spec["feat_set"])
    rows = [
        ("INC-001 ReqID ?곗＜ 李⑥씠", "Requirements.csv", "Summary of key features.csv", f"ReqID 吏묓빀 遺덉씪移?feature-only {len(f_only)}, req-only {len(r_only)})", "Requirements 湲곗? + feature 蹂묓빀 異붿쟻", "91_異붿쟻?깅ℓ?몃┃??, "?좉퇋 ID 異붽? ????臾몄꽌 ?숈떆 媛깆떊"),
        ("INC-002 CUSTOMER SSE 踰붿쐞", "?붽뎄?ы빆 UI-004/UI-005", "API-STREAM-SSE 沅뚰븳=AGENT", "怨좉컼 ?꾩젽 SSE 踰붿쐞 ?곸땐", "MVP??怨좉컼 baseline UI, 怨좉툒 SSE??Phase2", "10_CUS001,11_CUS002", "CUSTOMER SSE 沅뚰븳 API 寃??),
        ("INC-003 SSE ?⑹뼱", "Requirements token/chunk", "AGENTS/API token/tool/citation", "?대깽??紐낆묶 ?쇱옱", "?쒖? ?대깽?????AGENTS/API 湲곗? ?듭씪", "05_AGT003,06_AGT004", "?꾨줎???뚯꽌 蹂꾩묶 ?덉슜"),
        ("INC-004 ?쒗뵆由?異붿쿇 ?ㅽ뻾", "Features TMP-001", "AGENTS 8.3 踰꾪듉 ?꾩슜", "?먮룞 ?ㅽ뻾 媛?μ꽦", "踰꾪듉 ?≪뀡 ?녿뒗 ?몄텧 403 李⑤떒", "07_AGT005", "TOOL-003 ?뚭??뚯뒪??),
        ("INC-005 SP 以묒떖 ?쒗뵆由?, "37_SP?꾩껜紐⑸줉", "DB 紐낆꽭(SP ?놁쓬)", "?쒗뵆由욧낵 援ы쁽 諛⑹떇 李⑥씠", "A?? DB Access/Repository 紐⑸줉?쇰줈 ?ы빐??, "37_SP?꾩껜紐⑸줉", "諛깆뿏??援ы쁽 異붿쟻???좎?"),
        ("INC-006 沅뚰븳 紐낆묶", "API role(ADMIN/OPS)", "?붽뎄?ы빆(Manager/System Admin)", "沅뚰븳 ?쒗쁽 遺덉씪移?, "API role 湲곗? + UI?먯꽌 ?몃텇???쒖떆", "38_沅뚰븳 蹂?UI", "RBAC ?댁쁺?뺤콉 ?숆린??),
        ("INC-007 secret_ref", "AGENTS(?됰Ц ??湲덉?)", "TB_PROVIDER_KEY 議댁옱", "?됰Ц ????꾪뿕", "UI?먯꽌 secret_ref留??낅젰 ?덉슜", "15_ADM003,01_?먮윭硫붿떆吏肄붾뱶", "?쒕쾭 寃利??뚯뒪???꾩닔"),
        ("INC-008 Fail-Closed 媛뺣룄", "AGENTS fail-closed 媛뺤젣", "?쇰? ?붽뎄 臾멸뎄 ?ъ떆??以묒떖", "fallback ?먯쑀?띿뒪???ъ?", "safe_response only ?뺤콉 怨좎젙", "09_AGT007", "移댄뵾/QA ?먭?"),
        ("INC-009 MCP ?곗꽑?쒖쐞", "Features MCP-001/002(Phase2)", "API MCP ?붾뱶?ъ씤??議댁옱", "湲곕뒫 議댁옱 vs MVP 踰붿쐞", "MVP???깅줉/?ъ뒪 以묒떖, 怨좉툒 transport??Phase2", "15_ADM003", "?댁쁺 以鍮꾨룄 ?먭?"),
        ("INC-010 API ?섎웾 vs ?붾㈃ ?섎웾", "API 85媛?, "?듭떖 ?붾㈃ 15媛?, "1:1 ?붾㈃ 留ㅽ븨 遺덇?", "ReqID 湲곗? ?ㅻ???異붿쟻?깆쑝濡??꾨씫 理쒖냼??, "91_異붿쟻?깅ℓ?몃┃??, "?좉퇋 ?붾㈃ ?꾩슂 ??ARCHIVE ?쒗듃 ?ы솢??),
    ]
    for i, row in enumerate(rows, start=4):
        for c, v in enumerate(row, start=1):
            setv(ws, i, c, v)


def defaults():
    screen = {
        "SYS": "AGT-001,AGT-003,CUS-001", "UI": "AGT-003,AGT-004,AGT-006,CUS-002", "API": "AGT-003,CUS-002,OPS-001",
        "AI": "AGT-004,AGT-007,ADM-003", "RAG": "AGT-004,AGT-007,ADM-003", "TMP": "AGT-005,AGT-006,ADM-002",
        "KB": "AGT-004,ADM-003", "ADM": "ADM-001,ADM-002,ADM-003", "OPS": "OPS-001,OPS-002", "INT": "CUS-003,OPS-002",
        "SEC": "AGT-001,AGT-007,OPS-002", "TOOL": "AGT-005,ADM-003", "LLM": "ADM-003,OPS-001", "MCP": "ADM-003,OPS-001",
        "ETL": "ADM-003,OPS-001", "PERF": "AGT-003,OPS-001", "CCH": "AGT-002,AGT-003",
    }
    api = {
        "SYS": "API-SESSION-BOOTSTRAP:/v1/chat/bootstrap; API-SESSION-CREATE:/v1/sessions",
        "UI": "API-MESSAGE-POST:/v1/sessions/{session_id}/messages; API-STREAM-SSE:/v1/sessions/{session_id}/messages/{message_id}/stream",
        "API": "API-MESSAGE-POST:/v1/sessions/{session_id}/messages; API-STREAM-SSE:/v1/sessions/{session_id}/messages/{message_id}/stream",
        "AI": "API-RAG-RETRIEVE:/v1/rag/retrieve; API-RAG-ANSWER:/v1/rag/answer; API-RAG-CITATIONS-GET:/v1/rag/answers/{answer_id}/citations",
        "RAG": "API-RAG-RETRIEVE:/v1/rag/retrieve; API-RAG-ANSWER:/v1/rag/answer; API-RAG-CITATIONS-GET:/v1/rag/answers/{answer_id}/citations",
        "TMP": "TMP-RECOMMEND-BUTTON:/v1/sessions/{session_id}/template-recommendations; TMP-TEMPLATE-LIST:/v1/admin/templates",
        "KB": "KB-DOC-UPLOAD:/v1/admin/kb/documents; KB-REINDEX-REQUEST:/v1/admin/kb/reindex",
        "ADM": "ADM-POLICY-UPDATE:/v1/admin/policies/{policy_id}; ADM-VERSION-BUNDLE-ACTIVATE:/v1/admin/version-bundles/{bundle_id}/activate",
        "OPS": "OPS-METRIC-SUMMARY:/v1/ops/metrics/summary; OPS-AUDIT-LOG-QUERY:/v1/admin/audit-logs",
        "INT": "INT-HANDOFF-SYNC:/v1/integrations/crm/handoffs; INT-WORKFLOW-REPORT:/v1/ops/workflow/reports",
        "SEC": "API-AUTH-LOGIN:/v1/auth/login; ADM-RBAC-MATRIX-UPSERT:/v1/admin/rbac/matrix/{resource_key}",
        "TOOL": "TOOL-CALL-EXECUTE:/v1/internal/tools/execute; TOOL-ALLOWLIST-UPDATE:/v1/admin/tools/allowlist/{tool_name}",
        "LLM": "ADM-MODEL-LIST:/v1/admin/models; LLM-PROVIDER-HEALTH:/v1/ops/llm/providers/health",
        "MCP": "MCP-SERVER-UPSERT:/v1/admin/mcp/servers/{server_id}; MCP-SERVER-HEALTH:/v1/ops/mcp/servers/{server_id}/health",
        "ETL": "KB-REINDEX-REQUEST:/v1/admin/kb/reindex; OPS-KB-INDEX-STATUS:/v1/admin/kb/index-operations",
        "PERF": "API-STREAM-SSE:/v1/sessions/{session_id}/messages/{message_id}/stream; OPS-ADMIN-DASHBOARD-SERIES:/v1/admin/dashboard/series",
        "CCH": "API-MESSAGE-LIST:/v1/sessions/{session_id}/messages; OPS-METRIC-SUMMARY:/v1/ops/metrics/summary",
    }
    db = {
        "SYS": "TB_AUTH_SESSION,TB_CONVERSATION,TB_WIDGET_INSTANCE", "UI": "TB_MESSAGE,TB_STREAM_EVENT,TB_ATTACHMENT,TB_MESSAGE_ATTACHMENT",
        "API": "TB_API_METRIC_HOURLY,TB_MESSAGE,TB_AUDIT_LOG", "AI": "TB_RAG_SEARCH_LOG,TB_RAG_CITATION,TB_GUARDRAIL_EVENT",
        "RAG": "TB_RAG_SEARCH_LOG,TB_RAG_CITATION,TB_KB_CHUNK", "TMP": "TB_TEMPLATE,TB_TEMPLATE_VERSION,TB_TEMPLATE_PLACEHOLDER,TB_TEMPLATE_POLICY_MAP",
        "KB": "TB_KB_DOCUMENT,TB_KB_DOCUMENT_VERSION,TB_KB_INDEX_VERSION,TB_KB_INGEST_JOB", "ADM": "TB_POLICY_VERSION,TB_PROMPT_VERSION,TB_TEMPLATE_VERSION,TB_AUDIT_LOG",
        "OPS": "TB_OPS_EVENT,TB_AUDIT_LOG,TB_TENANT_QUOTA,TB_API_METRIC_HOURLY", "INT": "TB_INTEGRATION_LOG,TB_WEBHOOK_ENDPOINT,TB_AUDIT_LOG",
        "SEC": "TB_ROLE,TB_PERMISSION,TB_ROLE_PERMISSION,TB_TENANT_QUOTA", "TOOL": "TB_TOOL_DEFINITION,TB_TOOL_CALL_LOG,TB_AUDIT_LOG",
        "LLM": "TB_LLM_MODEL,TB_LLM_PROVIDER,TB_LLM_ROUTE,TB_PROVIDER_KEY", "MCP": "TB_TOOL_DEFINITION,TB_TOOL_CALL_LOG,TB_AUDIT_LOG",
        "ETL": "TB_KB_INGEST_JOB,TB_VECTOR_INDEX_STATUS,TB_KB_INDEX_VERSION", "PERF": "TB_API_METRIC_HOURLY,TB_TENANT_USAGE_DAILY,TB_TENANT_USAGE_MONTHLY",
        "CCH": "TB_ANSWER_BANK,TB_SEMANTIC_CACHE,TB_TENANT_USAGE_DAILY",
    }
    tel = {
        "SYS": "trace_id,tenant_id,session_id,message_id,session_create_ms,trace_coverage_pct", "UI": "trace_id,tenant_id,session_id,message_id,first_token_ms,reconnect_count",
        "API": "trace_id,tenant_id,endpoint,http_status,latency_ms,error_code", "AI": "trace_id,tenant_id,message_id,citation_count,evidence_score,fail_closed_count",
        "RAG": "trace_id,tenant_id,message_id,citation_count,contract_pass_rate", "TMP": "trace_id,tenant_id,session_id,template_id,cooldown_block_count,placeholder_fill_rate",
        "KB": "trace_id,tenant_id,kb_id,index_version,reindex_latency_ms", "ADM": "trace_id,tenant_id,deploy_id,approval_lead_time_ms,rollback_count",
        "OPS": "trace_id,tenant_id,event_type,p95_ms,error_rate,cost", "INT": "trace_id,tenant_id,handoff_id,delivery_status,retry_count",
        "SEC": "trace_id,tenant_id,user_role,auth_fail_count,quota_block_count", "TOOL": "trace_id,tenant_id,tool_name,tool_call_count,tool_error_rate",
        "LLM": "trace_id,tenant_id,provider_id,model_id,provider_health_status", "MCP": "trace_id,tenant_id,mcp_server_id,mcp_health_status,tool_error_rate",
        "ETL": "trace_id,tenant_id,index_version,chunk_count,index_build_ms", "PERF": "trace_id,tenant_id,first_token_ms,p50_ms,p95_ms,p99_ms",
        "CCH": "trace_id,tenant_id,cache_hit_rate,token_per_response,latency_ms",
    }
    return screen, api, db, tel


def fill_trace(ws, spec: dict) -> None:
    clear(ws, 1, 2500, 1, 8)
    setv(ws, 1, 1, "?붽뎄?ы빆 異붿쟻??留ㅽ듃由?뒪 (ReqID ??Screen ??API ??DB ??Telemetry ??TC)")
    setv(ws, 2, 1, "??? Requirements Statement + Summary of key features ?꾩껜 ReqID")
    setv(ws, 3, 1, "ReqID")
    setv(ws, 3, 2, "Screen ID")
    setv(ws, 3, 3, "API(ProgramID/Endpoint)")
    setv(ws, 3, 4, "DB(Table)")
    setv(ws, 3, 5, "Telemetry(?꾩닔??")
    setv(ws, 3, 6, "?뚯뒪??耳?댁뒪(TC)")
    d_screen, d_api, d_db, d_tel = defaults()
    r = 4
    for rid in spec["all_req"]:
        pfx = rid.split("-", 1)[0]
        apis = spec["api_by_req"].get(rid, [])
        if apis:
            api_text = "; ".join(f"{x['pid']}:{x['method']} {x['ep']}" for x in apis[:4])
        else:
            api_text = d_api.get(pfx, "")
        setv(ws, r, 1, rid)
        setv(ws, r, 2, d_screen.get(pfx, "AGT-002"))
        setv(ws, r, 3, api_text)
        setv(ws, r, 4, d_db.get(pfx, "TB_AUDIT_LOG"))
        setv(ws, r, 5, d_tel.get(pfx, "trace_id,tenant_id,event_type"))
        setv(ws, r, 6, f"TC-{rid}-01")
        r += 1


def fill_template_analysis(ws, wb, active: list[tuple[str, str, str]]) -> None:
    clear(ws, 1, 500, 1, 8)
    setv(ws, 1, 1, "?쒗뵆由?遺꾩꽍 ?붿빟")
    setv(ws, 3, 1, "??ぉ")
    setv(ws, 3, 2, "遺꾩꽍 寃곌낵")
    setv(ws, 3, 3, "?곸슜 寃곗젙")
    setv(ws, 3, 4, "鍮꾧퀬")
    rows = [
        ("?듯빀紐⑹감 援ъ“", "No/?쒗듃紐?Phase/?ㅻ챸", "?숈씪 ?뺤떇 ?좎?", "00_?듯빀紐⑹감"),
        ("?먮윭肄붾뱶 援ъ“", "硫붿떆吏肄붾뱶/硫붿떆吏 ?댁슜/HTTP/?ㅻ챸", "?숈씪 ?뺤떇 ?좎?", "01_?먮윭硫붿떆吏肄붾뱶"),
        ("醫낇빀肄붾뱶 援ъ“", "肄붾뱶洹몃９/肄붾뱶媛?肄붾뱶紐??ㅻ챸", "?숈씪 ?뺤떇 ?좎?", "02_異붽?醫낇빀肄붾뱶 + 35_異붽?醫낇빀肄붾뱶_2"),
        ("?붾㈃?쒗듃 ?⑦꽩", "?꾨줈洹몃옩?뺣낫/紐⑹쟻/?낅젰/踰꾪듉/?뚯씠釉??덉쇅/TC", "?숈씪 ?⑦꽩 ?ъ궗??, "03~17"),
        ("SP ?쒗듃", "Stored Procedure 紐⑸줉", "A?? DB Access/Repository 紐⑸줉?쇰줈 ?ы빐??, "37_SP?꾩껜紐⑸줉"),
        ("沅뚰븳蹂?UI", "沅뚰븳蹂?硫붾돱 留ㅽ븨", "ROLE 湲곕컲 留ㅽ븨 + ?쒕쾭 403 紐낆떆", "38_沅뚰븳 蹂?UI"),
        ("?꾩닔 ?뺤옣", "?쒗뵆由?誘명룷??, "90_遺덉씪移섎ぉ濡? 91_異붿쟻?깅ℓ?몃┃??異붽?", "?곸슜??異붿쟻??蹂닿컯"),
    ]
    for i, row in enumerate(rows, start=4):
        for c, v in enumerate(row, start=1):
            setv(ws, i, c, v)
    setv(ws, 20, 1, "?쒖꽦 ?쒗듃 紐⑸줉")
    setv(ws, 21, 1, "No")
    setv(ws, 21, 2, "?쒗듃紐?)
    setv(ws, 21, 3, "Phase")
    setv(ws, 21, 4, "?ㅻ챸")
    for i, row in enumerate(active, start=22):
        setv(ws, i, 1, i - 21)
        for c, v in enumerate(row, start=2):
            setv(ws, i, c, v)


def fill_toc(ws, entries: list[tuple[str, str, str]]) -> None:
    clear(ws, 1, 800, 1, 6)
    setv(ws, 1, 1, "CS RAG UI/UX ?ㅺ퀎??(?듯빀蹂?")
    setv(ws, 3, 1, "No")
    setv(ws, 3, 2, "?쒗듃紐?)
    setv(ws, 3, 3, "Phase")
    setv(ws, 3, 4, "?ㅻ챸")
    for i, (sheet, phase, desc) in enumerate(entries, start=4):
        setv(ws, i, 1, f"{i-3:02d}")
        setv(ws, i, 2, sheet)
        setv(ws, i, 3, phase)
        setv(ws, i, 4, desc)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    template = find_template()
    shutil.copy2(template, OUT_PATH)
    spec = load_spec()
    wb = load_workbook(OUT_PATH)

    ws00 = by_prefix(wb, "00_")
    ws01 = by_prefix(wb, "01_")
    ws02 = by_prefix(wb, "02_")
    ws35 = by_prefix(wb, "35_")
    ws37 = by_prefix(wb, "37_")
    ws38 = by_prefix(wb, "38_")

    screens_old = [
        n for n in wb.sheetnames
        if re.match(r"^\d{2}_", n) and not n.startswith(("00_", "01_", "02_", "35_", "37_", "38_"))
    ]
    screens_old.sort(key=num_prefix)

    if len(screens_old) < len(SCREENS):
        raise RuntimeError("?쒗뵆由??붾㈃ ?쒗듃媛 遺議깊빀?덈떎.")

    for sc, old in zip(SCREENS, screens_old):
        ws = wb[old]
        ws.title = uniq_name(wb, sc["sheet"])
        sc["sheet"] = ws.title

    # Drop leftover template screens to avoid carrying unrelated legacy content.
    for old in screens_old[len(SCREENS):]:
        if old in wb.sheetnames:
            wb.remove(wb[old])

    for sc in SCREENS:
        fill_screen(wb[sc["sheet"]], sc, spec["api_by_pid"])

    fill_errors(ws01)
    fill_codes(ws02, ws35)
    fill_sp(ws37)
    fill_rbac(ws38)

    ws90 = wb.copy_worksheet(ws37); ws90.title = uniq_name(wb, "90_遺덉씪移섎ぉ濡?)
    ws91 = wb.copy_worksheet(ws37); ws91.title = uniq_name(wb, "91_異붿쟻?깅ℓ?몃┃??)
    ws92 = wb.copy_worksheet(ws37); ws92.title = uniq_name(wb, "92_?쒗뵆由용텇?앹슂??)
    fill_inconsistency(ws90, spec)
    fill_trace(ws91, spec)

    active = [
        (ws01.title, "?꾩껜", "?먮윭/?깃났 硫붿떆吏 肄붾뱶 移댄깉濡쒓렇"),
        (ws02.title, "?꾩껜", "怨듯넻 肄붾뱶 ?뺤쓽"),
    ]
    for sc in SCREENS:
        active.append((sc["sheet"], sc["phase"], f"{sc['sid']} {sc['name']}"))
    active += [
        (ws35.title, "PHASE2", "異붽? 醫낇빀肄붾뱶"),
        (ws37.title, "?꾩껜", "DB Access / Repository / Query 紐⑸줉"),
        (ws38.title, "?꾩껜", "沅뚰븳蹂?UI ?묎렐 留ㅽ븨"),
        (ws90.title, "?꾩껜", "?ㅽ럺 遺덉씪移?紐⑸줉 諛?寃곗젙"),
        (ws91.title, "?꾩껜", "ReqID?뭆creen?묨PI?묭B 異붿쟻??留ㅽ듃由?뒪"),
        (ws92.title, "?꾩껜", "?쒗뵆由?遺꾩꽍 ?붿빟"),
    ]

    fill_template_analysis(ws92, wb, active)
    fill_toc(ws00, active)
    wb.save(OUT_PATH)
    print(f"Generated: {OUT_PATH}")


if __name__ == "__main__":
    main()
