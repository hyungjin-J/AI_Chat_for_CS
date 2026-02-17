#!/usr/bin/env python3
"""Refresh UIUX workbook design and replace legacy screenshots.

Applies a polished table style inspired by docs/references/CS_AI_CHATBOT_DB.xlsx,
removes unrelated ARCHIVE sheets, and regenerates project-specific UI mock images.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
UIUX_DIR = ROOT / "docs" / "uiux"
ASSET_DIR = UIUX_DIR / "assets"


@dataclass(frozen=True)
class ScreenSpec:
    file_name: str
    title: str
    subtitle: str
    chips: tuple[str, ...]
    left_menu: tuple[str, ...]
    timeline: tuple[str, ...]
    right_panel: tuple[str, ...]


SCREEN_SPECS: dict[str, ScreenSpec] = {
    "AGT_001_Bootstrap.png": ScreenSpec(
        "AGT_001_Bootstrap.png",
        "AGT-001 Agent Bootstrap",
        "Tenant routing + session restore + policy hydration",
        ("trace_id", "tenant_key", "rbac=agent"),
        ("Dashboard", "Conversations", "Templates", "Evidence", "Ops"),
        ("session bootstrap", "policy bundle load", "recent messages"),
        ("X-Trace-Id", "X-Tenant-Key", "locale=ko-KR"),
    ),
    "AGT_002_Home.png": ScreenSpec(
        "AGT_002_Home.png",
        "AGT-002 Home",
        "Recent sessions, alerts, and quality gates",
        ("P95 1.8s", "fail_closed 0.9%", "csat 4.6"),
        ("Inbox", "SLA", "Escalations", "KB", "Audit"),
        ("priority queue", "pending reviews", "safe_response alerts"),
        ("tenant quota", "tool calls", "ops notice"),
    ),
    "AGT_003_Conversation_Stream.png": ScreenSpec(
        "AGT_003_Conversation_Stream.png",
        "AGT-003 Conversation Streaming",
        "token -> citation -> done with resume-safe UX",
        ("SSE", "last_event_id", "resume"),
        ("Conversation", "Citations", "Templates", "History"),
        ("token stream", "tool event", "citation mapped", "done"),
        ("first token", "retry-after", "contract status"),
    ),
    "AGT_004_Citation_Panel.png": ScreenSpec(
        "AGT_004_Citation_Panel.png",
        "AGT-004 Citation Panel",
        "Evidence-grounded answer with source/version mapping",
        ("RAG", "evidence>=0.82", "citation required"),
        ("Answer", "Citations", "Policy", "KB Chunks"),
        ("answer draft", "source mapping", "policy check"),
        ("doc_id", "chunk_id", "version"),
    ),
    "AGT_005_Template_Recommend.png": ScreenSpec(
        "AGT_005_Template_Recommend.png",
        "AGT-005 Template Recommendation",
        "Button-triggered recommendation only with cooldown/budget",
        ("button only", "cooldown", "budget guard"),
        ("Templates", "Variables", "Policy Map", "History"),
        ("button action", "candidate rank", "placeholder fill"),
        ("remaining calls", "reason code", "approval version"),
    ),
    "AGT_006_Editor_and_Gate.png": ScreenSpec(
        "AGT_006_Editor_and_Gate.png",
        "AGT-006 Editor and Gate",
        "Draft editing with contract validation and policy lint",
        ("schema", "policy", "pii masked"),
        ("Editor", "Preview", "Checks", "Send"),
        ("draft", "lint pass", "citation verify"),
        ("error_code", "trace link", "retry path"),
    ),
    "AGT_007_Blocked_Safe_Response.png": ScreenSpec(
        "AGT_007_Blocked_Safe_Response.png",
        "AGT-007 Blocked / Safe Response",
        "Fail-closed when schema/citation/evidence checks fail",
        ("fail-closed", "safe_response", "audit log"),
        ("Gate", "Violation", "Safe Reply", "Escalate"),
        ("validation fail", "delivery blocked", "safe_response"),
        ("reason", "required action", "audit id"),
    ),
    "CUS_001_Widget_Bootstrap.png": ScreenSpec(
        "CUS_001_Widget_Bootstrap.png",
        "CUS-001 Widget Bootstrap",
        "Customer widget init with tenant skin and trace propagation",
        ("widget", "session create", "locale"),
        ("Help", "Orders", "Delivery", "Refund"),
        ("widget mount", "session create", "history restore"),
        ("channel_id", "session_id", "expires_at"),
    ),
    "CUS_002_Widget_Attachment.png": ScreenSpec(
        "CUS_002_Widget_Attachment.png",
        "CUS-002 Attachment and Message",
        "Presign upload + completion + message send",
        ("attachment", "presign", "complete"),
        ("Chat", "Upload", "Preview", "Send"),
        ("select file", "upload chunk", "message post"),
        ("mime/type", "size guard", "virus check"),
    ),
    "CUS_003_Widget_CSAT_Handoff.png": ScreenSpec(
        "CUS_003_Widget_CSAT_Handoff.png",
        "CUS-003 CSAT and Handoff",
        "CSAT capture and CRM handoff with trace continuity",
        ("csat", "handoff", "crm sync"),
        ("CSAT", "Resolved", "Need Agent", "History"),
        ("csat submit", "handoff request", "crm sync"),
        ("handoff_id", "status", "timeline"),
    ),
    "ADM_001_Ops_Dashboard.png": ScreenSpec(
        "ADM_001_Ops_Dashboard.png",
        "ADM-001 Governance Dashboard",
        "Policy/prompt approval workflow with rollback controls",
        ("draft", "approved", "rollback"),
        ("Policies", "Prompts", "Approvals", "Audit"),
        ("review queue", "typed confirm", "activate version"),
        ("approver", "change diff", "published at"),
    ),
    "ADM_002_Governance_Approval.png": ScreenSpec(
        "ADM_002_Governance_Approval.png",
        "ADM-002 Approval and Deploy",
        "Versioned template deploy flow with safety checkpoints",
        ("versioning", "approval", "deploy"),
        ("Template", "Version", "Approval", "Deploy"),
        ("candidate select", "approval gate", "deploy done"),
        ("bundle id", "rollback target", "event id"),
    ),
    "ADM_003_KB_Model_Tools.png": ScreenSpec(
        "ADM_003_KB_Model_Tools.png",
        "ADM-003 KB / Model / Tool",
        "Knowledge index and model/tool allowlist operations",
        ("kb", "llm", "tool allowlist"),
        ("KB Docs", "Models", "Tools", "MCP"),
        ("upload doc", "index job", "activate model"),
        ("index status", "provider health", "tool errors"),
    ),
    "OPS_001_Audit_Killswitch.png": ScreenSpec(
        "OPS_001_Audit_Killswitch.png",
        "OPS-001 Audit and Kill Switch",
        "trace_id drill-down, emergency block, and provider kill switch",
        ("ops", "audit", "kill-switch"),
        ("Trace", "Alerts", "Blocks", "Rollbacks"),
        ("trace query", "incident triage", "block apply"),
        ("provider", "block scope", "rollback id"),
    ),
}


SHEET_IMAGE_MAP = {
    "03_AGT001_": ("AGT_001_Bootstrap.png", "A10"),
    "04_AGT002_": ("AGT_002_Home.png", "A7"),
    "05_AGT003_": ("AGT_003_Conversation_Stream.png", "A7"),
    "06_AGT004_": ("AGT_004_Citation_Panel.png", "A7"),
    "07_AGT005_": ("AGT_005_Template_Recommend.png", "A9"),
    "08_AGT006_": ("AGT_006_Editor_and_Gate.png", "A9"),
    "09_AGT007_": ("AGT_007_Blocked_Safe_Response.png", "A9"),
    "10_CUS001_": ("CUS_001_Widget_Bootstrap.png", "A9"),
    "11_CUS002_": ("CUS_002_Widget_Attachment.png", "A9"),
    "12_CUS003_": ("CUS_003_Widget_CSAT_Handoff.png", "A10"),
    "13_ADM001_": ("ADM_001_Ops_Dashboard.png", "A9"),
    "14_ADM002_": ("ADM_002_Governance_Approval.png", "A6"),
    "15_ADM003_": ("ADM_003_KB_Model_Tools.png", "A9"),
    "16_OPS001_": ("ADM_001_Ops_Dashboard.png", "A13"),
    "17_OPS002_": ("OPS_001_Audit_Killswitch.png", "A9"),
}


THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT_BASE = Font(name="Malgun Gothic", size=10, bold=False, color="0F172A")
FONT_BOLD = Font(name="Malgun Gothic", size=10, bold=True, color="0F172A")
FONT_TITLE = Font(name="Malgun Gothic", size=13, bold=True, color="FFFFFF")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
FILL_TITLE = PatternFill("solid", fgColor="1D4ED8")
FILL_SECTION = PatternFill("solid", fgColor="EAF1FF")
FILL_TABLE_HEADER = PatternFill("solid", fgColor="F1F5F9")
FILL_ALT = PatternFill("solid", fgColor="F8FAFC")


def find_workbook() -> Path:
    files = sorted(UIUX_DIR.glob("CS_RAG_UI_UX_*.xlsx"))
    if not files:
        raise FileNotFoundError("UIUX workbook not found under docs/uiux")
    return files[0]


def sheet_phase(sheet_name: str) -> str:
    if sheet_name.startswith(("03_", "04_", "05_", "06_", "07_", "08_", "09_")):
        return "PHASE1"
    if sheet_name.startswith(("10_", "11_", "12_", "15_")):
        return "PHASE2"
    if sheet_name.startswith(("13_", "14_", "16_", "17_")):
        return "PHASE1"
    if sheet_name.startswith("35_"):
        return "PHASE2"
    return "전체"


def is_section_label(value: str) -> bool:
    return bool(
        re.match(r"^\d+(\-\d+)?\.", value.strip())
        or value.strip().startswith("부록")
        or "체크리스트" in value
    )


def is_table_header(values: list[str]) -> bool:
    if len(values) < 3:
        return False
    short_text = [v for v in values if isinstance(v, str) and len(v) <= 24]
    return len(short_text) >= 3


def iter_data_rows(ws) -> Iterable[int]:
    last = min(ws.max_row, 420)
    for row in range(1, last + 1):
        if any(ws.cell(row=row, column=col).value not in (None, "") for col in range(1, 9)):
            yield row


def apply_sheet_style(ws) -> None:
    for col, width in zip("ABCDEFGH", (30, 26, 18, 30, 22, 18, 16, 16)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"

    section_rows: set[int] = set()
    table_header_rows: set[int] = set()

    for row in iter_data_rows(ws):
        values = [ws.cell(row=row, column=col).value for col in range(1, 9)]
        text_values = [v for v in values if isinstance(v, str) and v.strip()]
        if row == 1:
            continue
        if text_values and is_section_label(text_values[0]):
            section_rows.add(row)
        elif is_table_header(text_values):
            table_header_rows.add(row)

    for row in iter_data_rows(ws):
        has_data = False
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if cell.value in (None, ""):
                continue
            has_data = True
            cell.border = BORDER
            if row == 1:
                cell.fill = FILL_TITLE
                cell.font = FONT_TITLE
                cell.alignment = ALIGN_LEFT
            elif row in section_rows:
                cell.fill = FILL_SECTION
                cell.font = FONT_BOLD
                cell.alignment = ALIGN_LEFT
            elif row in table_header_rows:
                cell.fill = FILL_TABLE_HEADER
                cell.font = FONT_BOLD
                cell.alignment = ALIGN_CENTER if col > 1 else ALIGN_LEFT
            else:
                cell.fill = FILL_ALT if row % 2 == 0 else PatternFill(fill_type=None)
                cell.font = FONT_BASE
                cell.alignment = ALIGN_LEFT if col <= 2 else ALIGN_CENTER

        if row == 1 and has_data:
            ws.row_dimensions[row].height = 28
        elif has_data:
            ws.row_dimensions[row].height = 21


def font_paths() -> tuple[Path | None, Path | None]:
    regular = Path("C:/Windows/Fonts/malgun.ttf")
    bold = Path("C:/Windows/Fonts/malgunbd.ttf")
    return (regular if regular.exists() else None, bold if bold.exists() else None)


def load_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    regular, bold_path = font_paths()
    chosen = bold_path if bold else regular
    if chosen:
        return ImageFont.truetype(str(chosen), size=size)
    return ImageFont.load_default()


def rounded(draw: ImageDraw.ImageDraw, xy: tuple[int, int, int, int], fill: str, outline: str = "") -> None:
    draw.rounded_rectangle(xy, radius=12, fill=fill, outline=outline if outline else None, width=1)


def draw_mock(spec: ScreenSpec, out_path: Path) -> None:
    w, h = 1600, 900
    img = Image.new("RGB", (w, h), "#E2E8F0")
    draw = ImageDraw.Draw(img)

    f_title = load_font(34, bold=True)
    f_sub = load_font(18, bold=False)
    f_chip = load_font(16, bold=True)
    f_body = load_font(15, bold=False)
    f_menu = load_font(16, bold=True)

    # App shell
    rounded(draw, (30, 24, w - 30, h - 24), "#F8FAFC", "#CBD5E1")
    rounded(draw, (46, 40, w - 46, 98), "#1E3A8A")
    draw.text((70, 57), "CS RAG Support Console", fill="white", font=f_menu)
    draw.text((w - 380, 57), "trace_id: 8ce4...  tenant: kr-main", fill="#DBEAFE", font=f_body)

    # Sidebar
    rounded(draw, (46, 114, 280, h - 40), "#0F172A")
    y = 140
    for idx, item in enumerate(spec.left_menu):
        color = "#1E293B" if idx != 1 else "#1E3A8A"
        rounded(draw, (60, y, 266, y + 42), color)
        draw.text((74, y + 12), item, fill="#E2E8F0", font=f_body)
        y += 52

    # Main panel and side panel
    rounded(draw, (300, 114, 1090, h - 40), "#FFFFFF", "#CBD5E1")
    rounded(draw, (1110, 114, w - 46, h - 40), "#F1F5F9", "#CBD5E1")
    draw.text((330, 140), spec.title, fill="#0F172A", font=f_title)
    draw.text((330, 188), spec.subtitle, fill="#334155", font=f_sub)

    # Chips
    chip_x = 330
    for chip in spec.chips:
        rounded(draw, (chip_x, 228, chip_x + 190, 268), "#DBEAFE")
        draw.text((chip_x + 16, 242), chip, fill="#1E3A8A", font=f_chip)
        chip_x += 206

    # Timeline cards
    card_y = 292
    for idx, step in enumerate(spec.timeline, start=1):
        rounded(draw, (330, card_y, 1048, card_y + 92), "#F8FAFC", "#CBD5E1")
        draw.text((352, card_y + 16), f"Step {idx}", fill="#1E3A8A", font=f_chip)
        draw.text((460, card_y + 18), step, fill="#0F172A", font=f_body)
        draw.text((460, card_y + 46), "policy pass · pii safe · trace linked", fill="#64748B", font=f_body)
        card_y += 108

    # Composer strip
    rounded(draw, (330, h - 130, 1048, h - 68), "#FFFFFF", "#94A3B8")
    draw.text((352, h - 108), "Compose evidence-based answer...", fill="#64748B", font=f_body)
    rounded(draw, (930, h - 122, 1032, h - 76), "#1E3A8A")
    draw.text((958, h - 108), "Send", fill="white", font=f_chip)

    # Right panel blocks
    draw.text((1134, 142), "Operational Context", fill="#0F172A", font=f_chip)
    block_y = 176
    for key in spec.right_panel:
        rounded(draw, (1128, block_y, w - 64, block_y + 66), "#FFFFFF", "#CBD5E1")
        draw.text((1146, block_y + 24), key, fill="#334155", font=f_body)
        block_y += 78

    rounded(draw, (1128, h - 180, w - 64, h - 82), "#DCFCE7", "#86EFAC")
    draw.text((1148, h - 150), "Answer Contract: PASS", fill="#166534", font=f_chip)
    draw.text((1148, h - 122), "citations=3  evidence=0.91", fill="#166534", font=f_body)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    img.save(out_path)


def regenerate_assets() -> None:
    for spec in SCREEN_SPECS.values():
        draw_mock(spec, ASSET_DIR / spec.file_name)


def rebuild_toc(ws, names: list[str]) -> None:
    for row in range(1, 450):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
    if not isinstance(ws["A1"], MergedCell):
        ws["A1"] = "CS RAG UI/UX 설계서 (정제본)"
    if not isinstance(ws["A3"], MergedCell):
        ws["A3"] = "No"
    if not isinstance(ws["B3"], MergedCell):
        ws["B3"] = "시트명"
    if not isinstance(ws["C3"], MergedCell):
        ws["C3"] = "Phase"
    if not isinstance(ws["D3"], MergedCell):
        ws["D3"] = "설명"
    r = 4
    for i, name in enumerate(names, start=1):
        for col, value in (
            (1, f"{i:02d}"),
            (2, name),
            (3, sheet_phase(name)),
            (4, "UIUX 설계 문서 구성 시트"),
        ):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = value
        r += 1


def purge_images(ws) -> None:
    if hasattr(ws, "_images"):
        ws._images = []


def match_sheet_prefix(sheet_name: str) -> tuple[str, str] | None:
    for prefix, mapping in SHEET_IMAGE_MAP.items():
        if sheet_name.startswith(prefix):
            return mapping
    return None


def apply_images(wb) -> None:
    for ws in wb.worksheets:
        mapping = match_sheet_prefix(ws.title)
        if not mapping:
            continue
        file_name, anchor = mapping
        image_path = ASSET_DIR / file_name
        if not image_path.exists():
            continue
        xl_img = XLImage(str(image_path))
        xl_img.width = 1120
        xl_img.height = 620
        ws.add_image(xl_img, anchor)


def refresh_workbook(path: Path) -> None:
    wb = load_workbook(path)

    # 1) Drop legacy archive sheets.
    archive_sheets = [name for name in wb.sheetnames if name.startswith("ARCHIVE_")]
    for name in archive_sheets:
        wb.remove(wb[name])

    # 2) Purge existing images first.
    for ws in wb.worksheets:
        purge_images(ws)

    # 3) Rebuild TOC based on remaining sheets.
    if "00_통합목차" in wb.sheetnames:
        keep_order = [name for name in wb.sheetnames if name != "00_통합목차"]
        rebuild_toc(wb["00_통합목차"], keep_order)

    # 4) Apply refreshed style.
    for ws in wb.worksheets:
        apply_sheet_style(ws)

    # 5) Add regenerated project-specific images.
    apply_images(wb)

    wb.save(path)


def main() -> None:
    workbook_path = find_workbook()
    regenerate_assets()
    refresh_workbook(workbook_path)
    print(f"Refreshed workbook: {workbook_path}")
    print(f"Regenerated assets: {ASSET_DIR}")


if __name__ == "__main__":
    main()
