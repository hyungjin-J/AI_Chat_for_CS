#!/usr/bin/env python3
"""Release gate checks for AI_Chatbot.

Checks:
1) If any of the 6 spec files changed, spec_sync_report.md must also be changed.
2) Requirements ReqID is the source of truth; references in other specs must be subsets.
3) Optional placeholder/TBD checks for critical fields.
4) Auto-generate Notion sync summary text when spec files changed.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REQ_PATH = ROOT / "docs" / "references" / "CS AI Chatbot_Requirements Statement.csv"
SUM_PATH = ROOT / "docs" / "references" / "Summary of key features.csv"
DEV_PATH = ROOT / "docs" / "references" / "Development environment.csv"
API_XLSX = ROOT / "docs" / "references" / "google_ready_api_spec_v0.3_20260216.xlsx"
DB_XLSX = ROOT / "docs" / "references" / "CS_AI_CHATBOT_DB.xlsx"
UIUX_GLOB = "docs/uiux/CS_RAG_UI_UX_*.xlsx"
UIUX_MD = ROOT / "docs" / "uiux" / "UIUX_Spec.md"
SPEC_SYNC_REPORT = ROOT / "spec_sync_report.md"

REQ_ID_RE = re.compile(r"\b[A-Z]{2,5}-\d{3}\b")
PLACEHOLDER_RE = re.compile(r"\bTBD\b|coverage-gap|placeholder|todo", re.IGNORECASE)

NOTION_LINKS: dict[str, str] = {
    "docs/references/Summary of key features.csv": "https://www.notion.so/2ed405a3a72081d594b2c3738b3c8149",
    "docs/references/CS AI Chatbot_Requirements Statement.csv": "https://www.notion.so/2ed405a3a720816594e4dc34972174ec",
    "docs/references/Development environment.csv": "https://www.notion.so/2ed405a3a72081d198e6f648e508b6e7",
    "docs/references/google_ready_api_spec_v0.3_20260216.xlsx": "https://www.notion.so/2ed405a3a720816594e4dc34972174ec",
    "docs/references/CS_AI_CHATBOT_DB.xlsx": "https://www.notion.so/2ed405a3a720812180d9d508b77f31a4",
}

CRITICAL_REQ_IDS = {
    "AI-004",
    "AI-005",
    "AI-009",
    "RAG-003",
    "PERF-001",
    "SYS-004",
    "OPS-001",
    "OPS-100",
    "OPS-102",
    "API-007",
    "SEC-004",
}


@dataclass(slots=True)
class CheckResult:
    name: str
    passed: bool
    detail: str
    samples: list[str]


def to_repo_rel(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def discover_uiux_spec() -> Path:
    candidates = sorted(ROOT.glob(UIUX_GLOB))
    if not candidates:
        raise FileNotFoundError(f"UIUX spec not found by pattern: {UIUX_GLOB}")
    return candidates[0]


def spec_files() -> set[str]:
    return {
        to_repo_rel(SUM_PATH),
        to_repo_rel(REQ_PATH),
        to_repo_rel(DEV_PATH),
        to_repo_rel(API_XLSX),
        to_repo_rel(DB_XLSX),
        to_repo_rel(discover_uiux_spec()),
    }


def run_git(args: list[str]) -> str:
    completed = subprocess.run(
        ["git", *args],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
    )
    if completed.returncode != 0:
        raise RuntimeError(f"git {' '.join(args)} failed: {completed.stderr.strip()}")
    return completed.stdout


def changed_files_from_git(base_ref: str | None, head_ref: str | None) -> set[str]:
    if base_ref:
        range_ref = f"{base_ref}...{head_ref or 'HEAD'}"
        output = run_git(["diff", "--name-only", range_ref])
        return {line.strip() for line in output.splitlines() if line.strip()}

    try:
        output = run_git(["diff", "--name-only", "HEAD~1...HEAD"])
    except RuntimeError:
        output = run_git(["diff", "--name-only"])
    return {line.strip() for line in output.splitlines() if line.strip()}


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.reader(fp)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def read_req_master() -> tuple[set[str], CheckResult]:
    headers, rows = read_csv(REQ_PATH)
    if not headers:
        return set(), CheckResult("requirements_master_exists", False, "empty file", [])

    req_index = headers.index("ReqID") if "ReqID" in headers else 0
    req_ids: list[str] = []
    malformed: list[str] = []

    for idx, row in enumerate(rows, start=2):
        if req_index >= len(row):
            malformed.append(f"row={idx}: missing ReqID column")
            continue
        req_id = row[req_index].strip()
        if not req_id:
            continue
        req_ids.append(req_id)
        if not REQ_ID_RE.fullmatch(req_id):
            malformed.append(f"row={idx}: {req_id}")

    duplicates = sorted({value for value in req_ids if req_ids.count(value) > 1})
    samples = malformed[:10] + [f"duplicate:{value}" for value in duplicates[:10]]
    passed = len(malformed) == 0 and len(duplicates) == 0 and len(req_ids) > 0
    detail = f"req_ids={len(set(req_ids))} malformed={len(malformed)} duplicates={len(duplicates)}"
    return set(req_ids), CheckResult("requirements_master_valid", passed, detail, samples)


def check_summary_reqids(req_master: set[str]) -> CheckResult:
    headers, rows = read_csv(SUM_PATH)
    if not headers:
        return CheckResult("summary_reqid_subset", False, "summary file is empty", [])

    id_col = None
    for index, name in enumerate(headers):
        if name.strip() in {"요구사항ID", "ReqID"}:
            id_col = index
            break

    if id_col is None:
        return CheckResult("summary_reqid_subset", False, "요구사항ID column not found", headers[:10])

    refs = {row[id_col].strip() for row in rows if id_col < len(row) and row[id_col].strip()}
    missing = sorted(refs - req_master)
    return CheckResult(
        "summary_reqid_subset",
        len(missing) == 0,
        f"summary_refs={len(refs)} missing={len(missing)}",
        missing[:20],
    )


def check_devenv_reqids(req_master: set[str]) -> CheckResult:
    headers, rows = read_csv(DEV_PATH)
    if not headers:
        return CheckResult("devenv_reqid_subset", False, "development environment file is empty", [])

    desc_col = None
    for index, name in enumerate(headers):
        if name.strip().lower() == "description":
            desc_col = index
            break

    if desc_col is None:
        return CheckResult("devenv_reqid_subset", False, "description column not found", headers[:10])

    refs: set[str] = set()
    for row in rows:
        if desc_col < len(row):
            refs.update(REQ_ID_RE.findall(row[desc_col]))

    missing = sorted(refs - req_master)
    return CheckResult(
        "devenv_reqid_subset",
        len(missing) == 0,
        f"devenv_refs={len(refs)} missing={len(missing)}",
        missing[:20],
    )


def find_api_sheet(path: Path):
    wb = load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        if ws.title.strip() == "전체API목록":
            return wb, ws
    if len(wb.worksheets) >= 2:
        return wb, wb.worksheets[1]
    return wb, wb.worksheets[0]


def check_api_reqids(req_master: set[str]) -> CheckResult:
    wb, ws = find_api_sheet(API_XLSX)
    refs: set[str] = set()

    max_row = min(ws.max_row, 600)
    for row in range(2, max_row + 1):
        note = ws.cell(row, 11).value
        if isinstance(note, str) and "ReqID" in note:
            refs.update(REQ_ID_RE.findall(note))

    wb.close()
    missing = sorted(refs - req_master)
    return CheckResult(
        "api_reqid_subset",
        len(missing) == 0,
        f"api_refs={len(refs)} missing={len(missing)}",
        missing[:20],
    )


def extract_uiux_reqids_from_markdown(path: Path) -> set[str]:
    lines = path.read_text(encoding="utf-8").splitlines()

    start_index = None
    for idx, line in enumerate(lines):
        if line.startswith("## ") and "ReqID 목록" in line:
            start_index = idx + 1
            break

    if start_index is None:
        return set()

    refs: set[str] = set()
    for line in lines[start_index:]:
        if line.startswith("## "):
            break
        if not line.startswith("|"):
            continue
        parts = [part.strip() for part in line.split("|")]
        if len(parts) < 3:
            continue
        candidate = parts[1]
        if REQ_ID_RE.fullmatch(candidate):
            refs.add(candidate)

    return refs


def check_uiux_reqids_from_xlsx(req_master: set[str], uiux_path: Path) -> CheckResult:
    wb = load_workbook(uiux_path, data_only=True, read_only=True)
    ws = None
    for candidate in wb.worksheets:
        if candidate.title.startswith("91_"):
            ws = candidate
            break

    if ws is None:
        wb.close()
        return CheckResult("uiux_reqid_subset", False, "sheet starting with 91_ not found", [])

    refs: set[str] = set()
    max_row = min(ws.max_row, 4000)
    for row in range(4, max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, str):
            candidate = value.strip()
            if REQ_ID_RE.fullmatch(candidate):
                refs.add(candidate)

    wb.close()
    missing = sorted(refs - req_master)
    return CheckResult(
        "uiux_reqid_subset",
        len(missing) == 0,
        f"uiux_refs={len(refs)} missing={len(missing)} source={uiux_path.name}:91_*",
        missing[:20],
    )


def check_uiux_reqids(req_master: set[str], uiux_path: Path, uiux_xlsx_changed: bool) -> CheckResult:
    if not uiux_xlsx_changed and UIUX_MD.exists():
        refs = extract_uiux_reqids_from_markdown(UIUX_MD)
        if refs:
            missing = sorted(refs - req_master)
            return CheckResult(
                "uiux_reqid_subset",
                len(missing) == 0,
                f"uiux_refs={len(refs)} missing={len(missing)} source=UIUX_Spec.md",
                missing[:20],
            )

    return check_uiux_reqids_from_xlsx(req_master, uiux_path)


def check_spec_sync_report_changed(
    changed_files: set[str],
    spec_changed: set[str],
    base_ref: str | None,
    head_ref: str | None,
) -> CheckResult:
    if not spec_changed:
        return CheckResult("spec_sync_report_required", True, "no spec file changes", [])

    report_rel = to_repo_rel(SPEC_SYNC_REPORT)
    if report_rel not in changed_files:
        return CheckResult(
            "spec_sync_report_required",
            False,
            "spec files changed but spec_sync_report.md not changed",
            sorted(spec_changed),
        )

    if base_ref:
        range_ref = f"{base_ref}...{head_ref or 'HEAD'}"
        diff_out = run_git(["diff", "--name-only", range_ref, "--", report_rel]).strip()
        if not diff_out:
            return CheckResult(
                "spec_sync_report_required",
                False,
                "spec_sync_report.md appears unchanged in diff range",
                sorted(spec_changed),
            )

    return CheckResult(
        "spec_sync_report_required",
        True,
        f"spec_sync_report.md updated for {len(spec_changed)} changed spec file(s)",
        sorted(spec_changed),
    )


def check_placeholders(uiux_path: Path) -> CheckResult:
    samples: list[str] = []

    req_headers, req_rows = read_csv(REQ_PATH)
    req_col = req_headers.index("ReqID") if "ReqID" in req_headers else 0
    detail_col = req_headers.index("상세 구현 가이드") if "상세 구현 가이드" in req_headers else 4

    for idx, row in enumerate(req_rows, start=2):
        if req_col >= len(row) or detail_col >= len(row):
            continue
        req_id = row[req_col].strip()
        detail = row[detail_col]
        if req_id in CRITICAL_REQ_IDS and isinstance(detail, str) and PLACEHOLDER_RE.search(detail):
            samples.append(f"requirements:{req_id}:row{idx}")

    wb_api, ws_api = find_api_sheet(API_XLSX)
    critical_programs = {
        "OPS-ADMIN-DASHBOARD-SUMMARY",
        "OPS-ADMIN-DASHBOARD-SERIES",
        "OPS-METRIC-SUMMARY",
        "OPS-TRACE-QUERY",
    }

    max_api_row = min(ws_api.max_row, 600)
    for row in range(2, max_api_row + 1):
        program_id = ws_api.cell(row, 3).value
        if program_id not in critical_programs:
            continue
        for col in (7, 8, 9, 11):
            value = ws_api.cell(row, col).value
            if isinstance(value, str) and PLACEHOLDER_RE.search(value):
                samples.append(f"api:{program_id}:r{row}c{col}")
    wb_api.close()

    wb_ui = load_workbook(uiux_path, data_only=True, read_only=True)
    ws_err = None
    for ws in wb_ui.worksheets:
        if ws.title.startswith("01_"):
            ws_err = ws
            break

    if ws_err is not None:
        critical_codes = {
            "AI-009-422-SCHEMA",
            "AI-009-409-CITATION",
            "AI-009-409-EVIDENCE",
            "SYS-004-409-TRACE",
            "API-008-429-BUDGET",
            "SEC-003-409-PII",
        }
        max_err_row = min(ws_err.max_row, 400)
        for row in range(4, max_err_row + 1):
            code = ws_err.cell(row, 1).value
            message = ws_err.cell(row, 2).value
            if code in critical_codes and isinstance(message, str) and PLACEHOLDER_RE.search(message):
                samples.append(f"uiux_error:{code}:row{row}")

    wb_ui.close()
    return CheckResult(
        "placeholder_tbd_critical_fields",
        len(samples) == 0,
        f"placeholder_hits={len(samples)}",
        samples[:30],
    )


def generate_notion_summary(spec_changed: set[str]) -> str:
    lines: list[str] = []
    lines.append("### Notion Update Summary (Auto-generated)")
    lines.append(f"- Generated at: {dt.datetime.now(tz=dt.timezone.utc).isoformat()}")
    lines.append("- Changed spec files:")
    for spec in sorted(spec_changed):
        lines.append(f"  - `{spec}`")

    lines.append("- Required Notion sync pages:")
    for spec in sorted(spec_changed):
        notion_url = NOTION_LINKS.get(spec)
        if notion_url:
            lines.append(f"  - `{spec}` -> {notion_url}")
        elif spec.endswith(".xlsx") and "CS_RAG_UI_UX_" in spec:
            lines.append(f"  - `{spec}` -> https://www.notion.so/UI-UX-2ee405a3a72080a58c93d967ef0f2444")
        else:
            lines.append(f"  - `{spec}` -> (check AGENTS.md mapping)")

    lines.append("- Required metadata updates on each page:")
    lines.append("  - Last synced at")
    lines.append("  - Source file")
    lines.append("  - Version or commit")
    lines.append("  - Change summary (3-10 lines)")
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run release gate checks.")
    parser.add_argument("--base-ref", help="Base ref for diff range. e.g. origin/main")
    parser.add_argument("--head-ref", default="HEAD", help="Head ref for diff range. default: HEAD")
    parser.add_argument(
        "--changed-files",
        help="Optional comma-separated changed file paths (bypass git diff).",
    )
    parser.add_argument(
        "--check-placeholders",
        action="store_true",
        help="Fail when placeholder/TBD remnants are found in critical fields.",
    )
    parser.add_argument(
        "--report-json",
        default="docs/release/release_check_report.json",
        help="Output JSON report path.",
    )
    parser.add_argument(
        "--notion-summary-out",
        help="Optional file path to write generated Notion summary text when spec changed.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    uiux_spec = discover_uiux_spec()

    if args.changed_files:
        changed_files = {item.strip() for item in args.changed_files.split(",") if item.strip()}
    else:
        changed_files = changed_files_from_git(args.base_ref, args.head_ref)

    specs = spec_files()
    spec_changed = {path for path in changed_files if path in specs}
    uiux_xlsx_changed = to_repo_rel(uiux_spec) in spec_changed

    req_master, req_master_check = read_req_master()
    checks: list[CheckResult] = [
        req_master_check,
        check_summary_reqids(req_master),
        check_devenv_reqids(req_master),
        check_api_reqids(req_master),
        check_uiux_reqids(req_master, uiux_spec, uiux_xlsx_changed),
        check_spec_sync_report_changed(changed_files, spec_changed, args.base_ref, args.head_ref),
    ]

    if args.check_placeholders:
        checks.append(check_placeholders(uiux_spec))

    if spec_changed:
        notion_summary = generate_notion_summary(spec_changed)
        print(notion_summary)
        if args.notion_summary_out:
            out_path = (ROOT / args.notion_summary_out).resolve()
            out_path.parent.mkdir(parents=True, exist_ok=True)
            out_path.write_text(notion_summary + "\n", encoding="utf-8")

    failed = [check for check in checks if not check.passed]
    passed_count = len(checks) - len(failed)

    report_payload = {
        "generated_at": dt.datetime.now(tz=dt.timezone.utc).isoformat(),
        "base_ref": args.base_ref,
        "head_ref": args.head_ref,
        "changed_files_count": len(changed_files),
        "spec_changed_count": len(spec_changed),
        "spec_changed": sorted(spec_changed),
        "pass_count": passed_count,
        "fail_count": len(failed),
        "checks": [
            {
                "name": check.name,
                "passed": check.passed,
                "detail": check.detail,
                "samples": check.samples,
            }
            for check in checks
        ],
    }

    report_path = ROOT / args.report_json
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"PASS={passed_count} FAIL={len(failed)}")
    for check in checks:
        status = "PASS" if check.passed else "FAIL"
        print(f"[{status}] {check.name}: {check.detail}")
        if check.samples:
            print(f"  samples: {check.samples[:5]}")
    print(f"report: {report_path.relative_to(ROOT).as_posix()}")

    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
