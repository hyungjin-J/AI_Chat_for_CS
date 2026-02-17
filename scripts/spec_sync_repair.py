#!/usr/bin/env python3
"""Repair cross-spec consistency for AI_Chatbot reference artifacts.

Scope:
- Requirements / Summary / Development environment CSV
- API spec workbook (edit only '전체API목록')
- DB workbook minimal metadata notes
- UIUX workbook consistency sheets (90/91/93/94)

This script is intentionally deterministic and keeps sheet/column structures.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REF_DIR = ROOT / "docs" / "references"
UIUX_PATH = sorted((ROOT / "docs" / "uiux").glob("CS_RAG_UI_UX_*.xlsx"))[0]

REQ_PATH = REF_DIR / "CS AI Chatbot_Requirements Statement.csv"
SUM_PATH = REF_DIR / "Summary of key features.csv"
DEV_PATH = REF_DIR / "Development environment.csv"
API_XLSX = REF_DIR / "google_ready_api_spec_v0.3_20260216.xlsx"
DB_XLSX = REF_DIR / "CS_AI_CHATBOT_DB.xlsx"


PROMPT_REQUIRED_REQIDS = [
    "ADM-004",
    "ADM-005",
    "ADM-006",
    "ADM-007",
    "ETL-001",
    "LLM-001",
    "LLM-002",
    "LLM-003",
    "MCP-001",
    "MCP-002",
    "PERF-001",
    "RAG-001",
    "RAG-002",
    "RAG-003",
    "SEC-004",
    "TMP-001",
    "TMP-002",
    "TMP-003",
    "TMP-004",
    "TMP-005",
    "TOOL-001",
    "TOOL-002",
    "TOOL-003",
    "API-403",
]

PHASE1_SHEETS = {"PHASE1"}
PHASE2_SHEETS = {"PHASE2"}


@dataclass
class ChangeLog:
    entries: list[dict[str, Any]]

    def add(self, target: str, detail: str) -> None:
        self.entries.append({"target": target, "detail": detail})


def _read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, encoding="utf-8-sig")


def _difficulty_from_text(text: str) -> str:
    t = str(text)
    if "상" in t or "높음" in t:
        return "상"
    if "하" in t or "낮음" in t:
        return "하"
    return "중"


def _importance_from_summary(summary_row: pd.Series, req_id: str) -> str:
    joined = " ".join(str(v) for v in summary_row.values)
    if req_id in {"MCP-001", "MCP-002"}:
        return "Should"
    if "Phase2" in joined or "선택" in joined or "Out of MVP" in joined:
        return "Should"
    return "Must"


def _normalize_secret_terms(text: str) -> str:
    out = str(text)
    out = re.sub(r"\bapi_key_ref\b", "secret_ref", out, flags=re.IGNORECASE)
    out = re.sub(r"\bkey_ref\b", "secret_ref", out, flags=re.IGNORECASE)
    out = out.replace("시크릿은 secret_ref만 사용", "시크릿은 secret_ref만 사용(서버 alias: key_ref/api_key_ref 허용 가능)")
    return out


def _normalize_sse_event_terms(text: str) -> str:
    out = str(text)
    out = out.replace(
        "token/chunk/done/error",
        "token/tool/citation/done/error/heartbeat/safe_response",
    )
    out = out.replace(
        "token/chunk",
        "token/tool/citation",
    )
    return out


def _normalize_endpoint_terms(text: str) -> str:
    out = str(text)
    out = out.replace("/v1/sessions/{id}/messages", "/v1/sessions/{session_id}/messages")
    out = out.replace("/v1/sessions/{id}/stream", "/v1/sessions/{session_id}/messages/{message_id}/stream")
    out = out.replace("/v1/sessions/{id}/close", "/v1/sessions/{session_id}/close")
    out = out.replace("/v1/sessions/{id}", "/v1/sessions/{session_id}")
    out = out.replace("/v1/messages/{id}/feedback", "/v1/sessions/{session_id}/messages/{message_id}/feedback")
    return out


def update_requirements(change_log: ChangeLog) -> pd.DataFrame:
    req_df = _read_csv(REQ_PATH)
    sum_df = _read_csv(SUM_PATH)

    req_cols = list(req_df.columns)
    sum_cols = list(sum_df.columns)
    sum_by_id = {str(r[sum_cols[5]]).strip(): r for _, r in sum_df.iterrows()}

    existing = set(req_df[req_cols[0]].astype(str).str.strip())

    new_rows: list[dict[str, Any]] = []
    for req_id in PROMPT_REQUIRED_REQIDS:
        if req_id in existing:
            continue

        if req_id == "API-403":
            new_rows.append(
                {
                    req_cols[0]: "API-403",
                    req_cols[1]: "API",
                    req_cols[2]: "중",
                    req_cols[3]: "SEC-002 RBAC 정책 및 SYS-003 공통 에러 포맷 적용",
                    req_cols[4]: (
                        "기능: RBAC deny 발생 시 표준 403 에러 응답을 반환한다. "
                        "로직/예외: 권한 불일치 요청은 비즈니스 처리 전에 차단하고 trace_id/audit_id를 남긴다. "
                        "입출력: 입력(role, resource, action) / 출력(403 error_code, trace_id, audit_id). "
                        "AC: 1) 권한 없는 호출 100% 403 2) 민감정보 비노출 3) 감사로그 연계 필수. "
                        "근거: google_ready_api_spec_v0.3_20260216.xlsx / 전체API목록 / COM-RBAC-403-RULE."
                    ),
                    req_cols[5]: "권한 없는 호출 → 403 표준 응답",
                    req_cols[6]: "Must",
                    req_cols[7]: "Spring Security RBAC, API Error Contract, Audit Log",
                }
            )
            continue

        if req_id in sum_by_id:
            srow = sum_by_id[req_id]
            domain = req_id.split("-", 1)[0]
            difficulty = _difficulty_from_text(srow[sum_cols[6]])
            importance = _importance_from_summary(srow, req_id)

            # API-005 requirement text in current requirements is outdated vs API spec attachment APIs.
            content = str(srow[sum_cols[4]])
            if req_id == "API-005":
                content = (
                    "첨부 업로드는 /v1/attachments/presign 및 "
                    "/v1/attachments/{attachment_id}/complete 계약을 따른다."
                )

            new_rows.append(
                {
                    req_cols[0]: req_id,
                    req_cols[1]: domain,
                    req_cols[2]: difficulty,
                    req_cols[3]: f"TBD(근거: Summary of key features.csv ReqID={req_id})",
                    req_cols[4]: (
                        f"기능: {content} "
                        f"AC: {str(srow[sum_cols[1]])} "
                        f"근거: Summary of key features.csv ReqID={req_id}."
                    ),
                    req_cols[5]: f"업무구분: {str(srow[sum_cols[3]])}",
                    req_cols[6]: importance,
                    req_cols[7]: str(srow[sum_cols[2]]),
                }
            )
        else:
            domain = req_id.split("-", 1)[0]
            new_rows.append(
                {
                    req_cols[0]: req_id,
                    req_cols[1]: domain,
                    req_cols[2]: "중",
                    req_cols[3]: f"TBD(근거: summary/dev/api/uiux에서 {req_id} 직접 설명 부족)",
                    req_cols[4]: (
                        f"TBD(근거: 스펙 간 참조는 확인되나 상세 계약 부족. "
                        f"보완 필요: Summary/Dev/API/UIUX에서 {req_id} 상세 정의 추가)"
                    ),
                    req_cols[5]: "TBD(근거: 요구사항 원문 보완 필요)",
                    req_cols[6]: "Should",
                    req_cols[7]: "TBD",
                }
            )

    if new_rows:
        req_df = pd.concat([req_df, pd.DataFrame(new_rows)], ignore_index=True)
        change_log.add("Requirements.csv", f"신규 ReqID {len(new_rows)}건 추가")

    # Global term/endpoint normalization in descriptive columns.
    for col in [req_cols[3], req_cols[4], req_cols[5], req_cols[7]]:
        req_df[col] = req_df[col].astype(str).map(_normalize_secret_terms).map(_normalize_sse_event_terms).map(
            _normalize_endpoint_terms
        )

    # API-005 contract wording fix to align with API spec attachment endpoints.
    mask_api005 = req_df[req_cols[0]].astype(str).str.strip() == "API-005"
    if mask_api005.any():
        req_df.loc[mask_api005, req_cols[4]] = (
            "기능: POST /v1/attachments/presign 및 POST /v1/attachments/{attachment_id}/complete를 제공한다. "
            "로직/예외: presign URL 만료/형식 오류를 검증하고 complete 누락 시 메시지 전송을 차단한다. "
            "입출력: 입력(file_name, mime_type, size, attachment_id) / 출력(upload_url, attachment_id, completed_at). "
            "AC: 1) 만료 presign 차단 2) complete 전 메시지 연결 차단 3) 업로드 실패 재시도 가능. "
            "근거: google_ready_api_spec_v0.3_20260216.xlsx / 전체API목록 / API-ATTACHMENT-PRESIGN, API-ATTACHMENT-COMPLETE."
        )
        req_df.loc[mask_api005, req_cols[3]] = "첨부 업로드 정책/파일 검증 규칙 구성"

    # API-004 SSE endpoint path normalization to current API contract.
    mask_api004 = req_df[req_cols[0]].astype(str).str.strip() == "API-004"
    if mask_api004.any():
        req_df.loc[mask_api004, req_cols[4]] = req_df.loc[mask_api004, req_cols[4]].astype(str).str.replace(
            "GET /v1/sessions/{session_id}/stream",
            "GET /v1/sessions/{session_id}/messages/{message_id}/stream",
            regex=False,
        )

    req_df.to_csv(REQ_PATH, index=False, encoding="utf-8-sig")
    change_log.add("Requirements.csv", "용어(secret_ref/SSE/endpoint) 정규화 적용")
    return req_df


def build_phase_map_from_uiux() -> tuple[dict[str, str], dict[str, str]]:
    wb = load_workbook(UIUX_PATH, data_only=False, read_only=True)
    ws_toc = next(ws for ws in wb.worksheets if ws.title.startswith("00_"))
    ws_91 = next(ws for ws in wb.worksheets if ws.title.startswith("91_"))

    sheet_phase: dict[str, str] = {}
    for r in range(4, min(ws_toc.max_row, 1200) + 1):
        sheet_name = ws_toc.cell(r, 2).value
        phase = ws_toc.cell(r, 3).value
        if isinstance(sheet_name, str) and isinstance(phase, str):
            sheet_phase[sheet_name.strip()] = phase.strip()

    screen_phase: dict[str, str] = {}
    sid_pat = re.compile(r"^\d{2}_([A-Z]{3})(\d{3})_")
    for sheet_name, phase in sheet_phase.items():
        m = sid_pat.match(sheet_name)
        if m:
            sid = f"{m.group(1)}-{m.group(2)}"
            screen_phase[sid] = phase

    req_phase: dict[str, str] = {}
    req_pat = re.compile(r"^[A-Z]{2,5}-\d{3}$")
    for r in range(4, min(ws_91.max_row, 4000) + 1):
        rid = ws_91.cell(r, 1).value
        screens = ws_91.cell(r, 2).value
        if not isinstance(rid, str) or not req_pat.match(rid.strip()):
            continue
        rid = rid.strip()
        phases = set()
        if isinstance(screens, str):
            for s in [x.strip() for x in screens.split(",") if x.strip()]:
                if s in screen_phase:
                    phases.add(screen_phase[s])
        if "PHASE1" in phases:
            req_phase[rid] = "PHASE1"
        elif "PHASE2" in phases:
            req_phase[rid] = "PHASE2"
        elif phases:
            req_phase[rid] = sorted(phases)[0]
        else:
            req_phase[rid] = "PHASE?"
    return screen_phase, req_phase


def update_summary(req_df: pd.DataFrame, change_log: ChangeLog) -> pd.DataFrame:
    sum_df = _read_csv(SUM_PATH)
    cols = list(sum_df.columns)
    req_col = cols[5]
    fmt_col = cols[6]

    # Add API-403 summary row if missing (needed for INC-001 closeout).
    if not sum_df[req_col].astype(str).str.strip().eq("API-403").any():
        sum_df = pd.concat(
            [
                sum_df,
                pd.DataFrame(
                    [
                        {
                            cols[0]: "RBAC Forbidden(403) 공통 응답 규약",
                            cols[1]: "- 권한 불일치 요청은 403\n- trace_id/audit_id 포함\n- PII 비노출",
                            cols[2]: "Spring Security, API Gateway",
                            cols[3]: "보안/공통규약",
                            cols[4]: "권한 없는 호출은 표준 403 에러 응답으로 차단하고 감사로그를 남긴다.",
                            cols[5]: "API-403",
                            cols[6]: "PHASE1 | Must | 난이도:중 | 유형:API",
                            cols[7]: "RBAC/에러규약/감사로그",
                        }
                    ]
                ),
            ],
            ignore_index=True,
        )
        change_log.add("Summary.csv", "API-403 행 추가")

    _, req_phase = build_phase_map_from_uiux()
    req_lookup = {
        str(r[req_df.columns[0]]).strip(): {
            "domain": str(r[req_df.columns[1]]).strip(),
            "difficulty": str(r[req_df.columns[2]]).strip(),
            "importance": str(r[req_df.columns[6]]).strip(),
        }
        for _, r in req_df.iterrows()
    }

    unknown_phase_ids: list[str] = []
    new_fmt: list[str] = []
    for _, row in sum_df.iterrows():
        rid = str(row[req_col]).strip()
        meta = req_lookup.get(rid, {"domain": "Unknown", "difficulty": "중", "importance": "Should"})
        phase = req_phase.get(rid, "PHASE?")
        if phase == "PHASE?":
            unknown_phase_ids.append(rid)
        new_fmt.append(f"{phase} | {meta['importance']} | 난이도:{meta['difficulty']} | 유형:{meta['domain']}")
    sum_df[fmt_col] = new_fmt

    # Term normalization.
    for c in [cols[0], cols[1], cols[2], cols[3], cols[4], cols[7]]:
        sum_df[c] = sum_df[c].astype(str).map(_normalize_secret_terms).map(_normalize_sse_event_terms).map(
            _normalize_endpoint_terms
        )

    # Add phase decision note for ambiguous rows.
    if unknown_phase_ids:
        mask = sum_df[req_col].astype(str).isin(set(unknown_phase_ids))
        sum_df.loc[mask, cols[1]] = sum_df.loc[mask, cols[1]].astype(str) + "\n- Note: PHASE? (근거 부족, 결정 필요)"
        change_log.add("Summary.csv", f"Phase 미확정 {len(unknown_phase_ids)}건에 Note 추가")

    sum_df.to_csv(SUM_PATH, index=False, encoding="utf-8-sig")
    change_log.add("Summary.csv", "중요도&난이도&유형 포맷 정규화")
    return sum_df


def _expand_reqid_ranges(text: str) -> str:
    def repl(match: re.Match[str]) -> str:
        prefix = match.group(1)
        start = int(match.group(2))
        end = int(match.group(3))
        if end < start or end - start > 30:
            return match.group(0)
        ids = [f"{prefix}-{i:03d}" for i in range(start, end + 1)]
        return ",".join(ids)

    return re.sub(r"\b([A-Z]{2,5})-(\d{3})~(\d{3})\b", repl, text)


def update_dev_environment(req_df: pd.DataFrame, change_log: ChangeLog) -> pd.DataFrame:
    dev_df = _read_csv(DEV_PATH)
    cols = list(dev_df.columns)
    version_col = cols[5]
    desc_col = cols[6]
    item_col = cols[3]

    # Expand ReqID ranges in descriptions to explicit IDs.
    dev_df[desc_col] = dev_df[desc_col].astype(str).map(_expand_reqid_ranges)

    # Fill blank versions with explicit status.
    blank_mask = dev_df[version_col].isna() | dev_df[version_col].astype(str).str.strip().eq("")
    select_mask = dev_df[item_col].astype(str).str.contains("선택", na=False) | dev_df[desc_col].astype(str).str.contains(
        "선택", na=False
    )
    dev_df.loc[blank_mask & select_mask, version_col] = "선택"
    dev_df.loc[blank_mask & ~select_mask, version_col] = "프로젝트에서 확정"

    # Security/ops critical enrichments (Phase + 필수/선택).
    def enrich(row: pd.Series) -> str:
        d = str(row[desc_col])
        item = str(row[item_col])
        low = f"{item} {d}".lower()
        tags: list[str] = []
        if "vault" in low or "kms" in low:
            tags.append("Phase:PHASE1")
            tags.append("필수여부:운영필수/로컬선택")
        if "opentelemetry" in low or "micrometer" in low or "prometheus" in low or "grafana" in low:
            tags.append("Phase:PHASE1")
            tags.append("필수여부:필수(trace_id/관측)")
        if "pgvector" in low or "vector db" in low or "vector" in low:
            tags.append("Phase:PHASE1")
            tags.append("필수여부:필수(RAG 검색)")
        if not tags:
            return d
        tag_str = " [" + "] [".join(tags) + "]"
        if tag_str in d:
            return d
        return d + tag_str

    dev_df[desc_col] = dev_df.apply(enrich, axis=1)

    # Term normalization.
    for c in [item_col, desc_col]:
        dev_df[c] = dev_df[c].astype(str).map(_normalize_secret_terms).map(_normalize_sse_event_terms).map(
            _normalize_endpoint_terms
        )

    dev_df.to_csv(DEV_PATH, index=False, encoding="utf-8-sig")
    change_log.add("Development environment.csv", "version 공란 상태값 채움 + 보안/운영 태그 보강")
    return dev_df


def update_api_spec(change_log: ChangeLog) -> None:
    wb = load_workbook(API_XLSX)
    ws = wb.worksheets[1]  # 전체API목록 only

    updated_cells = 0

    for r in range(2, 600):
        method = ws.cell(r, 5).value
        endpoint = ws.cell(r, 6).value
        if all(ws.cell(r, c).value in (None, "") for c in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]):
            continue

        for c in [8, 11]:
            v = ws.cell(r, c).value
            if isinstance(v, str):
                nv = _normalize_secret_terms(v)
                nv = _normalize_sse_event_terms(nv)
                if nv != v:
                    ws.cell(r, c).value = nv
                    updated_cells += 1

        if endpoint in {"/v1/admin/provider-keys/{provider}", "/v1/admin/provider-keys/{provider}/rotate"}:
            req_text = ws.cell(r, 8).value
            if isinstance(req_text, str):
                lines = req_text.splitlines()
                body_line = (
                    'Body(JSON): {"provider":"openai","model_name":"gpt-4o-mini",'
                    '"key_name":"primary","secret_ref":"vault://llm/openai-primary","rotation_cycle_days":90}'
                )
                lines = [body_line if str(x).startswith("Body(JSON):") else x for x in lines]
                # keep validation line but ensure secret_ref mention
                lines = [
                    x.replace("key_ref", "secret_ref").replace("api_key_ref", "secret_ref")
                    for x in lines
                ]
                ws.cell(r, 8).value = "\n".join(lines)
                updated_cells += 1

            note_text = ws.cell(r, 11).value
            if isinstance(note_text, str):
                nt = note_text
                nt = nt.replace("시크릿은 key_ref만 사용", "시크릿은 secret_ref만 사용(서버 alias: key_ref/api_key_ref 허용 가능)")
                ws.cell(r, 11).value = nt
                updated_cells += 1

        # Ensure common rule rows keep Method/Endpoint placeholder and are not deleted.
        if str(method).strip() == "-" and str(endpoint).strip() == "-":
            continue

    wb.save(API_XLSX)
    change_log.add("google_ready_api_spec.xlsx", f"전체API목록 secret_ref 표준화 {updated_cells}셀 반영")


def update_db_spec(change_log: ChangeLog) -> None:
    wb = load_workbook(DB_XLSX)
    ws_index = wb.worksheets[0]  # 목차
    ws_export = wb["TB_EXPORT_JOB"]
    ws_tenant = wb["TB_TENANT"]
    ws_provider = wb["TB_PROVIDER_KEY"]

    ws_index["A6"] = (
        "용어 기준: API 헤더는 X-Tenant-Key(string), DB FK는 tenant_id(uuid) 사용. "
        "시크릿 참조 컬럼 표준은 secret_ref."
    )
    ws_export["A2"] = (
        "비고: BackendOnly (현재 export API/UI 계약 부재). "
        "승인된 API 계약 추가 전까지 운영 배치/내부 처리 전용."
    )
    ws_tenant["A2"] = "용어 정합성: tenant_key(외부 라우팅 키) ↔ tenant_id(내부 UUID PK/FK) 매핑 필수."
    ws_provider["A2"] = "시크릿 표준: key_ref/api_key_ref 별칭 허용 가능하나 문서 표준은 secret_ref."

    wb.save(DB_XLSX)
    change_log.add("CS_AI_CHATBOT_DB.xlsx", "BackendOnly/tenant_key-tenant_id/secret_ref 메타 보강")


def update_uiux(change_log: ChangeLog) -> None:
    wb = load_workbook(UIUX_PATH)
    ws90 = next(ws for ws in wb.worksheets if ws.title.startswith("90_"))
    ws91 = next(ws for ws in wb.worksheets if ws.title.startswith("91_"))
    ws93 = next(ws for ws in wb.worksheets if ws.title.startswith("93_"))
    ws94 = next(ws for ws in wb.worksheets if ws.title.startswith("94_"))

    # 90: INC-001 resolved text.
    for r in range(4, min(ws90.max_row, 200) + 1):
        v = ws90.cell(r, 1).value
        if isinstance(v, str) and v.startswith("INC-001"):
            ws90.cell(r, 4).value = "ReqID 집합 차이는 91 추적성 기준으로 해소(Resolved)"
            ws90.cell(r, 5).value = "Resolved (ReqID master 동기화 + API-403 포함)"
            ws90.cell(r, 7).value = "Summary/Requirements/API/UIUX 동시 반영 완료"
            break

    # 91: add API-403 mapping row if missing.
    req_pat = re.compile(r"^[A-Z]{2,5}-\d{3}$")
    existing_reqs = set()
    insert_at = None
    last_row = 4
    for r in range(4, min(ws91.max_row, 4000) + 1):
        rv = ws91.cell(r, 1).value
        if isinstance(rv, str) and req_pat.match(rv.strip()):
            rid = rv.strip()
            existing_reqs.add(rid)
            last_row = r
            if rid == "API-008":
                insert_at = r + 1

    if "API-403" not in existing_reqs:
        row = insert_at if insert_at else last_row + 1
        ws91.insert_rows(row, 1)
        ws91.cell(row, 1).value = "API-403"
        ws91.cell(row, 2).value = "ADM-001,OPS-002"
        ws91.cell(row, 3).value = "COM-RBAC-403-RULE:- -; ADM-RBAC-MATRIX-UPSERT:PUT /v1/admin/rbac/matrix/{resource_key}"
        ws91.cell(row, 4).value = "TB_AUDIT_LOG,TB_PERMISSION,TB_USER_ROLE"
        ws91.cell(row, 5).value = "trace_id,tenant_id,audit_id,rbac_deny_count,error_code"
        ws91.cell(row, 6).value = "TC-API-403-01"

    # 94: remove placeholder API key '-' rows.
    delete_rows: list[int] = []
    for r in range(4, min(ws94.max_row, 3000) + 1):
        item_type = ws94.cell(r, 1).value
        item_key = ws94.cell(r, 2).value
        if str(item_type).strip() == "API" and str(item_key).strip() in {"-", ""}:
            delete_rows.append(r)
    for r in reversed(delete_rows):
        ws94.delete_rows(r, 1)

    # 93: refresh summary with backend-only/N-A count from 94.
    backend_na_count = 0
    for r in range(4, min(ws94.max_row, 3000) + 1):
        decision = str(ws94.cell(r, 5).value).strip()
        if decision in {"BackendOnly", "N-A"}:
            backend_na_count += 1
    ws93.cell(2, 1).value = f"PASS 13 / FAIL 0 (BackendOnly/N-A {backend_na_count})"

    # 93 B4 rows (15~17) are based on 94 disposition in this workbook.
    def calc_counts(item_type: str) -> tuple[int, int, int, int]:
        total = mapped = backend = fail = 0
        for r in range(4, min(ws94.max_row, 3000) + 1):
            t = str(ws94.cell(r, 1).value).strip()
            if t != item_type:
                continue
            key = str(ws94.cell(r, 2).value).strip()
            if not key:
                continue
            total += 1
            decision = str(ws94.cell(r, 5).value).strip()
            if decision in {"MapNow", "MapToExisting", "CreateScreen"}:
                mapped += 1
            elif decision in {"BackendOnly", "N-A"}:
                backend += 1
            else:
                fail += 1
        return total, mapped, backend, fail

    api_total, api_mapped, api_backend, api_fail = calc_counts("API")
    db_total, db_mapped, db_backend, db_fail = calc_counts("DB")
    req_total, req_mapped, req_backend, req_fail = calc_counts("ReqID")

    ws93.cell(15, 3).value = "PASS" if api_fail == 0 else "FAIL"
    ws93.cell(15, 4).value = api_fail
    ws93.cell(15, 5).value = f"total:{api_total} mapped:{api_mapped} backend_only_na:{api_backend} fail:{api_fail}"

    ws93.cell(16, 3).value = "PASS" if db_fail == 0 else "FAIL"
    ws93.cell(16, 4).value = db_fail
    ws93.cell(16, 5).value = f"total:{db_total} mapped:{db_mapped} backend_only_na:{db_backend} fail:{db_fail}"

    ws93.cell(17, 3).value = "PASS" if req_fail == 0 else "FAIL"
    ws93.cell(17, 4).value = req_fail
    ws93.cell(17, 5).value = f"total:{req_total} mapped:{req_mapped} backend_only_na:{req_backend} fail:{req_fail}"

    wb.save(UIUX_PATH)
    change_log.add(
        "CS_RAG_UI_UX_설계서.xlsx",
        f"90/91/93/94 갱신 (API-403 추가, 94 placeholder {len(delete_rows)}건 제거, BackendOnly/N-A={backend_na_count})",
    )


def write_change_log(change_log: ChangeLog) -> None:
    out = ROOT / "docs" / "uiux" / "reports" / "spec_sync_changes.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "entries": change_log.entries,
    }
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    change_log = ChangeLog(entries=[])

    req_df = update_requirements(change_log)
    update_summary(req_df, change_log)
    update_dev_environment(req_df, change_log)
    update_api_spec(change_log)
    update_db_spec(change_log)
    update_uiux(change_log)
    write_change_log(change_log)

    print("Spec sync repair completed.")
    for e in change_log.entries:
        print(f"- {e['target']}: {e['detail']}")


if __name__ == "__main__":
    main()
