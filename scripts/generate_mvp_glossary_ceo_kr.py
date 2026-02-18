from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT = Path("docs/review/mvp_verification_pack/07_MVP_GLOSSARY_CEO_KR.xlsx")


def base_style():
    colors = {
        "bg": "F4F7FB",
        "title": "1F3A5F",
        "title_text": "FFFFFF",
        "header": "DDEAF8",
        "header_text": "1B2A41",
        "row_even": "F9FCFF",
        "row_odd": "FFFFFF",
        "priority": "EAF6EE",
        "line": "D8E1EB",
        "sub": "EEF4FA",
    }
    thin = Side(style="thin", color=colors["line"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return colors, border


def fill_background(ws, rows: int, cols: int, color: str) -> None:
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=color)


def create_glossary(ws, colors: dict, border: Border) -> None:
    ws.title = "용어사전"
    fill_background(ws, 500, 12, colors["bg"])

    ws.merge_cells("A1:J1")
    ws["A1"] = "CS 서포팅 AI 챗봇 MVP 용어사전 (CEO/비개발자 설명용)"
    ws["A1"].font = Font(name="Malgun Gothic", size=18, bold=True, color=colors["title_text"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor=colors["title"])

    ws.merge_cells("A2:J2")
    ws["A2"] = (
        "목적: 기술 용어를 쉬운 한국어로 이해하고, 경영진이 리스크/통제 포인트를 즉시 판단하도록 지원 "
        f"(생성 시각: {datetime.now().strftime('%Y-%m-%d %H:%M')})"
    )
    ws["A2"].font = Font(name="Malgun Gothic", size=11, color="1B2A41")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["A2"].fill = PatternFill("solid", fgColor=colors["sub"])

    headers = [
        "구분",
        "용어(원문)",
        "쉬운 한국어 설명",
        "프로젝트에서 주로 쓰는 곳",
        "왜 중요한가(경영 관점)",
        "현장 예시",
        "주의 포인트",
        "관련 에러코드/지표",
        "빠른 확인 방법",
        "우선순위",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(4, i, h)
        c.font = Font(name="Malgun Gothic", size=11, bold=True, color=colors["header_text"])
        c.fill = PatternFill("solid", fgColor=colors["header"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    terms = [
        ("보안", "PII", "개인을 식별할 수 있는 정보(전화번호/이메일/주소/주문번호)", "입력 처리, 로그, DB, 인용문", "유출 시 법적 책임과 브랜드 신뢰 손상이 큼", "010-1234-5678 → 010-****-****로 저장", "화면에서 가려도 서버 원문 저장되면 실패", "SEC-001-401, RAG-002-422-POLICY", "샘플 3개 입력 후 로그/응답 원문 검색", "최상"),
        ("스트리밍", "SSE", "서버가 결과를 실시간으로 순서대로 보내는 방식", "답변 스트림, 하트비트, 재연결", "상담원이 기다리지 않고 초안 확인 가능", "token → citation → done 순서", "끊김 복구(Last-Event-ID) 없으면 현장 불편", "API-008-429-SSE, SYS-003-503", "중간에 네트워크 끊고 /resume 확인", "최상"),
        ("테스트", "E2E", "로그인부터 답변 완료까지 실제 사용자 흐름을 통째로 검증", "API 연동, 권한, 스트림 전체", "단위 테스트 통과만으로는 실사용 품질 보장 불가", "curl로 헤더 포함 전체 시나리오 실행", "중간 모킹만으로 PASS 처리 금지", "API-003-409, API-003-422", "실제 서버로 1회 완주 테스트", "최상"),
        ("품질계약", "Answer Contract", "모델 출력이 반드시 지켜야 하는 JSON 규격", "스키마 검증, 인용/근거 점수 검증", "근거 없는 답변을 구조적으로 차단", "인용 0개면 즉시 안전응답", "자연스러워 보여도 계약 미통과면 폐기", "AI-009-422-SCHEMA, AI-009-409-CITATION", "유효/무효 JSON 각각 주입", "최상"),
        ("안전정책", "Fail-Closed", "검증 실패 시 자유문장 대신 안전응답으로 종료", "최종 응답 게이트", "잘못된 답변 노출을 원천 차단", "근거 부족 시 안내문만 출력", "검증 전 토큰이 먼저 나가면 치명적", "AI-009-200-SAFE", "실패 시 token 0건 여부 확인", "최상"),
        ("근거", "Citation", "답변의 출처 문서 조각 정보", "근거 패널, TB_RAG_CITATION", "답변 신뢰도와 감사 가능성 확보", "chunk_id/rank/excerpt_masked 저장", "원문 개인정보 노출 금지", "AI-009-409-CITATION", "모든 답변에 인용 최소 1개", "상"),
        ("근거강도", "Evidence Score", "답변 근거가 충분한지 수치로 보는 점수", "evidence.score vs threshold", "신뢰도 낮은 답변 자동 차단", "0.62 < 0.70이면 차단", "항상 0.99만 나오면 지표 신뢰성 의심", "AI-009-409-EVIDENCE", "낮은 관련도 질문으로 차단 확인", "상"),
        ("추적", "trace_id", "요청 전체를 끝까지 연결하는 공통 식별자", "요청/로그/DB/SSE 전 구간", "장애 재현, 감사, SLA 분석의 핵심", "응답·로그·이벤트가 같은 trace_id", "중간 계층에서 새로 만들면 추적 단절", "SYS-004-409-TRACE", "단일 요청의 trace_id 일치 검증", "최상"),
        ("격리", "Tenant Isolation", "고객사 데이터가 서로 섞이지 않도록 분리", "X-Tenant-Key, DB 조건", "교차 유출 방지(계약/법무 핵심)", "A 테넌트 토큰으로 B 세션 조회 시 403", "UI 숨김은 보안이 아님", "SEC-002-403, SYS-002-403", "교차 조회 시도 후 차단 확인", "최상"),
        ("권한", "RBAC", "역할(AGENT/ADMIN 등) 기반 접근 통제", "서버 권한 검사", "내부 오남용/오작동 위험 축소", "권한 없는 호출은 403", "프론트 버튼 숨김만으로 완료 아님", "SEC-002-403", "권한 없는 JWT로 호출", "최상"),
        ("중복방지", "Idempotency", "같은 요청을 여러 번 보내도 1번만 처리", "POST /messages", "재시도 시 중복 생성/중복 비용 방지", "동일 키 재요청 시 409 또는 기존값", "멀티노드 저장전략 없으면 중복 위험", "API-003-409", "같은 키 2회 호출 비교", "상"),
        ("예산통제", "Budget Guard", "토큰/도구호출/top_k/동시연결 상한 보호장치", "요청 전 가드 검사", "비용 폭증과 악용 트래픽 차단", "한도 초과 시 429 + Retry-After", "클라이언트 단 제한은 우회 가능", "API-008-429-BUDGET, API-008-429-SSE", "의도적으로 초과 요청", "최상"),
        ("검색", "RAG", "문서 근거를 찾고 그 근거 기반으로 답변 생성", "RetrievalService", "환각 답변 감소", "top_k 근거를 찾아 모델에 전달", "검색 실패 시 안전응답 전환 필수", "RAG-002-422-POLICY", "무근거 질문으로 동작 확인", "상"),
        ("데이터", "PostgreSQL", "운영 데이터의 기본 저장소", "세션/메시지/인용/스트림 이벤트", "감사·복구·추적의 기반", "TB_MESSAGE, TB_RAG_CITATION 저장", "스키마 불일치 시 런타임 장애", "SYS-003-500", "스펙과 실제 컬럼 대조", "상"),
        ("캐시", "Redis", "빠른 임시 저장소(카운터/세션/락)", "레이트리밋, 동시성, 세션 보조", "성능 향상과 부하 완화", "429 카운터 관리", "tenant prefix 누락 시 격리 실패", "API-008-429-*", "키 패턴에 tenant 포함 확인", "중"),
        ("관측", "OpenTelemetry", "로그/메트릭/트레이스를 표준 방식으로 수집", "지표 대시보드, 장애 분석", "문제 탐지 시간 단축", "first-token 지연 P95 추적", "trace_id 없는 로그는 분석 가치 낮음", "운영지표", "대시보드 KPI 점검", "중"),
        ("모델", "Ollama/외부 LLM", "답변을 생성하는 모델 실행 주체", "LlmClient 구현", "비용/속도/품질 전략 선택", "개발: Ollama, 운영: 정책에 따라 선택", "모델 출력을 신뢰하지 말고 검증 우선", "AI-009 계열", "고의로 잘못된 JSON 유도", "상"),
        ("오류표준", "COM-ERROR-FORMAT", "모든 오류를 같은 JSON 형태로 반환하는 규칙", "error_code/message/trace_id/details", "운영 커뮤니케이션 비용 절감", "모든 4xx/5xx 형식 일치", "엔드포인트별 형식 분산 금지", "공통", "오류 API 3개 비교", "상"),
        ("안전응답", "safe_response", "검증 실패 시 제공되는 제한된 안전 문구", "SSE safe_response 이벤트", "위험정보 노출 방지", "근거 부족으로 현재 답변 제한 안내", "safe_response 전에 token 누출 금지", "AI-009-200-SAFE", "실패 로그에서 token 0건 확인", "최상"),
        ("복원력", "Resilience", "외부 장애가 전체 장애로 번지는 것을 막는 설계", "timeout/retry/circuit breaker", "가용성 유지", "외부 지연 시 제한된 재시도 후 안전종료", "무한 재시도 금지", "SYS-003-503", "지연 주입 테스트", "중"),
    ]

    row = 5
    for i, item in enumerate(terms):
        fill = colors["row_even"] if i % 2 == 0 else colors["row_odd"]
        for c, value in enumerate(item, 1):
            cell = ws.cell(row, c, value)
            cell.font = Font(name="Malgun Gothic", size=10, color="25364A")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = border
        if ws.cell(row, 10).value == "최상":
            for c in range(1, 11):
                ws.cell(row, c).fill = PatternFill("solid", fgColor=colors["priority"])
        ws.cell(row, 10).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 68
        row += 1

    widths = {1: 12, 2: 20, 3: 32, 4: 29, 5: 31, 6: 28, 7: 28, 8: 24, 9: 28, 10: 10}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[4].height = 42
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:J{row - 1}"


def create_flow_sheet(wb: Workbook, colors: dict, border: Border) -> None:
    ws = wb.create_sheet("흐름설명")
    fill_background(ws, 220, 8, colors["bg"])

    ws.merge_cells("A1:G1")
    ws["A1"] = "사용자 관점 MVP 흐름(로그인 → 질문 → 근거 → 답변/차단)"
    ws["A1"].font = Font(name="Malgun Gothic", size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="2B4D6E")

    headers = ["단계", "사용자 화면", "시스템 내부 동작", "실패 시 보이는 것", "핵심 통제", "확인 질문", "증빙 위치"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(3, i, h)
        c.font = Font(name="Malgun Gothic", size=11, bold=True, color="1B2A41")
        c.fill = PatternFill("solid", fgColor=colors["header"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    rows = [
        ("1. 로그인", "로그인 성공/실패 메시지", "JWT 발급 + 역할 확인", "401 또는 403 오류코드", "RBAC", "권한 없는 계정은 차단되는가?", "artifacts/rbac_401_403_checks.txt"),
        ("2. 세션 생성", "새 대화방 생성", "tenant_key 매핑 + trace_id 생성", "403(테넌트 불일치)", "Tenant Filter", "다른 테넌트 세션 접근이 막히는가?", "05_E2E_EVIDENCE.md"),
        ("3. 질문 전송", "전송 중 표시", "PII 마스킹 후 검색/모델 입력", "422 정책 오류 또는 safe_response", "PiiMaskingFilter", "개인정보 원문이 남지 않는가?", "artifacts/pii_masking_checks.txt"),
        ("4. 근거 검색", "근거 패널 로딩", "top_k 검색 + 근거 저장", "근거 부족 시 차단", "Evidence Threshold", "근거 없는 답변이 노출되지 않는가?", "artifacts/sse_stream_fail_closed.log"),
        ("5. 생성 + 검증", "실시간 출력 또는 안전응답", "Answer Contract 검증", "AI-009 계열 코드", "Fail-Closed", "검증 전 token 유출이 없는가?", "artifacts/sse_stream_fail_closed.log"),
        ("6. 스트리밍 완료", "done 이벤트 수신", "event 저장 + resume 지원", "연결 오류 시 재연결", "SSE Resume", "중단 후 이어받기가 되는가?", "05_E2E_EVIDENCE.md"),
    ]

    r = 4
    for i, row in enumerate(rows):
        fill = colors["row_even"] if i % 2 == 0 else colors["row_odd"]
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name="Malgun Gothic", size=10, color="25364A")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = border
        ws.row_dimensions[r].height = 60
        r += 1

    widths = {1: 14, 2: 24, 3: 28, 4: 23, 5: 18, 6: 28, 7: 40}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 38
    ws.freeze_panes = "A4"


def create_error_sheet(wb: Workbook, colors: dict, border: Border) -> None:
    ws = wb.create_sheet("에러코드해설")
    fill_background(ws, 250, 7, colors["bg"])

    ws.merge_cells("A1:F1")
    ws["A1"] = "주요 에러코드 한글 해설(현업/경영진 공용)"
    ws["A1"].font = Font(name="Malgun Gothic", size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="355E3B")

    headers = ["에러코드", "발생 상황", "쉽게 말하면", "즉시 조치", "재발 방지 포인트", "중요도"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(3, i, h)
        c.font = Font(name="Malgun Gothic", size=11, bold=True, color="1B2A41")
        c.fill = PatternFill("solid", fgColor=colors["header"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    rows = [
        ("SEC-001-401", "인증 토큰 없음/만료", "로그인이 필요합니다", "재로그인 후 재시도", "토큰 만료·갱신 정책 점검", "높음"),
        ("SEC-002-403", "권한 없는 역할 호출", "이 기능에 접근할 권한이 없습니다", "계정 역할 확인", "서버 RBAC 강제 유지", "높음"),
        ("SYS-002-403", "유효하지 않은 테넌트 키", "해당 고객사 영역 접근이 차단되었습니다", "헤더/매핑 확인", "테넌트 키 발급·회수 절차 표준화", "높음"),
        ("API-003-409", "중복 요청", "같은 요청이 이미 처리되었습니다", "중복 전송 확인", "Idempotency 전략 점검", "중간"),
        ("API-008-429-BUDGET", "예산/한도 초과", "사용량 제한을 넘었습니다", "잠시 후 재시도 안내", "테넌트 한도 재조정", "높음"),
        ("API-008-429-SSE", "동시 스트림 초과", "동시 연결이 많아 대기 필요", "기존 연결 정리", "연결 누수 모니터링", "중간"),
        ("AI-009-422-SCHEMA", "답변 형식 검증 실패", "모델 답변 형식이 규격과 다릅니다", "안전응답 전환 확인", "프롬프트/리페어 점검", "높음"),
        ("AI-009-409-CITATION", "인용 누락", "근거가 없어 답변을 제한했습니다", "근거 데이터 점검", "검색/인용 매핑 개선", "높음"),
        ("AI-009-409-EVIDENCE", "근거 점수 미달", "신뢰도가 낮아 답변을 제한했습니다", "질문 재구성 안내", "evidence threshold 점검", "높음"),
        ("AI-009-200-SAFE", "안전모드 전환", "안전한 안내문으로 대신 응답했습니다", "정상 보호 동작 안내", "차단 비율 모니터링", "높음"),
        ("SYS-004-409-TRACE", "trace_id 누락/불일치", "요청 추적이 끊겼습니다", "전파 경로 점검", "필터 강제 주입", "높음"),
        ("SYS-003-500", "내부 예외", "일시적 시스템 오류입니다", "로그 확인/재시도", "예외 분류와 재발 방지", "높음"),
        ("SYS-003-503", "외부 의존성 장애", "잠시 후 다시 시도해 주세요", "복구 후 재시도", "타임아웃/서킷 조정", "중간"),
    ]

    r = 4
    for i, row in enumerate(rows):
        fill = "FFFDF8" if row[-1] == "높음" else (colors["row_even"] if i % 2 == 0 else colors["row_odd"])
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name="Malgun Gothic", size=10, color="25364A")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = border
        ws.cell(r, 6).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 46
        r += 1

    widths = {1: 20, 2: 22, 3: 26, 4: 22, 5: 26, 6: 10}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 36
    ws.freeze_panes = "A4"


def create_checklist_sheet(wb: Workbook, colors: dict, border: Border) -> None:
    ws = wb.create_sheet("10분점검표")
    fill_background(ws, 220, 6, colors["bg"])

    ws.merge_cells("A1:E1")
    ws["A1"] = "데모 전 10분 점검표 (CEO 보고용)"
    ws["A1"].font = Font(name="Malgun Gothic", size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="5E548E")

    headers = ["체크 항목", "확인 질문", "성공 기준", "증빙 파일", "체크"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(3, i, h)
        c.font = Font(name="Malgun Gothic", size=11, bold=True, color="1B2A41")
        c.fill = PatternFill("solid", fgColor=colors["header"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    rows = [
        ("Fail-Closed", "검증 실패 시 안전응답만 나오는가?", "safe_response 후 done, token 누출 0건", "artifacts/sse_stream_fail_closed.log", "□"),
        ("PII 마스킹", "전화/이메일/주문번호 원문이 남는가?", "입력/로그/응답 모두 마스킹", "artifacts/pii_masking_checks.txt", "□"),
        ("trace_id 일관성", "같은 trace_id로 끝까지 추적되는가?", "응답/DB/SSE에서 동일 ID 확인", "artifacts/trace_id_checks.txt", "□"),
        ("테넌트 격리", "다른 고객사 데이터 접근이 차단되는가?", "교차 조회 시 403", "05_E2E_EVIDENCE.md", "□"),
        ("예산/동시성 제한", "한도 초과 시 429가 일관되게 반환되는가?", "error_code + Retry-After 일치", "artifacts/budget_429_checks.txt", "□"),
        ("권한 통제", "권한 없는 호출이 서버에서 차단되는가?", "401/403 규약 준수", "artifacts/rbac_401_403_checks.txt", "□"),
    ]

    r = 4
    for i, row in enumerate(rows):
        fill = colors["row_even"] if i % 2 == 0 else colors["row_odd"]
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name="Malgun Gothic", size=10, color="25364A")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = border
        ws.cell(r, 5).font = Font(name="Malgun Gothic", size=12, bold=True, color="34495E")
        ws.cell(r, 5).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 42
        r += 1

    widths = {1: 20, 2: 35, 3: 30, 4: 50, 5: 8}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 36
    ws.freeze_panes = "A4"


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    colors, border = base_style()

    create_glossary(wb.active, colors, border)
    create_flow_sheet(wb, colors, border)
    create_error_sheet(wb, colors, border)
    create_checklist_sheet(wb, colors, border)

    wb.save(OUTPUT)
    print(f"created: {OUTPUT}")


if __name__ == "__main__":
    main()
