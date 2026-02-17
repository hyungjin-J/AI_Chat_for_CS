#!/usr/bin/env python3
"""Validate UI/UX traceability across requirements, API, and DB inventories."""

from __future__ import annotations

import csv
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REQ_CSV = ROOT / 'docs' / 'references' / 'CS AI Chatbot_Requirements Statement.csv'
FEAT_CSV = ROOT / 'docs' / 'references' / 'Summary of key features.csv'
API_XLSX = ROOT / 'docs' / 'references' / 'google_ready_api_spec_v0.3_20260216.xlsx'
DB_XLSX = ROOT / 'docs' / 'references' / 'CS_AI_CHATBOT_DB.xlsx'
SPEC_MD = ROOT / 'docs' / 'uiux' / 'UIUX_Spec.md'
REPORT_JSON = ROOT / 'docs' / 'uiux' / 'reports' / 'trace_report.json'


def read_req_ids() -> set[str]:
    ids: set[str] = set()
    with REQ_CSV.open('r', encoding='utf-8-sig', newline='') as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if row and row[0].strip():
                ids.add(row[0].strip())
    return ids


def read_feature_ids() -> set[str]:
    ids: set[str] = set()
    with FEAT_CSV.open('r', encoding='utf-8-sig', newline='') as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if len(row) >= 6 and row[5].strip():
                ids.add(row[5].strip())
    return ids


def read_endpoints() -> set[str]:
    wb = load_workbook(API_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[1]]
    endpoints: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        method = str(row[4] or '').strip().upper() if len(row) > 4 else ''
        endpoint = str(row[5] or '').strip() if len(row) > 5 else ''
        if method in {'GET', 'POST', 'PUT', 'PATCH', 'DELETE'} and endpoint:
            endpoints.add(endpoint)
    return endpoints


def read_db_tables() -> set[str]:
    wb = load_workbook(DB_XLSX, read_only=True, data_only=True)
    return {name for name in wb.sheetnames if str(name).startswith('TB_')}


def find_missing(items: set[str], text: str) -> list[str]:
    return sorted(item for item in items if item not in text)


def print_table(rows: list[list[str]]) -> None:
    widths = [max(len(str(row[i])) for row in rows) for i in range(len(rows[0]))]
    for idx, row in enumerate(rows):
        line = ' | '.join(str(cell).ljust(widths[i]) for i, cell in enumerate(row))
        print(line)
        if idx == 0:
            print('-+-'.join('-' * w for w in widths))


def main() -> int:
    if not SPEC_MD.exists():
        print(f'Spec file not found: {SPEC_MD}')
        return 1

    spec_text = SPEC_MD.read_text(encoding='utf-8')
    req_ids = read_req_ids() | read_feature_ids()
    endpoints = read_endpoints()
    tables = read_db_tables()

    missing_req = find_missing(req_ids, spec_text)
    missing_api = find_missing(endpoints, spec_text)
    missing_db = find_missing(tables, spec_text)

    summary_rows = [
        ['Category', 'Total', 'Missing', 'Pass'],
        ['ReqID', str(len(req_ids)), str(len(missing_req)), str(len(req_ids) - len(missing_req))],
        ['API Endpoint', str(len(endpoints)), str(len(missing_api)), str(len(endpoints) - len(missing_api))],
        ['DB Table', str(len(tables)), str(len(missing_db)), str(len(tables) - len(missing_db))],
    ]
    print_table(summary_rows)

    if missing_req:
        print('\nMissing ReqID:', ', '.join(missing_req[:20]), ('...(+%d)' % (len(missing_req)-20) if len(missing_req) > 20 else ''))
    if missing_api:
        print('\nMissing API Endpoint:', ', '.join(missing_api[:20]), ('...(+%d)' % (len(missing_api)-20) if len(missing_api) > 20 else ''))
    if missing_db:
        print('\nMissing DB Table:', ', '.join(missing_db[:20]), ('...(+%d)' % (len(missing_db)-20) if len(missing_db) > 20 else ''))

    report = {
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'summary': {
            'req_id': {'total': len(req_ids), 'missing': len(missing_req)},
            'api_endpoint': {'total': len(endpoints), 'missing': len(missing_api)},
            'db_table': {'total': len(tables), 'missing': len(missing_db)},
        },
        'missing': {
            'req_ids': missing_req,
            'api_endpoints': missing_api,
            'db_tables': missing_db,
        },
        'sources': {
            'requirements_csv': str(REQ_CSV),
            'features_csv': str(FEAT_CSV),
            'api_xlsx': str(API_XLSX),
            'db_xlsx': str(DB_XLSX),
            'spec_md': str(SPEC_MD),
        },
    }

    REPORT_JSON.parent.mkdir(parents=True, exist_ok=True)
    REPORT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'\nReport written: {REPORT_JSON}')

    if missing_req or missing_api or missing_db:
        return 1
    return 0


if __name__ == '__main__':
    sys.exit(main())
