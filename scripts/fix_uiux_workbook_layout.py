#!/usr/bin/env python3
"""Workbook style/layout fixer for CS_RAG_UI_UX workbook.

What this script enforces on all sheets:
1) Three-level visual hierarchy (section > header > data).
2) Remove meaningless legacy highlight colors.
3) Remove borders/fills from empty cells.
4) Auto-wrap and auto-fit row heights to reduce clipped text.
5) Hide default sheet gridlines for cleaner presentation.
"""

from __future__ import annotations

import math
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
UIUX_DIR = ROOT / "docs" / "uiux"


# Typography
FONT_TITLE = Font(name="Malgun Gothic", size=13, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Malgun Gothic", size=10, bold=True, color="0F172A")
FONT_HEADER = Font(name="Malgun Gothic", size=10, bold=True, color="1F2937")
FONT_DATA = Font(name="Malgun Gothic", size=10, bold=False, color="0F172A")

# Border
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NO_BORDER = Border()

# Gradient hierarchy
FILL_TITLE = PatternFill("solid", fgColor="1D4ED8")        # strongest
FILL_SECTION = PatternFill("solid", fgColor="C8D9F1")      # darker than header
FILL_HEADER = PatternFill("solid", fgColor="DFE8F7")       # medium
FILL_DATA = PatternFill("solid", fgColor="F7FAFF")         # lighter
FILL_NONE = PatternFill(fill_type=None)

# Legacy colors to purge
LEGACY_FILL_CODES = {
    "FFFFF2CC",
    "00FFF2CC",
    "FFFFE699",
    "00FFE699",
    "FFFFD966",
    "00FFD966",
    "FF4472C4",
    "FFD9E2F3",
    "FF34495E",
}

HEADER_KEYWORDS = {
    "no",
    "항목",
    "항목유형",
    "항목키",
    "설명",
    "비고",
    "결과",
    "검증 항목",
    "누락 개수",
    "시트명",
    "phase",
    "reqid",
    "api",
    "db",
    "telemetry",
    "screen id",
    "프로그램 id",
    "메시지코드",
    "메시지 내용",
    "http",
    "권한",
    "처리",
    "단위테스트",
}

IDENTIFIER_PATTERNS = (
    re.compile(r"^[A-Z]{2,}(?:-[A-Z0-9_]+)+$"),  # e.g., AI-009-422-SCHEMA
    re.compile(r"^TB_[A-Z0-9_]+$"),              # e.g., TB_MESSAGE
    re.compile(r"^/[a-z0-9_/{\\}-]+$"),          # e.g., /v1/admin/...
)


def find_workbook() -> Path:
    files = sorted(UIUX_DIR.glob("CS_RAG_UI_UX_*.xlsx"))
    if not files:
        raise FileNotFoundError("CS_RAG_UI_UX workbook not found under docs/uiux")
    return files[0]


def normalize_fill_rgb(cell) -> str:
    rgb = getattr(cell.fill.fgColor, "rgb", None)
    return str(rgb).upper() if rgb else ""


def is_empty_value(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def iter_active_rows(ws, scan_cols: int = 8, max_rows: int = 2200) -> set[int]:
    active: set[int] = set()
    for row in range(1, min(ws.max_row, max_rows) + 1):
        if any(not is_empty_value(ws.cell(row=row, column=col).value) for col in range(1, scan_cols + 1)):
            active.add(row)
    return active


def is_section_row(text: str) -> bool:
    t = text.strip()
    return bool(
        re.match(r"^\d+(\-\d+)?\.", t)     # e.g., "5. DB ..."
        or t.lower().startswith("appendix")
        or "checklist" in t.lower()
    )


def _is_identifier_like(text: str) -> bool:
    t = text.strip()
    if not t:
        return False
    for pat in IDENTIFIER_PATTERNS:
        if pat.match(t):
            return True
    return False


def _header_keyword_hits(text_values: list[str]) -> int:
    return sum(1 for v in text_values if v.strip().lower() in HEADER_KEYWORDS)


def is_header_row(text_values: list[str]) -> bool:
    if len(text_values) < 3:
        return False

    first = text_values[0].strip()
    if _is_identifier_like(first):
        return False

    lowered = " ".join(text_values).lower()

    # Data-like rows must never be classified as header.
    data_markers = (
        "tb_",
        "api-",
        "/v1/",
        "tenant_id",
        "session_id",
        "message_id",
        "trace_id",
        "error_code",
        "onclick",
    )
    if any(marker in lowered for marker in data_markers):
        return False

    # Explicit header keywords are the strongest signal.
    keyword_hits = _header_keyword_hits(text_values)
    if keyword_hits >= 2:
        return True

    # Fallback: short-label row with little numeric content.
    short_count = sum(1 for v in text_values if len(v.strip()) <= 24)
    numeric_like = sum(1 for v in text_values if re.fullmatch(r"\d{1,4}", v.strip() or ""))
    return short_count >= 3 and numeric_like <= 1


def set_column_widths(ws) -> None:
    # Wider widths reduce clipping; keep per-sheet special handling.
    if ws.title.startswith("91_"):
        widths = (14, 28, 82, 34, 30, 26, 20, 18)
    elif ws.title.startswith("00_"):
        widths = (10, 36, 16, 36, 24, 18, 16, 16)
    elif ws.title.startswith(("01_", "02_", "35_", "37_", "38_", "90_", "92_")):
        widths = (34, 30, 24, 38, 30, 24, 20, 18)
    else:
        widths = (34, 30, 24, 38, 30, 24, 20, 18)

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def get_col_width(ws, col: int) -> float:
    width = ws.column_dimensions[get_column_letter(col)].width
    return float(width) if width else 8.43


def estimate_wrapped_lines(text: str, width: float) -> int:
    chars_per_line = max(8, int(width * 1.05))
    total = 0
    for line in text.splitlines() or [""]:
        ln = len(line) if line else 1
        total += max(1, math.ceil(ln / chars_per_line))
    return max(1, total)


def classify_rows(ws, active_rows: set[int]) -> tuple[set[int], set[int]]:
    section_rows: set[int] = set()
    header_rows: set[int] = set()

    sorted_rows = sorted(active_rows)

    for row in sorted_rows:
        if row == 1:
            continue
        first_cell = ws.cell(row=row, column=1).value
        if isinstance(first_cell, str) and is_section_row(first_cell):
            section_rows.add(row)

    # Primary rule: row immediately after each section label is a table header.
    max_row = max(sorted_rows) if sorted_rows else 1
    for section_row in sorted(section_rows):
        for candidate in range(section_row + 1, min(section_row + 4, max_row + 1)):
            if candidate not in active_rows:
                continue
            first_cell = ws.cell(row=candidate, column=1).value
            if isinstance(first_cell, str) and is_section_row(first_cell):
                continue
            header_rows.add(candidate)
            break

    # Secondary rule: catch standalone tables not preceded by section title.
    for row in sorted_rows:
        if row in section_rows or row in header_rows or row == 1:
            continue

        values = [ws.cell(row=row, column=col).value for col in range(1, 9)]
        text_values = [v for v in values if isinstance(v, str) and v.strip()]
        if not text_values:
            continue

        prev = row - 1
        prev_non_empty_count = 0
        if prev >= 1:
            prev_non_empty_count = sum(
                1 for col in range(1, 9) if not is_empty_value(ws.cell(prev, col).value)
            )
        # Header candidates usually start right after a section line or a blank gap.
        start_like = prev in section_rows or prev_non_empty_count <= 1
        if start_like and is_header_row(text_values):
            header_rows.add(row)
            continue

        # Top preamble rows (2~4) must be explicit keyword headers to avoid
        # misclassifying first data rows as headers.
        if row <= 4 and _header_keyword_hits(text_values) >= 2:
            header_rows.add(row)

    return section_rows, header_rows


def style_cell(ws, row: int, col: int, row_style: str) -> None:
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return

    value_exists = not is_empty_value(cell.value)
    current_fill = normalize_fill_rgb(cell)
    if current_fill in LEGACY_FILL_CODES:
        cell.fill = FILL_NONE

    if row_style == "title":
        cell.fill = FILL_TITLE
        cell.font = FONT_TITLE
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return

    # For non-title rows, empty cells should not show grid-like borders.
    if not value_exists:
        cell.fill = FILL_NONE
        cell.border = NO_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return

    if row_style == "section":
        cell.fill = FILL_SECTION
        cell.font = FONT_SECTION
    elif row_style == "header":
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
    else:
        cell.fill = FILL_DATA
        cell.font = FONT_DATA

    cell.border = BORDER
    cell.alignment = Alignment(
        horizontal="left" if col <= 5 else "center",
        vertical="center",
        wrap_text=True,
    )


def auto_fit_rows(ws, active_rows: set[int], max_row_scan: int = 2200) -> None:
    if not active_rows:
        return

    last_active = max(active_rows)
    scan_to = min(max(last_active + 120, ws.max_row), max_row_scan)

    merged_widths: dict[tuple[int, int], float] = {}
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        merged_widths[(min_row, min_col)] = sum(get_col_width(ws, c) for c in range(min_col, max_col + 1))

    for row in range(1, scan_to + 1):
        # Use compact default for non-active rows.
        if row not in active_rows and row != 1:
            ws.row_dimensions[row].height = 24
            continue

        max_lines = 1
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.strip():
                width = merged_widths.get((row, col), get_col_width(ws, col))
                max_lines = max(max_lines, estimate_wrapped_lines(cell.value, width))

        if row == 1:
            ws.row_dimensions[row].height = 32
        else:
            ws.row_dimensions[row].height = min(240, max(30, 20 + (max_lines - 1) * 16))


def cleanup_empty_cells(ws, max_row_scan: int = 2200, max_col_scan: int = 20) -> None:
    scan_to = min(max(ws.max_row, 1), max_row_scan)
    for row in range(1, scan_to + 1):
        for col in range(1, min(max(ws.max_column, 1), max_col_scan) + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if is_empty_value(cell.value):
                cell.fill = FILL_NONE
                cell.border = NO_BORDER


def remove_empty_merged_ranges(ws, max_row_scan: int = 2200) -> None:
    # Empty merged blocks often leave visual artifacts; safely unmerge them.
    ranges = list(ws.merged_cells.ranges)
    for merged in ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        if min_row > max_row_scan:
            continue

        all_empty = True
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if not is_empty_value(ws.cell(r, c).value):
                    all_empty = False
                    break
            if not all_empty:
                break

        if not all_empty:
            continue

        ws.unmerge_cells(str(merged))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(r, c)
                cell.fill = FILL_NONE
                cell.border = NO_BORDER


def apply_sheet(ws) -> None:
    set_column_widths(ws)
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False

    active_rows = iter_active_rows(ws)
    if not active_rows:
        cleanup_empty_cells(ws)
        return

    section_rows, header_rows = classify_rows(ws, active_rows)
    last_active = max(active_rows)
    style_scan_to = min(max(last_active + 120, ws.max_row), 2200)

    for row in range(1, style_scan_to + 1):
        if row == 1:
            row_style = "title"
        elif row in section_rows:
            row_style = "section"
        elif row in header_rows:
            row_style = "header"
        elif row in active_rows:
            row_style = "data"
        else:
            row_style = "empty"

        for col in range(1, 9):
            style_cell(ws, row, col, row_style)

    auto_fit_rows(ws, active_rows)
    remove_empty_merged_ranges(ws)
    cleanup_empty_cells(ws)


def main() -> None:
    workbook = find_workbook()
    wb = load_workbook(workbook)

    for ws in wb.worksheets:
        apply_sheet(ws)

    wb.save(workbook)
    print(f"Updated workbook: {workbook}")


if __name__ == "__main__":
    main()
