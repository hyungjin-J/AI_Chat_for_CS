#!/usr/bin/env python3
"""Shared gate logic for UIUX workbook QA and validation."""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "docs" / "uiux" / "CS_RAG_UI_UX_설계서.xlsx"
GATE_REPORT_PATH = ROOT / "docs" / "uiux" / "reports" / "xlsx_gate_report.json"

REQ_CSV_PATH = ROOT / "docs" / "references" / "CS AI Chatbot_Requirements Statement.csv"
FEATURE_CSV_PATH = ROOT / "docs" / "references" / "Summary of key features.csv"
API_XLSX_PATH = ROOT / "docs" / "references" / "google_ready_api_spec_v0.3_20260216.xlsx"
DB_XLSX_PATH = ROOT / "docs" / "references" / "CS_AI_CHATBOT_DB.xlsx"

REQ_ID_RE = re.compile(r"\b[A-Z]{2,5}-\d{3}\b")
ERROR_CODE_RE = re.compile(r"\b[A-Z]{2,5}-\d{3}(?:-[A-Z0-9_]+){1,3}\b")
TABLE_RE = re.compile(r"\bTB_[A-Z0-9_]+\b")
METHOD_ENDPOINT_RE = re.compile(r"\b(GET|POST|PUT|PATCH|DELETE)\s+(/v1/[A-Za-z0-9_{}\-/]+)")
ENDPOINT_RE = re.compile(r"(/v1/[A-Za-z0-9_{}\-/]+)")

SCREEN_NAME_RE = re.compile(r"^(?P<seq>\d{2})_(?P<token>[A-Z]{3}\d{3})_")
KNOWN_ROLES = {"AGENT", "CUSTOMER", "ADMIN", "OPS", "SYSTEM", "PUBLIC"}
DEFAULT_SSE_TYPES = {"token", "tool", "citation", "done", "error", "heartbeat", "safe_response"}


REQUIRED_SHEET_PREFIX = {
    "toc": "00_",
    "error": "01_",
    "code": "02_",
    "rbac": "38_",
    "inconsistency": "90_",
    "trace": "91_",
    "validation": "93_",
}

REQUIRED_DEFAULT_NAME = {
    "toc": "00_통합목차",
    "error": "01_에러메시지코드",
    "code": "02_추가종합코드",
    "rbac": "38_권한 별 UI",
    "inconsistency": "90_불일치목록",
    "trace": "91_추적성매트릭스",
    "validation": "93_검증결과",
}


@dataclass
class ScreenMeta:
    sheet_name: str
    expected_screen_id: str
    actual_screen_id: str
    program_text: str
    api_text: str
    role: str
    phase: str
    has_section_input: bool
    has_section_button: bool
    has_section_exception: bool
    has_section_test: bool
    has_constraints: bool
    used_error_codes: set[str]
    used_req_ids: set[str]
    used_tables: set[str]
    used_endpoints: set[str]
    used_sse_types: set[str]


def _sheet_by_prefix(wb: Workbook, prefix: str) -> str | None:
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return name
    return None


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _joined_row(ws: Worksheet, row: int, min_col: int = 1, max_col: int = 6) -> str:
    vals: list[str] = []
    for col in range(min_col, max_col + 1):
        v = _safe_str(ws.cell(row=row, column=col).value)
        if v:
            vals.append(v)
    return " | ".join(vals)


def _extract_field(ws: Worksheet, row: int, label_hint: str) -> str:
    a = _safe_str(ws.cell(row=row, column=1).value)
    b = _safe_str(ws.cell(row=row, column=2).value)
    if b:
        return b
    if label_hint in a and ":" in a:
        right = a.split(":", 1)[1].strip()
        if right:
            return right
    joined = _joined_row(ws, row)
    return joined


def _extract_role(text: str) -> str:
    up = text.upper()
    for role in sorted(KNOWN_ROLES):
        if re.search(rf"\b{role}\b", up):
            return role
    return ""


def _extract_phase(text: str) -> str:
    up = text.upper()
    match = re.search(r"\bPHASE\d+\b", up)
    if match:
        return match.group(0)
    if "MVP" in up:
        return "MVP"
    return ""


def list_screen_sheets(wb: Workbook) -> list[str]:
    screens: list[tuple[int, str]] = []
    for name in wb.sheetnames:
        m = SCREEN_NAME_RE.match(name)
        if not m:
            continue
        seq = int(m.group("seq"))
        screens.append((seq, name))
    screens.sort(key=lambda x: x[0])
    return [name for _, name in screens]


def _extract_expected_screen_id(sheet_name: str) -> str:
    m = SCREEN_NAME_RE.match(sheet_name)
    if not m:
        return ""
    token = m.group("token")
    return f"{token[:3]}-{token[3:]}"


def _scan_sheet_texts(ws: Worksheet, max_row: int = 220, max_col: int = 8) -> list[str]:
    texts: list[str] = []
    limit = min(ws.max_row, max_row)
    for row in range(1, limit + 1):
        for col in range(1, max_col + 1):
            v = ws.cell(row=row, column=col).value
            s = _safe_str(v)
            if s:
                texts.append(s)
    return texts


def _has_section(ws: Worksheet, section_prefix: str) -> bool:
    for row in range(1, min(ws.max_row, 220) + 1):
        v = _safe_str(ws.cell(row=row, column=1).value)
        if v.startswith(section_prefix):
            return True
    return False


def _has_constraints_table(ws: Worksheet) -> bool:
    for row in range(1, min(ws.max_row, 260) + 1):
        v = _safe_str(ws.cell(row=row, column=1).value)
        if "필수 제약" in v or "공통 제약" in v:
            return True
    return False


def _extract_endpoints_from_text(text: str) -> set[str]:
    endpoints: set[str] = set()
    for _, ep in METHOD_ENDPOINT_RE.findall(text):
        endpoints.add(ep.strip())
    if not endpoints:
        for ep in ENDPOINT_RE.findall(text):
            endpoints.add(ep.strip())
    return endpoints


def _extract_sse_types_from_text(text: str, allowed: set[str]) -> set[str]:
    found: set[str] = set()
    lower = text.lower()
    for token in allowed:
        if re.search(rf"\b{re.escape(token.lower())}\b", lower):
            found.add(token.lower())
    return found


def collect_screen_meta(wb: Workbook, sse_allowed: set[str]) -> list[ScreenMeta]:
    metas: list[ScreenMeta] = []
    for name in list_screen_sheets(wb):
        ws = wb[name]
        expected_id = _extract_expected_screen_id(name)
        actual_id = _extract_field(ws, 4, "화면")
        program_text = _extract_field(ws, 6, "프로그램")
        api_text = _extract_field(ws, 7, "API")
        role_text = _extract_field(ws, 8, "권한")
        phase_text = _extract_field(ws, 9, "개발")
        role = _extract_role(role_text)
        phase = _extract_phase(phase_text)

        texts = _scan_sheet_texts(ws)
        joined = "\n".join(texts)
        used_error_codes = set(ERROR_CODE_RE.findall(joined))
        used_req_ids = set(REQ_ID_RE.findall(joined))
        used_tables = set(TABLE_RE.findall(joined))

        used_endpoints: set[str] = set()
        used_sse: set[str] = set()
        for t in texts:
            used_endpoints |= _extract_endpoints_from_text(t)
            used_sse |= _extract_sse_types_from_text(t, sse_allowed)

        metas.append(
            ScreenMeta(
                sheet_name=name,
                expected_screen_id=expected_id,
                actual_screen_id=actual_id,
                program_text=program_text,
                api_text=api_text,
                role=role,
                phase=phase,
                has_section_input=_has_section(ws, "3."),
                has_section_button=_has_section(ws, "4."),
                has_section_exception=_has_section(ws, "8."),
                has_section_test=_has_section(ws, "9."),
                has_constraints=_has_constraints_table(ws),
                used_error_codes=used_error_codes,
                used_req_ids=used_req_ids,
                used_tables=used_tables,
                used_endpoints=used_endpoints,
                used_sse_types=used_sse,
            )
        )
    return metas


def load_reference_catalog() -> dict[str, Any]:
    req_ids: set[str] = set()
    with REQ_CSV_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        rd = csv.reader(f)
        next(rd, None)
        for row in rd:
            if row and _safe_str(row[0]):
                req_ids.add(_safe_str(row[0]))

    feat_req_ids: set[str] = set()
    with FEATURE_CSV_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        rd = csv.reader(f)
        next(rd, None)
        for row in rd:
            if len(row) >= 6 and _safe_str(row[5]):
                feat_req_ids.add(_safe_str(row[5]))

    from openpyxl import load_workbook

    api_wb = load_workbook(API_XLSX_PATH, read_only=True, data_only=True)
    api_ws = api_wb[api_wb.sheetnames[1]]
    api_entries: list[dict[str, str]] = []
    api_endpoints: set[str] = set()
    for row in api_ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            continue
        pid = _safe_str(row[2])
        method = _safe_str(row[4]).upper()
        ep = _safe_str(row[5])
        if not pid or not method or not ep:
            continue
        api_entries.append({"program_id": pid, "method": method, "endpoint": ep})
        api_endpoints.add(ep)

    db_wb = load_workbook(DB_XLSX_PATH, read_only=True, data_only=True)
    db_tables = {name for name in db_wb.sheetnames if name.startswith("TB_")}

    return {
        "req_ids": req_ids,
        "feature_req_ids": feat_req_ids,
        "all_req_ids": req_ids | feat_req_ids,
        "api_entries": api_entries,
        "api_endpoints": api_endpoints,
        "db_tables": db_tables,
    }


def extract_error_catalog(error_ws: Worksheet) -> set[str]:
    codes: set[str] = set()
    for row in range(1, min(error_ws.max_row, 400) + 1):
        for col in range(1, 4):
            v = _safe_str(error_ws.cell(row=row, column=col).value)
            if not v:
                continue
            for code in ERROR_CODE_RE.findall(v):
                codes.add(code)
    return codes


def extract_roles_and_sse(code_ws: Worksheet) -> tuple[set[str], set[str]]:
    roles: set[str] = set()
    sse_types: set[str] = set()
    for row in range(1, min(code_ws.max_row, 500) + 1):
        group = _safe_str(code_ws.cell(row=row, column=1).value).upper()
        code = _safe_str(code_ws.cell(row=row, column=2).value)
        if group == "ROLE" and code:
            roles.add(code.upper())
        if group == "SSE_EVENT_TYPE" and code:
            sse_types.add(code.lower())
    if not roles:
        roles = set(KNOWN_ROLES)
    if not sse_types:
        sse_types = set(DEFAULT_SSE_TYPES)
    return roles, sse_types


def parse_trace_matrix_incomplete(trace_ws: Worksheet) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for row in range(4, min(trace_ws.max_row, 3000) + 1):
        req = _safe_str(trace_ws.cell(row=row, column=1).value)
        if not REQ_ID_RE.fullmatch(req):
            continue
        b = _safe_str(trace_ws.cell(row=row, column=2).value)
        c = _safe_str(trace_ws.cell(row=row, column=3).value)
        d = _safe_str(trace_ws.cell(row=row, column=4).value)
        e = _safe_str(trace_ws.cell(row=row, column=5).value)
        missing = [name for name, val in (("Screen", b), ("API", c), ("DB", d), ("Telemetry", e)) if not val]
        if missing:
            issues.append({"row": row, "req_id": req, "missing": missing})
    return issues


def _parse_toc_entries(toc_ws: Worksheet) -> set[str]:
    entries: set[str] = set()
    for row in range(4, min(toc_ws.max_row, 1200) + 1):
        s = _safe_str(toc_ws.cell(row=row, column=2).value)
        if s:
            entries.add(s)
    return entries


def run_gate_checks(wb: Workbook) -> dict[str, Any]:
    sheet_map = {key: _sheet_by_prefix(wb, prefix) for key, prefix in REQUIRED_SHEET_PREFIX.items()}
    missing_required = [key for key, name in sheet_map.items() if not name]

    refs = load_reference_catalog()

    code_ws = wb[sheet_map["code"]] if sheet_map["code"] else None
    error_ws = wb[sheet_map["error"]] if sheet_map["error"] else None
    trace_ws = wb[sheet_map["trace"]] if sheet_map["trace"] else None
    toc_ws = wb[sheet_map["toc"]] if sheet_map["toc"] else None

    roles, sse_types = extract_roles_and_sse(code_ws) if code_ws else (set(KNOWN_ROLES), set(DEFAULT_SSE_TYPES))
    error_catalog = extract_error_catalog(error_ws) if error_ws else set()
    screen_meta = collect_screen_meta(wb, sse_types)

    screen_id_mismatch: list[dict[str, str]] = []
    missing_mandatory_fields: list[dict[str, Any]] = []
    missing_sections: list[dict[str, Any]] = []
    missing_constraints: list[str] = []
    role_mismatch: list[dict[str, str]] = []
    sse_mismatch: list[dict[str, Any]] = []

    used_error_codes: set[str] = set()
    used_req_ids: set[str] = set()
    used_tables: set[str] = set()
    used_endpoints: set[str] = set()

    for meta in screen_meta:
        if meta.expected_screen_id and meta.actual_screen_id and meta.expected_screen_id != meta.actual_screen_id:
            screen_id_mismatch.append(
                {
                    "sheet": meta.sheet_name,
                    "expected": meta.expected_screen_id,
                    "actual": meta.actual_screen_id,
                }
            )

        missing_fields: list[str] = []
        if not meta.program_text:
            missing_fields.append("program_id")
        if not meta.api_text:
            missing_fields.append("api")
        if not meta.role:
            missing_fields.append("role")
        if not meta.phase:
            missing_fields.append("phase")
        if missing_fields:
            missing_mandatory_fields.append({"sheet": meta.sheet_name, "missing_fields": missing_fields})

        missing_sec: list[str] = []
        if not meta.has_section_input:
            missing_sec.append("입력/조회 필드")
        if not meta.has_section_button:
            missing_sec.append("버튼 동작")
        if not meta.has_section_exception:
            missing_sec.append("예외사항")
        if not meta.has_section_test:
            missing_sec.append("단위테스트")
        if missing_sec:
            missing_sections.append({"sheet": meta.sheet_name, "missing_sections": missing_sec})

        if not meta.has_constraints:
            missing_constraints.append(meta.sheet_name)

        if meta.role and meta.role.upper() not in roles:
            role_mismatch.append({"sheet": meta.sheet_name, "role": meta.role})

        unsupported_sse = sorted(meta.used_sse_types - sse_types)
        if unsupported_sse:
            sse_mismatch.append({"sheet": meta.sheet_name, "unsupported_sse": unsupported_sse})

        used_error_codes |= meta.used_error_codes
        used_req_ids |= meta.used_req_ids
        used_tables |= meta.used_tables
        used_endpoints |= meta.used_endpoints

    error_code_mismatch = sorted(code for code in used_error_codes if code not in error_catalog)

    trace_incomplete = parse_trace_matrix_incomplete(trace_ws) if trace_ws else []

    toc_missing: list[str] = []
    toc_extra: list[str] = []
    if toc_ws:
        toc_entries = _parse_toc_entries(toc_ws)
        wb_sheet_set = {name for name in wb.sheetnames if name != sheet_map.get("toc")}
        toc_missing = sorted(wb_sheet_set - toc_entries)
        toc_extra = sorted(toc_entries - wb_sheet_set)
    else:
        toc_missing = sorted(wb.sheetnames)

    api_unmapped = sorted(refs["api_endpoints"] - used_endpoints)
    db_unmapped = sorted(refs["db_tables"] - used_tables)
    req_unmapped = sorted(refs["all_req_ids"] - used_req_ids)

    checks = [
        {
            "section": "B1",
            "item": "00_통합목차에 모든 시트가 등록되었는가",
            "status": "PASS" if not toc_missing and not toc_extra else "FAIL",
            "missing_count": len(toc_missing) + len(toc_extra),
            "note": f"missing:{len(toc_missing)}, extra:{len(toc_extra)}",
        },
        {
            "section": "B1",
            "item": "필수 시트가 존재하는가",
            "status": "PASS" if not missing_required else "FAIL",
            "missing_count": len(missing_required),
            "note": ", ".join(missing_required) if missing_required else "",
        },
        {
            "section": "B2",
            "item": "화면ID가 시트명 패턴과 일치하는가",
            "status": "PASS" if not screen_id_mismatch else "FAIL",
            "missing_count": len(screen_id_mismatch),
            "note": "mismatch list 참고",
        },
        {
            "section": "B2",
            "item": "프로그램ID/API/권한/개발일(Phase) 필드 누락이 없는가",
            "status": "PASS" if not missing_mandatory_fields else "FAIL",
            "missing_count": len(missing_mandatory_fields),
            "note": "missing field list 참고",
        },
        {
            "section": "B2",
            "item": "필수 섹션(입력/버튼/예외/단위테스트) 누락이 없는가",
            "status": "PASS" if not missing_sections else "FAIL",
            "missing_count": len(missing_sections),
            "note": "missing section list 참고",
        },
        {
            "section": "B2",
            "item": "프로젝트 공통 제약 표가 모든 화면 시트에 존재하는가",
            "status": "PASS" if not missing_constraints else "FAIL",
            "missing_count": len(missing_constraints),
            "note": "constraints missing list 참고",
        },
        {
            "section": "B3",
            "item": "화면 시트 에러코드가 01_에러메시지코드와 일치하는가",
            "status": "PASS" if not error_code_mismatch else "FAIL",
            "missing_count": len(error_code_mismatch),
            "note": "error mismatch list 참고",
        },
        {
            "section": "B3",
            "item": "화면 권한 값이 02_추가종합코드 ROLE 그룹과 일치하는가",
            "status": "PASS" if not role_mismatch else "FAIL",
            "missing_count": len(role_mismatch),
            "note": "role mismatch list 참고",
        },
        {
            "section": "B3",
            "item": "SSE 이벤트 타입이 02_추가종합코드(SSE_EVENT_TYPE)에 정의되어 있는가",
            "status": "PASS" if not sse_mismatch else "FAIL",
            "missing_count": len(sse_mismatch),
            "note": "sse mismatch list 참고",
        },
        {
            "section": "B3",
            "item": "91_추적성매트릭스 ReqID→Screen→API→DB→Telemetry 링크가 최소 연결되는가",
            "status": "PASS" if not trace_incomplete else "FAIL",
            "missing_count": len(trace_incomplete),
            "note": "trace incomplete list 참고",
        },
        {
            "section": "B4",
            "item": "API 미매핑 목록",
            "status": "PASS" if not api_unmapped else "FAIL",
            "missing_count": len(api_unmapped),
            "note": "api unmapped list 참고",
        },
        {
            "section": "B4",
            "item": "DB 미매핑 목록",
            "status": "PASS" if not db_unmapped else "FAIL",
            "missing_count": len(db_unmapped),
            "note": "db unmapped list 참고",
        },
        {
            "section": "B4",
            "item": "ReqID 미매핑 목록",
            "status": "PASS" if not req_unmapped else "FAIL",
            "missing_count": len(req_unmapped),
            "note": "req unmapped list 참고",
        },
    ]

    pass_count = sum(1 for c in checks if c["status"] == "PASS")
    fail_count = sum(1 for c in checks if c["status"] == "FAIL")

    return {
        "sheet_map": sheet_map,
        "missing_required": missing_required,
        "screen_sheets": [m.sheet_name for m in screen_meta],
        "screen_meta": [m.__dict__ for m in screen_meta],
        "screen_id_mismatch": screen_id_mismatch,
        "missing_mandatory_fields": missing_mandatory_fields,
        "missing_sections": missing_sections,
        "missing_constraints": missing_constraints,
        "error_code_mismatch": error_code_mismatch,
        "role_mismatch": role_mismatch,
        "sse_mismatch": sse_mismatch,
        "trace_incomplete": trace_incomplete,
        "toc_missing": toc_missing,
        "toc_extra": toc_extra,
        "api_unmapped": api_unmapped,
        "db_unmapped": db_unmapped,
        "req_unmapped": req_unmapped,
        "checks": checks,
        "pass_count": pass_count,
        "fail_count": fail_count,
    }


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if isinstance(value, set):
        return sorted(_json_safe(v) for v in value)
    return value


def write_json_report(report: dict[str, Any], path: Path = GATE_REPORT_PATH) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(_json_safe(report), f, ensure_ascii=False, indent=2)
