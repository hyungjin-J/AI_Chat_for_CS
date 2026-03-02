#!/usr/bin/env python3
"""Cross-spec consistency checks anchored on Requirements ReqID SSOT."""

from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


REQ_ID_STRICT_RE = re.compile(r"\b[A-Z]{2,5}-\d{3}\b")
REQ_ID_LAX_RE = re.compile(r"\b[A-Z]{2,5}-[A-Za-z0-9]{3}\b")
REQID_TAG_RE = re.compile(r"ReqID\+?\s*:\s*([^\n\r;]+)", re.IGNORECASE)
DEFAULT_TERMINOLOGY_SSOT = "docs/review/mvp_verification_pack/TERMINOLOGY_SSOT.json"

ROLE_TAXONOMY = {"AGENT", "CUSTOMER", "ADMIN", "OPS", "SYSTEM"}
ACCESS_LEVEL_TAXONOMY = {"PUBLIC", "AUTHENTICATED"}
NOTE_KEY_VALUE_RE = re.compile(r"\b([a-z_][a-z0-9_]*)\s*=\s*([a-z0-9_./:-]+)", re.IGNORECASE)
AUTH_REFRESH_ENDPOINT = "/v1/auth/refresh"
PLACEHOLDER_VALUES = {"", "-", "N/A", "NA", "NONE", "NULL"}
UIUX_PLACEHOLDER_VALUES = {"", "-", "—", "TBD", "TODO", "TEMP", "PLACEHOLDER", "N/A", "NA"}


# Keep Korean labels as escapes to avoid shell/codepage drift on Windows.
KOR_REQ_ID = "\uc694\uad6c\uc0ac\ud56dID"      # 요구사항ID
KOR_API_LIST = "\uc804\uccb4API\ubaa9\ub85d"   # 전체API목록
KOR_NOTE = "\ube44\uace0"                      # 비고
KOR_ROLE = "\uad8c\ud55c"                      # 권한
KOR_UIUX_UNMAPPED_PREFIX = "94_"
KOR_UIUX_MAPPING_KEY = "\ud56d\ubaa9\ud0a4"  # 항목키
KOR_UIUX_MAPPING_SCREEN = "\ub9e4\ud551\ub300\uc0c1 screen id"  # 매핑대상 screen id
KOR_UIUX_MAPPING_SHEET = "\ub9e4\ud551\ub300\uc0c1 \uc2dc\ud2b8\uba85"  # 매핑대상 시트명
KOR_UIUX_ERROR_PREFIX = "01_"
KOR_UIUX_INCONSISTENCY_PREFIX = "90_"

UIUX_TARGET_ERROR_CONTRACT = {
    "SEC-003-409-PII": {"http_status": "409", "assume_id": "ASSUME-001"},
    "SYS-001-404": {"http_status": "404", "assume_id": "ASSUME-002"},
}


@dataclass(frozen=True)
class TerminologyCheckSpec:
    code_suffix: str
    required_tokens: tuple[str, ...]
    optional_tokens: tuple[str, ...] = ()
    forbidden_variants: tuple[tuple[str, str], ...] = ()


@dataclass(frozen=True)
class TerminologySsot:
    required_exact_tokens: dict[str, tuple[str, ...]]
    forbidden_variants: dict[str, tuple[tuple[str, str], ...]]
    required_headers_required: tuple[str, ...]
    required_headers_optional: tuple[str, ...]
    sse_event_types: tuple[str, ...]
    error_payload_fields: tuple[str, ...]


def _as_token_tuple(values: object, field_name: str) -> tuple[str, ...]:
    if not isinstance(values, list):
        raise ValueError(f"{field_name} must be a list of tokens")

    normalized: list[str] = []
    seen: set[str] = set()
    for item in values:
        if not isinstance(item, str):
            raise ValueError(f"{field_name} must contain only strings")
        token = item.strip()
        if not token:
            raise ValueError(f"{field_name} must not contain empty tokens")
        if token in seen:
            continue
        normalized.append(token)
        seen.add(token)
    return tuple(normalized)


def _as_forbidden_variant_pairs(values: object, field_name: str) -> tuple[tuple[str, str], ...]:
    if not isinstance(values, list):
        raise ValueError(f"{field_name} must be a list of [variant, canonical] pairs")

    normalized: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for item in values:
        if not isinstance(item, list) or len(item) != 2:
            raise ValueError(f"{field_name} entries must be [variant, canonical]")
        variant_raw, canonical_raw = item
        if not isinstance(variant_raw, str) or not isinstance(canonical_raw, str):
            raise ValueError(f"{field_name} entries must contain strings")
        variant = variant_raw.strip()
        canonical = canonical_raw.strip()
        if not variant or not canonical:
            raise ValueError(f"{field_name} entries must not contain empty strings")
        pair = (variant, canonical)
        if pair in seen:
            continue
        normalized.append(pair)
        seen.add(pair)
    return tuple(normalized)


def resolve_terminology_ssot_path(root: Path, value: str) -> Path:
    candidate = resolve_path(root, value)
    if candidate.exists():
        return candidate

    repo_root = Path(__file__).resolve().parent.parent
    fallback = resolve_path(repo_root, value)
    if fallback.exists():
        return fallback

    raise FileNotFoundError(f"terminology SSOT not found: {value}")


def load_terminology_ssot(path: Path) -> TerminologySsot:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(payload, dict):
        raise ValueError("terminology SSOT root must be a JSON object")

    required_exact_raw = payload.get("required_exact_tokens")
    forbidden_raw = payload.get("forbidden_variants")
    headers_raw = payload.get("required_headers")
    sse_raw = payload.get("sse_event_types")
    error_payload_raw = payload.get("error_payload_fields")

    if not isinstance(required_exact_raw, dict):
        raise ValueError("required_exact_tokens must be an object")
    if not isinstance(forbidden_raw, dict):
        raise ValueError("forbidden_variants must be an object")
    if not isinstance(headers_raw, dict):
        raise ValueError("required_headers must be an object")

    required_exact_tokens = {
        key: _as_token_tuple(required_exact_raw.get(key), f"required_exact_tokens.{key}")
        for key in sorted(required_exact_raw)
    }
    forbidden_variants = {
        key: _as_forbidden_variant_pairs(forbidden_raw.get(key), f"forbidden_variants.{key}")
        for key in sorted(forbidden_raw)
    }
    required_headers_required = _as_token_tuple(headers_raw.get("required"), "required_headers.required")
    required_headers_optional = _as_token_tuple(headers_raw.get("optional"), "required_headers.optional")
    sse_event_types = _as_token_tuple(sse_raw, "sse_event_types")
    error_payload_fields = _as_token_tuple(error_payload_raw, "error_payload_fields")

    for required_key in ("secret_ref", "snake_case_contract_fields"):
        if required_key not in required_exact_tokens:
            raise ValueError(f"required_exact_tokens.{required_key} is required")

    for required_key in ("error_payload", "sse_event_types", "snake_case_contract_fields"):
        if required_key not in forbidden_variants:
            raise ValueError(f"forbidden_variants.{required_key} is required")

    return TerminologySsot(
        required_exact_tokens=required_exact_tokens,
        forbidden_variants=forbidden_variants,
        required_headers_required=required_headers_required,
        required_headers_optional=required_headers_optional,
        sse_event_types=sse_event_types,
        error_payload_fields=error_payload_fields,
    )


def compile_sse_event_regex(event_types: tuple[str, ...]) -> re.Pattern[str]:
    escaped = "|".join(re.escape(item) for item in sorted(event_types))
    if not escaped:
        raise ValueError("sse_event_types must not be empty")
    return re.compile(rf"\b({escaped})\b", re.IGNORECASE)


def build_terminology_specs(config: TerminologySsot) -> tuple[TerminologyCheckSpec, ...]:
    return (
        TerminologyCheckSpec(
            code_suffix="SECRET_REF",
            required_tokens=config.required_exact_tokens["secret_ref"],
        ),
        TerminologyCheckSpec(
            code_suffix="ERROR_PAYLOAD",
            required_tokens=config.error_payload_fields,
            forbidden_variants=config.forbidden_variants["error_payload"],
        ),
        TerminologyCheckSpec(
            code_suffix="SSE_EVENT",
            required_tokens=tuple(sorted(config.sse_event_types)),
            forbidden_variants=config.forbidden_variants["sse_event_types"],
        ),
        TerminologyCheckSpec(
            code_suffix="TRACE_TENANT_HEADER",
            required_tokens=config.required_headers_required,
            optional_tokens=config.required_headers_optional,
        ),
        TerminologyCheckSpec(
            code_suffix="SNAKE_CASE",
            required_tokens=config.required_exact_tokens["snake_case_contract_fields"],
            forbidden_variants=config.forbidden_variants["snake_case_contract_fields"],
        ),
    )


@dataclass
class Violation:
    code: str
    file: str
    location: str
    token: str
    message: str


def normalize(path: str) -> str:
    return path.replace("\\", "/").strip()


def to_rel(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError:
        return normalize(path.as_posix())


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Cross-spec consistency checks (ReqID SSOT + terminology)")
    parser.add_argument("--root", default=".")
    parser.add_argument("--requirements", default="docs/references/CS AI Chatbot_Requirements Statement.csv")
    parser.add_argument("--summary", default="docs/references/Summary of key features.csv")
    parser.add_argument("--development", default="docs/references/Development environment.csv")
    parser.add_argument("--api-workbook", default="docs/references/google_ready_api_spec_v0.3_20260216.xlsx")
    parser.add_argument("--api-sheet", default=KOR_API_LIST)
    parser.add_argument("--db-workbook", default="docs/references/CS_AI_CHATBOT_DB.xlsx")
    parser.add_argument(
        "--uiux-workbook",
        help="Optional explicit UI/UX workbook path. If omitted, the first docs/uiux/CS_RAG_UI_UX_*.xlsx is used.",
    )
    parser.add_argument("--uiux-glob", default="docs/uiux/CS_RAG_UI_UX_*.xlsx")
    parser.add_argument("--terminology-ssot", default=DEFAULT_TERMINOLOGY_SSOT)
    parser.add_argument("--report-json", default="docs/uiux/reports/spec_consistency_check_report.json")
    parser.add_argument("--report-txt")
    parser.add_argument("--pass-artifact", help="Optional path for concise PASS summary text artifact.")
    return parser.parse_args()


def resolve_path(root: Path, value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else root / value


def read_csv_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file_pointer:
        rows = list(csv.reader(file_pointer))
    if not rows:
        return [], []
    return rows[0], rows[1:]


def find_column(headers: list[str], candidates: list[str]) -> int | None:
    normalized = [header.strip().lower() for header in headers]
    normalized_candidates = [candidate.strip().lower() for candidate in candidates]

    for candidate in normalized_candidates:
        if candidate in normalized:
            return normalized.index(candidate)

    for candidate in normalized_candidates:
        for index, header in enumerate(normalized):
            if candidate in header:
                return index
    return None


def extract_reqid_tokens(value: str) -> tuple[list[str], list[str]]:
    strict_tokens = sorted(set(REQ_ID_STRICT_RE.findall(value)))
    lax_tokens = sorted(set(REQ_ID_LAX_RE.findall(value)))
    malformed_tokens = sorted(token for token in lax_tokens if token not in strict_tokens)
    return strict_tokens, malformed_tokens


def extract_prefix(token: str) -> str:
    return token.split("-", 1)[0] if "-" in token else ""


def parse_note_key_values(note_value: str) -> dict[str, set[str]]:
    pairs: dict[str, set[str]] = {}
    for key_raw, value_raw in NOTE_KEY_VALUE_RE.findall(note_value):
        key = key_raw.strip().lower()
        value = value_raw.strip().upper()
        if not key or not value:
            continue
        values = pairs.setdefault(key, set())
        values.add(value)
    return pairs


def git_head(root: Path) -> str:
    result = subprocess.run(
        ["git", "rev-parse", "--short", "HEAD"],
        cwd=root,
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if result.returncode != 0:
        return "UNKNOWN"
    return result.stdout.strip() or "UNKNOWN"


def add_violation(
    collection: list[Violation],
    code: str,
    file_path: Path,
    root: Path,
    location: str,
    token: str,
    message: str,
) -> None:
    collection.append(
        Violation(
            code=code,
            file=to_rel(file_path, root),
            location=location,
            token=token,
            message=message,
        )
    )


def load_requirements(root: Path, requirements_path: Path, violations: list[Violation]) -> set[str]:
    headers, rows = read_csv_rows(requirements_path)
    if not headers:
        add_violation(
            violations,
            "REQID_REQUIREMENTS_EMPTY",
            requirements_path,
            root,
            "file",
            "",
            "requirements CSV is empty",
        )
        return set()

    reqid_col = find_column(headers, ["ReqID", KOR_REQ_ID])
    if reqid_col is None:
        reqid_col = 0

    reqids: set[str] = set()
    seen: dict[str, int] = {}
    for row_number, row in enumerate(rows, start=2):
        value = row[reqid_col].strip() if reqid_col < len(row) else ""
        if not value:
            continue
        if not REQ_ID_STRICT_RE.fullmatch(value):
            add_violation(
                violations,
                "REQID_REQUIREMENTS_MALFORMED",
                requirements_path,
                root,
                f"row={row_number}",
                value,
                "requirements ReqID must match [A-Z]{2,5}-\\d{3}",
            )
            continue
        reqids.add(value)
        seen[value] = seen.get(value, 0) + 1

    for token, count in sorted(seen.items()):
        if count > 1:
            add_violation(
                violations,
                "REQID_REQUIREMENTS_DUPLICATE",
                requirements_path,
                root,
                "ReqID",
                token,
                f"duplicate ReqID detected: {count}",
            )
    return reqids


def validate_summary(root: Path, summary_path: Path, req_master: set[str], violations: list[Violation]) -> int:
    headers, rows = read_csv_rows(summary_path)
    if not headers:
        add_violation(
            violations,
            "REQID_SUMMARY_EMPTY",
            summary_path,
            root,
            "file",
            "",
            "summary CSV is empty",
        )
        return 0

    reqid_col = find_column(headers, [KOR_REQ_ID, "ReqID"])
    if reqid_col is None:
        add_violation(
            violations,
            "REQID_SUMMARY_COLUMN_MISSING",
            summary_path,
            root,
            "header",
            KOR_REQ_ID,
            "summary CSV missing 요구사항ID/ReqID column",
        )
        return 0

    token_count = 0
    for row_number, row in enumerate(rows, start=2):
        raw_value = row[reqid_col].strip() if reqid_col < len(row) else ""
        if not raw_value:
            continue

        strict_tokens, malformed_tokens = extract_reqid_tokens(raw_value)
        token_count += len(strict_tokens)
        for token in strict_tokens:
            if token not in req_master:
                add_violation(
                    violations,
                    "REQID_SUMMARY_UNKNOWN",
                    summary_path,
                    root,
                    f"row={row_number}",
                    token,
                    "summary 요구사항ID must exist in Requirements SSOT",
                )
        for token in malformed_tokens:
            add_violation(
                violations,
                "REQID_SUMMARY_MALFORMED",
                summary_path,
                root,
                f"row={row_number}",
                token,
                "summary 요구사항ID contains malformed ReqID token",
            )

        if not strict_tokens and not malformed_tokens and raw_value.upper() not in PLACEHOLDER_VALUES:
            add_violation(
                violations,
                "REQID_SUMMARY_MALFORMED",
                summary_path,
                root,
                f"row={row_number}",
                raw_value,
                "summary 요구사항ID cell has no valid ReqID token",
            )
    return token_count


def select_api_sheet(workbook_path: Path, preferred_sheet_name: str):
    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    preferred_normalized = preferred_sheet_name.strip().lower()

    for worksheet in workbook.worksheets:
        if worksheet.title.strip().lower() == preferred_normalized:
            return workbook, worksheet

    # Fallback for localization/codepage edge cases.
    for worksheet in workbook.worksheets:
        title = worksheet.title.strip().lower()
        if "api" in title and ("list" in title or "\ubaa9\ub85d" in worksheet.title):  # 목록
            return workbook, worksheet

    workbook.close()
    raise ValueError(f"sheet '{preferred_sheet_name}' not found in {workbook_path.as_posix()}")


def validate_api_notes_and_roles(
    root: Path,
    api_workbook: Path,
    api_sheet: str,
    req_master: set[str],
    violations: list[Violation],
) -> tuple[int, set[str], list[tuple[str, str, str]]]:
    workbook, worksheet = select_api_sheet(api_workbook, api_sheet)
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    note_col = find_column(headers, [KOR_NOTE, "remark", "note"])
    role_col = find_column(headers, ["role", KOR_ROLE])
    endpoint_col = find_column(headers, ["endpoint", "path", "uri"])
    if note_col is None:
        add_violation(
            violations,
            "REQID_API_COLUMN_MISSING",
            api_workbook,
            root,
            "header",
            KOR_NOTE,
            "API sheet missing 비고/remark/note column",
        )
        workbook.close()
        return 0, set(), []

    scan_targets = [note_col]
    if role_col is not None:
        scan_targets.append(role_col)
    if endpoint_col is not None:
        scan_targets.append(endpoint_col)
    scan_max_col = max(scan_targets) + 1

    token_count = 0
    role_tokens: set[str] = set()
    corpus: list[tuple[str, str, str]] = []

    for row_number, row_values in enumerate(
        worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=scan_max_col,
            values_only=True,
        ),
        start=2,
    ):
        for col_index, cell_value in enumerate(row_values, start=1):
            if isinstance(cell_value, str) and cell_value.strip():
                cell_ref = f"{get_column_letter(col_index)}{row_number}"
                corpus.append((to_rel(api_workbook, root), f"{worksheet.title}!{cell_ref}", cell_value))

        endpoint_value = ""
        if endpoint_col is not None and endpoint_col < len(row_values):
            endpoint_cell = row_values[endpoint_col]
            if isinstance(endpoint_cell, str):
                endpoint_value = endpoint_cell.strip()

        role_cell_text = ""
        role_cell_ref = ""
        if role_col is not None and role_col < len(row_values):
            raw_role = row_values[role_col]
            if isinstance(raw_role, str) and raw_role.strip():
                role_cell_text = raw_role.strip()
                role_cell_ref = f"{get_column_letter(role_col + 1)}{row_number}"
                normalized_roles = re.split(r"[,/| ]+", role_cell_text.upper())
                for token in normalized_roles:
                    if token:
                        role_tokens.add(token)
                        if token in ACCESS_LEVEL_TAXONOMY:
                            add_violation(
                                violations,
                                "ACCESS_LEVEL_ROLE_COLUMN_FORBIDDEN",
                                api_workbook,
                                root,
                                f"sheet={worksheet.title} cell={role_cell_ref}",
                                token,
                                "PUBLIC/AUTHENTICATED must be declared in 비고 as access_level=...",
                            )

        note_cell = row_values[note_col] if note_col < len(row_values) else None
        note_value = note_cell.strip() if isinstance(note_cell, str) else ""
        note_cell_ref = f"{get_column_letter(note_col + 1)}{row_number}"
        note_key_values = parse_note_key_values(note_value)

        access_level_values = note_key_values.get("access_level", set())
        invalid_access_levels = sorted(
            token for token in access_level_values if token not in ACCESS_LEVEL_TAXONOMY
        )
        for token in invalid_access_levels:
            add_violation(
                violations,
                "ACCESS_LEVEL_STANDARD_INVALID",
                api_workbook,
                root,
                f"sheet={worksheet.title} cell={note_cell_ref}",
                f"access_level={token}",
                "access_level must be one of PUBLIC/AUTHENTICATED",
            )

        if endpoint_value.lower() == AUTH_REFRESH_ENDPOINT:
            if "PUBLIC" not in access_level_values:
                add_violation(
                    violations,
                    "ACCESS_LEVEL_REFRESH_RULE_MISSING",
                    api_workbook,
                    root,
                    f"sheet={worksheet.title} cell={note_cell_ref}",
                    "access_level=PUBLIC",
                    "/v1/auth/refresh must include access_level=PUBLIC in 비고",
                )

            auth_mechanisms = note_key_values.get("auth_mechanism", set())
            if "REFRESH_COOKIE" not in auth_mechanisms:
                add_violation(
                    violations,
                    "ACCESS_LEVEL_REFRESH_RULE_MISSING",
                    api_workbook,
                    root,
                    f"sheet={worksheet.title} cell={note_cell_ref}",
                    "auth_mechanism=refresh_cookie",
                    "/v1/auth/refresh must include auth_mechanism=refresh_cookie in 비고",
                )

            if role_cell_text:
                location = role_cell_ref or f"row={row_number}"
                add_violation(
                    violations,
                    "ACCESS_LEVEL_REFRESH_ROLE_NOT_EMPTY",
                    api_workbook,
                    root,
                    f"sheet={worksheet.title} cell={location}",
                    role_cell_text,
                    "/v1/auth/refresh role column must be blank when access_level=PUBLIC is used",
                )

        if "reqid" not in note_value.lower():
            continue

        cell_ref = note_cell_ref
        segments = REQID_TAG_RE.findall(note_value)
        segments_to_scan = segments if segments else [note_value]
        found_in_cell = 0
        malformed_in_cell: set[str] = set()

        for segment in segments_to_scan:
            strict_tokens, malformed_tokens = extract_reqid_tokens(segment)
            token_count += len(strict_tokens)
            found_in_cell += len(strict_tokens)
            for token in strict_tokens:
                if token not in req_master:
                    add_violation(
                        violations,
                        "REQID_API_UNKNOWN",
                        api_workbook,
                        root,
                        f"sheet={worksheet.title} cell={cell_ref}",
                        token,
                        "API 비고 ReqID reference is missing in Requirements SSOT",
                    )
            for token in malformed_tokens:
                malformed_in_cell.add(token)

        for token in sorted(malformed_in_cell):
            add_violation(
                violations,
                "REQID_API_MALFORMED",
                api_workbook,
                root,
                f"sheet={worksheet.title} cell={cell_ref}",
                token,
                "API 비고 contains malformed ReqID token",
            )

        if found_in_cell == 0 and not malformed_in_cell:
            add_violation(
                violations,
                "REQID_API_TAG_WITHOUT_TOKEN",
                api_workbook,
                root,
                f"sheet={worksheet.title} cell={cell_ref}",
                note_value,
                "API 비고 contains ReqID marker but no ReqID token",
            )

    workbook.close()
    return token_count, role_tokens, corpus


def iter_workbook_corpus_rows(
    worksheet,
    max_row: int,
    max_col: int,
):
    for row_number, row_values in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True),
        start=1,
    ):
        yield row_number, row_values


def scan_workbook_for_reqids(
    root: Path,
    workbook_path: Path,
    req_master: set[str],
    violations: list[Violation],
    code_prefix: str,
) -> tuple[int, list[tuple[str, str, str]]]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    token_count = 0
    corpus: list[tuple[str, str, str]] = []
    req_prefixes = {extract_prefix(token) for token in req_master}

    for worksheet in workbook.worksheets:
        if code_prefix == "UIUX" and not worksheet.title.startswith(("91_", "92_", "93_", "94_")):
            continue

        if code_prefix == "UIUX":
            max_row = min(worksheet.max_row, 1500)
            max_col = min(worksheet.max_column, 16)
        elif code_prefix == "DB":
            max_row = min(worksheet.max_row, 800)
            max_col = min(worksheet.max_column, 20)
        else:
            max_row = min(worksheet.max_row, 5000)
            max_col = min(worksheet.max_column, 40)

        reqid_columns: set[int] = set()
        for row_number, row_values in iter_workbook_corpus_rows(
            worksheet=worksheet,
            max_row=min(5, max_row),
            max_col=max_col,
        ):
            _ = row_number
            for col_index, value in enumerate(row_values, start=1):
                if not isinstance(value, str):
                    continue
                header_text = value.strip().lower()
                if "reqid" in header_text or KOR_REQ_ID.lower() in header_text:
                    reqid_columns.add(col_index)

        for row_number, row_values in iter_workbook_corpus_rows(
            worksheet=worksheet,
            max_row=max_row,
            max_col=max_col,
        ):
            for col_index, value in enumerate(row_values, start=1):
                if not isinstance(value, str):
                    continue
                text = value.strip()
                if not text:
                    continue

                location = f"{worksheet.title}!{get_column_letter(col_index)}{row_number}"
                corpus.append((to_rel(workbook_path, root), location, text))

                strict_tokens, malformed_tokens = extract_reqid_tokens(text)
                if not strict_tokens and not malformed_tokens:
                    continue

                has_explicit_reqid_marker = "reqid" in text.lower()
                has_reqid_marker = has_explicit_reqid_marker or col_index in reqid_columns
                filtered_strict = [token for token in strict_tokens if extract_prefix(token) in req_prefixes]
                filtered_malformed = [token for token in malformed_tokens if extract_prefix(token) in req_prefixes]

                if has_reqid_marker:
                    strict_to_validate = strict_tokens
                    malformed_to_validate = malformed_tokens if has_explicit_reqid_marker else filtered_malformed
                else:
                    strict_to_validate = filtered_strict
                    malformed_to_validate = []

                if not strict_to_validate and not malformed_to_validate:
                    continue

                for token in strict_to_validate:
                    token_count += 1
                    if token not in req_master:
                        add_violation(
                            violations,
                            f"REQID_{code_prefix}_UNKNOWN",
                            workbook_path,
                            root,
                            f"sheet={location}",
                            token,
                            "ReqID token was found but is missing in Requirements SSOT",
                        )
                for token in malformed_to_validate:
                    add_violation(
                        violations,
                        f"REQID_{code_prefix}_MALFORMED",
                        workbook_path,
                        root,
                        f"sheet={location}",
                        token,
                        "Malformed ReqID-like token detected",
                    )

    workbook.close()
    return token_count, corpus


def scan_csv_for_reqids(
    root: Path,
    csv_path: Path,
    req_master: set[str],
    violations: list[Violation],
    code_prefix: str,
) -> tuple[int, list[tuple[str, str, str]]]:
    headers, rows = read_csv_rows(csv_path)
    token_count = 0
    corpus: list[tuple[str, str, str]] = []
    if not headers:
        return token_count, corpus

    for row_number, row in enumerate(rows, start=2):
        joined = " | ".join(item for item in row if item)
        if not joined.strip():
            continue
        location = f"row={row_number}"
        corpus.append((to_rel(csv_path, root), location, joined))
        if "reqid" not in joined.lower():
            continue

        strict_tokens, malformed_tokens = extract_reqid_tokens(joined)

        for token in strict_tokens:
            token_count += 1
            if token not in req_master:
                add_violation(
                    violations,
                    f"REQID_{code_prefix}_UNKNOWN",
                    csv_path,
                    root,
                    location,
                    token,
                    "ReqID token was found but is missing in Requirements SSOT",
                )
        for token in malformed_tokens:
            add_violation(
                violations,
                f"REQID_{code_prefix}_MALFORMED",
                csv_path,
                root,
                location,
                token,
                "Malformed ReqID-like token detected",
            )
    return token_count, corpus


def collect_csv_corpus(root: Path, csv_path: Path) -> list[tuple[str, str, str]]:
    headers, rows = read_csv_rows(csv_path)
    if not headers:
        return []
    corpus: list[tuple[str, str, str]] = []
    for row_number, row in enumerate(rows, start=2):
        joined = " | ".join(item for item in row if item)
        if not joined.strip():
            continue
        corpus.append((to_rel(csv_path, root), f"row={row_number}", joined))
    return corpus


def normalize_cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def compile_token_pattern(token: str) -> re.Pattern[str]:
    return re.compile(rf"(?<![A-Za-z0-9_]){re.escape(token)}(?![A-Za-z0-9_])", re.IGNORECASE)


def find_token_occurrences(corpus: list[tuple[str, str, str]], token: str) -> list[tuple[str, str]]:
    pattern = compile_token_pattern(token)
    occurrences: list[tuple[str, str]] = []
    for file_path, location, text in corpus:
        if pattern.search(text):
            occurrences.append((file_path, location))
    return occurrences


def detect_uiux_placeholder_rows(
    worksheet,
    header_row: int,
    target_columns: dict[str, int],
    max_col: int,
) -> list[tuple[int, int, str, str]]:
    hits: list[tuple[int, int, str, str]] = []
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        row_values = [
            normalize_cell_text(worksheet.cell(row=row_number, column=col_index).value)
            for col_index in range(1, max_col + 1)
        ]
        if not any(row_values):
            continue

        for header_name, col_index in target_columns.items():
            value = normalize_cell_text(worksheet.cell(row=row_number, column=col_index).value)
            if value.upper() in UIUX_PLACEHOLDER_VALUES:
                token = value if value else "<blank>"
                hits.append((row_number, col_index, header_name, token))
    return hits


def detect_uiux_mapping_placeholders(
    root: Path,
    uiux_workbook: Path,
    violations: list[Violation],
) -> dict[str, object]:
    workbook = load_workbook(uiux_workbook, data_only=False, read_only=True)
    scanned_targets: list[dict[str, object]] = []
    placeholder_hits = 0

    target_sheets = [worksheet for worksheet in workbook.worksheets if worksheet.title.startswith(KOR_UIUX_UNMAPPED_PREFIX)]
    if not target_sheets:
        add_violation(
            violations,
            "PLACEHOLDER_UIUX_TARGET_SHEET_MISSING",
            uiux_workbook,
            root,
            "workbook",
            KOR_UIUX_UNMAPPED_PREFIX,
            "UIUX workbook missing expected 94_* unmapped disposition sheet",
        )
        workbook.close()
        return {"scanned_targets": [], "placeholder_hits": 0}

    target_header_markers = {
        KOR_UIUX_MAPPING_KEY: [KOR_UIUX_MAPPING_KEY],
        KOR_UIUX_MAPPING_SCREEN: [KOR_UIUX_MAPPING_SCREEN],
        KOR_UIUX_MAPPING_SHEET: [KOR_UIUX_MAPPING_SHEET],
    }

    for worksheet in target_sheets:
        max_col = min(worksheet.max_column, 24)
        header_row: int | None = None
        target_columns: dict[str, int] = {}

        for row_number in range(1, min(30, worksheet.max_row) + 1):
            row_column_matches: dict[str, int] = {}
            for col_index in range(1, max_col + 1):
                header_text = normalize_cell_text(worksheet.cell(row=row_number, column=col_index).value).lower()
                if not header_text:
                    continue
                for canonical_header, aliases in target_header_markers.items():
                    if canonical_header in row_column_matches:
                        continue
                    if any(alias in header_text for alias in aliases):
                        row_column_matches[canonical_header] = col_index

            if len(row_column_matches) == len(target_header_markers):
                header_row = row_number
                target_columns = row_column_matches
                break

        if header_row is None:
            add_violation(
                violations,
                "PLACEHOLDER_UIUX_HEADER_NOT_FOUND",
                uiux_workbook,
                root,
                f"sheet={worksheet.title}",
                "",
                "unable to locate UIUX unmapped mapping headers for placeholder scan",
            )
            continue

        scanned_target = {
            "sheet": worksheet.title,
            "header_row": header_row,
            "columns": [
                {"header": header_name, "column": get_column_letter(target_columns[header_name])}
                for header_name in sorted(target_columns)
            ],
        }
        scanned_targets.append(scanned_target)

        hits = detect_uiux_placeholder_rows(
            worksheet=worksheet,
            header_row=header_row,
            target_columns=target_columns,
            max_col=max(max_col, max(target_columns.values())),
        )
        placeholder_hits += len(hits)
        for row_number, col_index, header_name, token in hits:
            cell_ref = f"{get_column_letter(col_index)}{row_number}"
            add_violation(
                violations,
                "PLACEHOLDER_UIUX_MAPPING_VALUE",
                uiux_workbook,
                root,
                f"sheet={worksheet.title} cell={cell_ref} header={header_name}",
                token,
                "placeholder value detected in UIUX mapping column",
            )

    workbook.close()
    scanned_targets_sorted = sorted(
        scanned_targets,
        key=lambda item: (str(item["sheet"]), int(item["header_row"])),
    )
    return {
        "scanned_targets": scanned_targets_sorted,
        "placeholder_hits": placeholder_hits,
    }


def _sheet_by_prefix(workbook, prefix: str):
    for worksheet in workbook.worksheets:
        if worksheet.title.startswith(prefix):
            return worksheet
    return None


def _is_uiux_placeholder(value: str) -> bool:
    normalized = value.strip()
    if not normalized:
        return True
    upper = normalized.upper()
    if upper in UIUX_PLACEHOLDER_VALUES:
        return True
    return "TBD" in upper


def detect_uiux_error_contract(
    root: Path,
    uiux_workbook: Path,
    violations: list[Violation],
) -> dict[str, object]:
    workbook = load_workbook(uiux_workbook, data_only=False, read_only=True)
    error_sheet = _sheet_by_prefix(workbook, KOR_UIUX_ERROR_PREFIX)
    inconsistency_sheet = _sheet_by_prefix(workbook, KOR_UIUX_INCONSISTENCY_PREFIX)

    summary: dict[str, object] = {
        "error_sheet": error_sheet.title if error_sheet is not None else None,
        "inconsistency_sheet": inconsistency_sheet.title if inconsistency_sheet is not None else None,
        "target_codes": {},
        "assumptions": {},
    }

    if error_sheet is None:
        add_violation(
            violations,
            "UIUX_ERROR_SHEET_MISSING",
            uiux_workbook,
            root,
            "workbook",
            KOR_UIUX_ERROR_PREFIX,
            "UIUX workbook missing 01_ error catalog sheet",
        )
    if inconsistency_sheet is None:
        add_violation(
            violations,
            "UIUX_INCONSISTENCY_SHEET_MISSING",
            uiux_workbook,
            root,
            "workbook",
            KOR_UIUX_INCONSISTENCY_PREFIX,
            "UIUX workbook missing 90_ inconsistency sheet",
        )
    if error_sheet is None or inconsistency_sheet is None:
        workbook.close()
        return summary

    error_rows: dict[str, int] = {}
    for row_number in range(1, min(error_sheet.max_row, 2000) + 1):
        code = normalize_cell_text(error_sheet.cell(row=row_number, column=1).value)
        if code in UIUX_TARGET_ERROR_CONTRACT:
            error_rows[code] = row_number

    for code, rule in UIUX_TARGET_ERROR_CONTRACT.items():
        row_number = error_rows.get(code)
        if row_number is None:
            add_violation(
                violations,
                "UIUX_ERROR_CODE_MISSING",
                uiux_workbook,
                root,
                f"sheet={error_sheet.title}",
                code,
                "required UIUX error code row is missing",
            )
            summary["target_codes"][code] = {"exists": False}
            continue

        message_value = normalize_cell_text(error_sheet.cell(row=row_number, column=2).value)
        http_value = normalize_cell_text(error_sheet.cell(row=row_number, column=3).value)
        expected_http = str(rule["http_status"])

        summary["target_codes"][code] = {
            "exists": True,
            "row": row_number,
            "message": message_value,
            "http": http_value,
            "expected_http": expected_http,
        }

        if _is_uiux_placeholder(message_value):
            add_violation(
                violations,
                "UIUX_ERROR_MESSAGE_PLACEHOLDER",
                uiux_workbook,
                root,
                f"sheet={error_sheet.title} cell=B{row_number}",
                message_value or "<blank>",
                "UIUX error message must not contain placeholder/TBD value",
            )

        if http_value != expected_http:
            add_violation(
                violations,
                "UIUX_ERROR_HTTP_MISMATCH",
                uiux_workbook,
                root,
                f"sheet={error_sheet.title} cell=C{row_number}",
                f"{http_value}->{expected_http}",
                "UIUX HTTP status must match error code token status",
            )

    assume_rows: dict[str, int] = {}
    for row_number in range(1, min(inconsistency_sheet.max_row, 2000) + 1):
        item_id = normalize_cell_text(inconsistency_sheet.cell(row=row_number, column=1).value)
        if item_id.startswith("ASSUME-"):
            assume_rows[item_id] = row_number

    for _, rule in UIUX_TARGET_ERROR_CONTRACT.items():
        assume_id = str(rule["assume_id"])
        row_number = assume_rows.get(assume_id)
        if row_number is None:
            add_violation(
                violations,
                "UIUX_ASSUME_MISSING",
                uiux_workbook,
                root,
                f"sheet={inconsistency_sheet.title}",
                assume_id,
                "required ASSUME row is missing in 90_ sheet",
            )
            summary["assumptions"][assume_id] = {"exists": False}
            continue

        decision_b = normalize_cell_text(inconsistency_sheet.cell(row=row_number, column=2).value)
        decision_d = normalize_cell_text(inconsistency_sheet.cell(row=row_number, column=4).value)
        decision_e = normalize_cell_text(inconsistency_sheet.cell(row=row_number, column=5).value)
        decision_g = normalize_cell_text(inconsistency_sheet.cell(row=row_number, column=7).value)

        summary["assumptions"][assume_id] = {
            "exists": True,
            "row": row_number,
            "col_b": decision_b,
            "col_d": decision_d,
            "col_e": decision_e,
            "col_g": decision_g,
        }

        resolved_marked = "resolved" in decision_b.lower() and "resolved" in decision_e.lower()
        has_tbd = any("TBD" in value.upper() for value in (decision_b, decision_d, decision_e, decision_g))
        if not resolved_marked or has_tbd:
            add_violation(
                violations,
                "UIUX_ASSUME_UNRESOLVED",
                uiux_workbook,
                root,
                f"sheet={inconsistency_sheet.title} row={row_number}",
                assume_id,
                "ASSUME row must be resolved and must not contain TBD placeholders",
            )

    workbook.close()
    return summary


def validate_terminology(
    root: Path,
    corpus: list[tuple[str, str, str]],
    role_tokens: set[str],
    violations: list[Violation],
    terminology_specs: tuple[TerminologyCheckSpec, ...],
    sse_event_re: re.Pattern[str],
) -> dict[str, object]:
    required_token_hits: dict[str, int] = {}
    optional_token_hits: dict[str, int] = {}
    forbidden_variant_hits: dict[str, int] = {}

    secret_alias_hits = 0
    sse_found: set[str] = set()

    for _, _, text in corpus:
        lowered = text.lower()
        if "key_ref" in lowered or "api_key_ref" in lowered:
            secret_alias_hits += 1

        for token in sse_event_re.findall(text):
            sse_found.add(token.lower())

    for spec in terminology_specs:
        for token in spec.required_tokens:
            occurrences = find_token_occurrences(corpus, token)
            required_token_hits[token] = len(occurrences)
            if not occurrences:
                add_violation(
                    violations,
                    f"TERMINOLOGY_{spec.code_suffix}_MISSING",
                    root / "docs/references",
                    root,
                    "corpus",
                    token,
                    f"required terminology token '{token}' was not found in scanned specs",
                )

        for token in spec.optional_tokens:
            occurrences = find_token_occurrences(corpus, token)
            optional_token_hits[token] = len(occurrences)

        for variant, canonical in spec.forbidden_variants:
            occurrences = find_token_occurrences(corpus, variant)
            forbidden_variant_hits[variant] = len(occurrences)
            if occurrences:
                file_path, location = occurrences[0]
                add_violation(
                    violations,
                    f"TERMINOLOGY_{spec.code_suffix}_VARIANT",
                    resolve_path(root, file_path),
                    root,
                    location,
                    variant,
                    f"use canonical token '{canonical}' instead of '{variant}'",
                )

    unknown_roles = sorted(token for token in role_tokens if token not in ROLE_TAXONOMY)
    for token in unknown_roles:
        add_violation(
            violations,
            "TERMINOLOGY_ROLE_UNKNOWN",
            root / "docs/references",
            root,
            "API role column",
            token,
            "role token is outside ROLE taxonomy AGENT/CUSTOMER/ADMIN/OPS/SYSTEM",
        )

    missing_roles = sorted(ROLE_TAXONOMY - {token for token in role_tokens if token in ROLE_TAXONOMY})
    if missing_roles:
        add_violation(
            violations,
            "TERMINOLOGY_ROLE_MISSING",
            root / "docs/references",
            root,
            "API role column",
            ",".join(missing_roles),
            "one or more ROLE taxonomy strings were not found",
        )

    secret_ref_hits = required_token_hits.get("secret_ref", 0)
    return {
        "secret_ref_hits": secret_ref_hits,
        "secret_alias_hits": secret_alias_hits,
        "role_tokens_found": sorted(role_tokens),
        "sse_events_found": sorted(sse_found),
        "required_token_hits": {key: required_token_hits[key] for key in sorted(required_token_hits)},
        "optional_token_hits": {key: optional_token_hits[key] for key in sorted(optional_token_hits)},
        "forbidden_variant_hits": {key: forbidden_variant_hits[key] for key in sorted(forbidden_variant_hits)},
    }


def sort_violations(items: list[Violation]) -> list[Violation]:
    return sorted(
        items,
        key=lambda item: (item.code, item.file, item.location, item.token, item.message),
    )


def build_pass_artifact(root: Path, output_path: Path, payload: dict[str, object]) -> None:
    lines = [
        f"timestamp_utc={datetime.now(timezone.utc).isoformat()}",
        f"commit={git_head(root)}",
        f"status={payload['status']}",
        f"total_reqids_in_requirements={payload['summary']['total_reqids_in_requirements']}",
        (
            "total_reqid_tokens_found_in_each_spec="
            + json.dumps(payload["summary"]["reqid_token_counts"], ensure_ascii=False, sort_keys=True)
        ),
        f"invalid_tokens_count={payload['summary']['invalid_tokens_count']}",
    ]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_report(path: Path | None, content: str) -> None:
    if path is None:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def main() -> int:
    args = parse_args()
    root = Path(args.root).resolve()

    requirements_path = resolve_path(root, args.requirements)
    summary_path = resolve_path(root, args.summary)
    development_path = resolve_path(root, args.development)
    api_workbook = resolve_path(root, args.api_workbook)
    db_workbook = resolve_path(root, args.db_workbook)

    if args.uiux_workbook:
        uiux_workbook = resolve_path(root, args.uiux_workbook)
    else:
        matches = sorted(root.glob(args.uiux_glob))
        if not matches:
            raise FileNotFoundError(f"UI/UX workbook not found by glob: {args.uiux_glob}")
        uiux_workbook = matches[0]

    report_json_path = resolve_path(root, args.report_json) if args.report_json else None
    report_txt_path = resolve_path(root, args.report_txt) if args.report_txt else None
    pass_artifact_path = resolve_path(root, args.pass_artifact) if args.pass_artifact else None
    terminology_ssot_path = resolve_terminology_ssot_path(root, args.terminology_ssot)
    terminology_ssot = load_terminology_ssot(terminology_ssot_path)
    terminology_specs = build_terminology_specs(terminology_ssot)
    sse_event_re = compile_sse_event_regex(terminology_ssot.sse_event_types)

    violations: list[Violation] = []
    req_master = load_requirements(root, requirements_path, violations)

    summary_count = validate_summary(root, summary_path, req_master, violations)
    development_count, development_corpus = scan_csv_for_reqids(
        root=root,
        csv_path=development_path,
        req_master=req_master,
        violations=violations,
        code_prefix="DEVELOPMENT",
    )
    api_count, api_role_tokens, api_corpus = validate_api_notes_and_roles(
        root=root,
        api_workbook=api_workbook,
        api_sheet=args.api_sheet,
        req_master=req_master,
        violations=violations,
    )
    uiux_count, uiux_corpus = scan_workbook_for_reqids(
        root=root,
        workbook_path=uiux_workbook,
        req_master=req_master,
        violations=violations,
        code_prefix="UIUX",
    )
    db_count, db_corpus = scan_workbook_for_reqids(
        root=root,
        workbook_path=db_workbook,
        req_master=req_master,
        violations=violations,
        code_prefix="DB",
    )

    req_count = len(req_master)
    req_corpus = collect_csv_corpus(root, requirements_path)
    summary_corpus = collect_csv_corpus(root, summary_path)

    corpus = req_corpus + summary_corpus + development_corpus + api_corpus + uiux_corpus + db_corpus
    terminology_summary = validate_terminology(
        root=root,
        corpus=corpus,
        role_tokens=api_role_tokens,
        violations=violations,
        terminology_specs=terminology_specs,
        sse_event_re=sse_event_re,
    )
    placeholder_summary = detect_uiux_mapping_placeholders(
        root=root,
        uiux_workbook=uiux_workbook,
        violations=violations,
    )
    uiux_error_contract_summary = detect_uiux_error_contract(
        root=root,
        uiux_workbook=uiux_workbook,
        violations=violations,
    )

    violations = sort_violations(violations)

    reqid_token_counts = {
        "requirements": req_count,
        "summary": summary_count,
        "development": development_count,
        "api_note": api_count,
        "uiux": uiux_count,
        "db": db_count,
    }
    invalid_tokens_count = sum(1 for violation in violations if violation.code.startswith("REQID_"))

    payload = {
        "status": "PASS" if not violations else "FAIL",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "commit": git_head(root),
        "summary": {
            "total_reqids_in_requirements": len(req_master),
            "total_reqid_tokens_found_in_each_spec": reqid_token_counts,
            "reqid_token_counts": reqid_token_counts,
            "invalid_tokens_count": invalid_tokens_count,
        },
        "terminology": terminology_summary,
        "placeholder_scan": placeholder_summary,
        "uiux_error_contract": uiux_error_contract_summary,
        "violation_count": len(violations),
        "violations": [asdict(item) for item in violations],
    }

    json_report = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    write_report(report_json_path, json_report)

    lines = [
        f"{payload['status']} spec_consistency_check",
        f"total_reqids_in_requirements={payload['summary']['total_reqids_in_requirements']}",
        "total_reqid_tokens_found_in_each_spec="
        + json.dumps(reqid_token_counts, ensure_ascii=False, sort_keys=True),
        f"invalid_tokens_count={invalid_tokens_count}",
        f"placeholder_hits={placeholder_summary['placeholder_hits']}",
        "placeholder_scan_targets="
        + json.dumps(placeholder_summary["scanned_targets"], ensure_ascii=False, sort_keys=True),
        "uiux_error_contract="
        + json.dumps(uiux_error_contract_summary, ensure_ascii=False, sort_keys=True),
        f"violation_count={payload['violation_count']}",
    ]
    for item in violations:
        lines.append(f"[{item.code}] {item.file} {item.location} token={item.token} :: {item.message}")
    text_report = "\n".join(lines) + "\n"

    write_report(report_txt_path, text_report)
    if pass_artifact_path is not None and payload["status"] == "PASS":
        build_pass_artifact(root, pass_artifact_path, payload)

    sys.stdout.write(text_report)
    return 0 if payload["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
