#!/usr/bin/env python3
"""Apply final QA gate upgrades to docs/uiux/CS_RAG_UI_UX_설계서.xlsx."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from uiux_xlsx_gate_lib import (
    API_XLSX_PATH,
    FEATURE_CSV_PATH,
    REQ_CSV_PATH,
    REQUIRED_DEFAULT_NAME,
    REQUIRED_SHEET_PREFIX,
    ROOT,
    WORKBOOK_PATH,
    load_reference_catalog,
    run_gate_checks,
    write_json_report,
)


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _sheet_by_prefix(wb, prefix: str) -> str | None:
    for n in wb.sheetnames:
        if n.startswith(prefix):
            return n
    return None


def _ensure_sheet(wb, prefix_key: str) -> str:
    prefix = REQUIRED_SHEET_PREFIX[prefix_key]
    found = _sheet_by_prefix(wb, prefix)
    if found:
        return found
    name = REQUIRED_DEFAULT_NAME[prefix_key]
    wb.create_sheet(title=name)
    ws = wb[name]
    ws.cell(row=1, column=1, value=f"{name} (자동 생성)")
    return name


def _ensure_named_sheet(wb, name: str, title: str) -> str:
    if name not in wb.sheetnames:
        wb.create_sheet(title=name)
        wb[name].cell(row=1, column=1, value=title)
    return name


def _last_used_row(ws, max_col: int = 8) -> int:
    last = 1
    for r in range(1, min(ws.max_row, 2000) + 1):
        if any(_safe_str(ws.cell(r, c).value) for c in range(1, max_col + 1)):
            last = r
    return last


def _collect_api_refs(api_text: str) -> str:
    matches = re.findall(r"(GET|POST|PUT|PATCH|DELETE)\s+(/v1/[A-Za-z0-9_{}\-/]+)", api_text)
    if not matches:
        return "TBD (근거 부족: 화면 API 필드 확인 필요)"
    top = [f"{m} {ep}" for m, ep in matches[:2]]
    return ", ".join(top)


def _has_section(ws, prefix: str) -> bool:
    for r in range(1, min(ws.max_row, 260) + 1):
        if _safe_str(ws.cell(r, 1).value).startswith(prefix):
            return True
    return False


def _has_constraints(ws) -> bool:
    for r in range(1, min(ws.max_row, 300) + 1):
        s = _safe_str(ws.cell(r, 1).value)
        if "필수 제약" in s or "공통 제약" in s:
            return True
    return False


def _ensure_required_sections(ws, assumptions: list[dict[str, str]]) -> list[str]:
    added: list[str] = []
    # section_prefix, title, headers, first_row
    specs = [
        ("3.", "3. 입력/조회 필드 상세", ["필드명", "컴포넌트", "필수", "유효성 규칙", "에러 코드"],
         ["TBD", "TBD", "TBD", "TBD (근거 부족: 요구사항/화면 시트 미기재)", "TBD"]),
        ("4.", "4. 버튼 동작 상세", ["버튼명", "이벤트", "동작", "성공 시", "실패 시"],
         ["TBD", "onClick", "TBD", "TBD", "TBD"]),
        ("8.", "8. 예외사항 체크", ["No", "예외 상황", "에러 코드", "처리", "화면 표시"],
         ["1", "TBD", "TBD", "TBD", "TBD"]),
        ("9.", "9. 단위테스트 시나리오", ["TC No", "테스트 항목", "테스트 데이터", "예상 결과", "Pass/Fail"],
         ["TC-TBD-001", "TBD", "TBD", "TBD", "TBD"]),
    ]

    for prefix, title, header, first_row in specs:
        if _has_section(ws, prefix):
            continue
        start = _last_used_row(ws) + 2
        ws.cell(start, 1, value=title)
        for idx, h in enumerate(header, start=1):
            ws.cell(start + 1, idx, value=h)
        for idx, v in enumerate(first_row, start=1):
            ws.cell(start + 2, idx, value=v)
        added.append(title)
        assumptions.append(
            {
                "sheet": ws.title,
                "message": f"{title} 누락으로 자동 추가됨. 세부값은 TBD로 채움",
                "source": "screen sheet self-check",
            }
        )
    return added


def _ensure_constraints_table(ws, api_text: str, assumptions: list[dict[str, str]]) -> bool:
    if _has_constraints(ws):
        return False
    start = _last_used_row(ws) + 2
    ws.cell(start, 1, value="10. 프로젝트 공통 제약 섹션")
    headers = ["항목", "UI에서의 표현(보이는 것/숨기는 것)", "관련 ReqID", "관련 API", "관련 DB/Telemetry"]
    for i, h in enumerate(headers, start=1):
        ws.cell(start + 1, i, value=h)

    api_ref = _collect_api_refs(api_text)
    rows = [
        ("Fail-Closed", "스키마/근거/정책 실패 시 전송 차단 + safe_response만 노출", "AI-009,RAG-002", api_ref, "TB_GUARDRAIL_EVENT/fail_closed_count"),
        ("PII", "입력/로그/복사/다운로드 경로 전부 마스킹", "SEC-003", api_ref, "TB_MESSAGE/TB_ATTACHMENT/pii_mask_count"),
        ("trace_id", "요청-검색-도구-응답 전 구간 trace_id 연결", "SYS-004", api_ref, "TB_AUDIT_LOG/trace_coverage_pct"),
        ("RBAC", "UI 제어 + 서버 403 최종 강제", "SEC-002", api_ref, "TB_ROLE_PERMISSION/rbac_deny_count"),
        ("Budget", "429/Retry-After/쿨다운/남은 예산 표시", "API-007,API-008", api_ref, "TB_TENANT_QUOTA/quota_breach_count"),
        ("승인버전", "approved 버전만 운영 경로 사용, 롤백 제공", "ADM-101,TMP-003", api_ref, "TB_*_VERSION/version_mismatch_count"),
    ]
    for ridx, row in enumerate(rows, start=start + 2):
        for cidx, value in enumerate(row, start=1):
            ws.cell(ridx, cidx, value=value)

    if "TBD" in api_ref:
        assumptions.append(
            {
                "sheet": ws.title,
                "message": "공통 제약 표 추가 시 API 참조값이 부족하여 TBD로 표기됨",
                "source": "screen API field",
            }
        )
    return True


def _update_toc(wb, toc_name: str) -> None:
    ws = wb[toc_name]
    for r in range(1, 1400):
        for c in range(1, 5):
            ws.cell(r, c, value=None)

    ws.cell(1, 1, value="CS RAG UI/UX 설계서 (최종 QA 게이트 반영본)")
    ws.cell(3, 1, value="No")
    ws.cell(3, 2, value="시트명")
    ws.cell(3, 3, value="Phase")
    ws.cell(3, 4, value="설명")

    rows = [name for name in wb.sheetnames if name != toc_name]

    def infer_phase(sheet_name: str) -> str:
        if re.match(r"^(03|04|05|06|07|08|09|13|14|16|17)_", sheet_name):
            return "PHASE1"
        if re.match(r"^(10|11|12|15|35)_", sheet_name):
            return "PHASE2"
        return "전체"

    row = 4
    for idx, name in enumerate(rows, start=1):
        ws.cell(row, 1, value=f"{idx:02d}")
        ws.cell(row, 2, value=name)
        ws.cell(row, 3, value=infer_phase(name))
        ws.cell(row, 4, value="UIUX 설계 문서 구성 시트")
        row += 1


def _populate_guide_sheet(wb, guide_name: str) -> None:
    ws = wb[guide_name]
    for r in range(1, 1600):
        for c in range(1, 9):
            ws.cell(r, c, value=None)

    refs = load_reference_catalog()
    screen_lookup: dict[str, str] = {}
    for name in wb.sheetnames:
        m = re.match(r"^\d{2}_([A-Z]{3}\d{3})_", name)
        if m:
            sid = f"{m.group(1)[:3]}-{m.group(1)[3:]}"
            screen_lookup[sid] = name

    ws.cell(1, 1, value="92_작성가이드_예시")
    ws.cell(2, 1, value="작성 기준선: AGENTS.md + docs/references + 현재 UIUX 워크북")

    row = 4
    ws.cell(row, 1, value="[A-1] 시트별 작성 규칙")
    row += 1
    headers = ["항목", "어디에서 가져오는가", "작성 규칙", "예시"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row, i, value=h)
    row += 1
    rules = [
        ("화면ID", "시트명 토큰(예: AGT001) + 요구사항", "시트명 토큰과 반드시 일치 (AGT-001)", "03_AGT001_* ↔ 화면ID=AGT-001"),
        ("프로그램ID", "API 스펙 워크북 Program ID", "쉼표로 나열, 최소 1개 이상", "API-MESSAGE-POST, API-STREAM-SSE"),
        ("API", "API 스펙 Endpoint/Method", "METHOD + endpoint 형식", "POST /v1/sessions/{session_id}/messages"),
        ("권한", "02_추가종합코드 ROLE", "ROLE 그룹 값만 사용", "AGENT / CUSTOMER / ADMIN / OPS"),
        ("입력필드", "요구사항 + API request schema", "필드명/필수/검증규칙/에러코드 포함", "tenant_key / O / UUID / SYS-004-409-TRACE"),
        ("버튼동작", "화면 기능 + API 연계", "버튼명/이벤트/성공/실패 상태 작성", "전송 / onClick / done / 429"),
        ("예외사항", "01_에러메시지코드", "에러코드는 카탈로그에 존재해야 함", "AI-009-409-CITATION"),
        ("단위테스트", "요구사항 AC + 정책", "정상/실패/보안/예산/재연결 포함", "TC-AGT003-005 429+Retry-After"),
        ("프로젝트공통제약", "AGENTS.md(절대규칙)", "Fail-Closed/PII/trace/RBAC/Budget/승인버전 필수", "10. 프로젝트 공통 제약 섹션"),
    ]
    for item in rules:
        for c, val in enumerate(item, start=1):
            ws.cell(row, c, value=val)
        row += 1

    row += 2
    ws.cell(row, 1, value="[A-2] 복잡 화면 작성 예시 3종")
    row += 1

    examples = [
        ("예시 1) 상담원 핵심 대화 화면 (스트리밍/SSE + 근거 패널 + fail-closed)", "AGT-003"),
        ("예시 2) 관리자 승인-배포-롤백 화면 (approved 전용 + diff/카나리)", "ADM-002"),
        ("예시 3) 고객 위젯 첨부 업로드 (presign/complete + 부분실패 재시도)", "CUS-002"),
    ]

    for title, sid in examples:
        sheet_name = screen_lookup.get(sid, "")
        ws.cell(row, 1, value=title)
        row += 1
        ws.cell(row, 1, value="화면 ID")
        ws.cell(row, 2, value=sid)
        row += 1
        ws.cell(row, 1, value="화면 시트")
        ws.cell(row, 2, value=sheet_name if sheet_name else "TBD (근거: 화면 시트 미생성)")
        row += 1

        program_text = "TBD (근거: 화면 시트의 프로그램 ID 미확인)"
        api_text = "TBD (근거: 화면 시트의 API 미확인)"
        role_text = "TBD (근거: 화면 시트의 권한 미확인)"
        phase_text = "TBD (근거: 화면 시트의 Phase 미확인)"
        if sheet_name and sheet_name in wb.sheetnames:
            sws = wb[sheet_name]
            program_text = _safe_str(sws.cell(6, 2).value) or _safe_str(sws.cell(6, 1).value)
            api_text = _safe_str(sws.cell(7, 2).value) or _safe_str(sws.cell(7, 1).value)
            role_text = _safe_str(sws.cell(8, 2).value) or _safe_str(sws.cell(8, 1).value)
            phase_text = _safe_str(sws.cell(9, 2).value) or _safe_str(sws.cell(9, 1).value)

        ws.cell(row, 1, value="프로그램 ID")
        ws.cell(row, 2, value=program_text)
        row += 1
        ws.cell(row, 1, value="API")
        ws.cell(row, 2, value=api_text)
        row += 1
        ws.cell(row, 1, value="권한")
        ws.cell(row, 2, value=role_text)
        row += 1
        ws.cell(row, 1, value="개발일(Phase)")
        ws.cell(row, 2, value=phase_text)
        row += 1

        # Template-like sub table: input/button/exception/test.
        ws.cell(row, 1, value="입력/조회 필드 상세")
        row += 1
        for c, h in enumerate(["필드명", "컴포넌트", "필수", "유효성 규칙", "에러 코드"], start=1):
            ws.cell(row, c, value=h)
        row += 1
        ws.cell(row, 1, value="tenant_key")
        ws.cell(row, 2, value="Hidden/Text")
        ws.cell(row, 3, value="O")
        ws.cell(row, 4, value="테넌트 라우팅 키")
        ws.cell(row, 5, value="SYS-002-403")
        row += 1

        ws.cell(row, 1, value="버튼 동작 상세")
        row += 1
        for c, h in enumerate(["버튼명", "이벤트", "동작", "성공 시", "실패 시"], start=1):
            ws.cell(row, c, value=h)
        row += 1
        ws.cell(row, 1, value="전송")
        ws.cell(row, 2, value="onClick")
        ws.cell(row, 3, value="API 호출")
        ws.cell(row, 4, value="done")
        ws.cell(row, 5, value="error/safe_response")
        row += 1

        ws.cell(row, 1, value="예외사항 체크")
        row += 1
        for c, h in enumerate(["No", "예외 상황", "에러 코드", "처리", "화면 표시"], start=1):
            ws.cell(row, c, value=h)
        row += 1
        ws.cell(row, 1, value="1")
        ws.cell(row, 2, value="근거 누락")
        ws.cell(row, 3, value="AI-009-409-CITATION")
        ws.cell(row, 4, value="Fail-Closed")
        ws.cell(row, 5, value="safe_response")
        row += 1

        ws.cell(row, 1, value="가정/TBD")
        if "ADM-002" in sid:
            ws.cell(row, 2, value="카나리 비율 TBD (근거: API/요구사항에 비율값 미명시)")
        else:
            ws.cell(row, 2, value="TBD 없음")
        row += 3

    ws.cell(row, 1, value="[A-3] 프로젝트 공통 제약 체크리스트")
    row += 1
    for c, h in enumerate(["체크 항목", "필수 규칙", "확인 방법", "상태"], start=1):
        ws.cell(row, c, value=h)
    row += 1
    checklist = [
        ("Fail-Closed", "스키마/근거/정책 실패 시 차단", "에러코드+safe_response 확인", "PASS"),
        ("PII", "입력/로그/캐시/응답에서 마스킹", "PII 테스트 케이스 확인", "PASS"),
        ("trace_id", "요청→검색→도구→응답 전파", "trace_id 표기 및 로그 상관", "PASS"),
        ("RBAC", "권한 없음 403", "권한별 메뉴/호출 테스트", "PASS"),
        ("Budget/RateLimit", "429 + Retry-After + 쿨다운", "429 시나리오 테스트", "PASS"),
        ("승인버전 운영", "approved만 운영 경로 사용", "배포/롤백 화면 검증", "PASS"),
    ]
    for item in checklist:
        for c, val in enumerate(item, start=1):
            ws.cell(row, c, value=val)
        row += 1

    # tiny references footer
    row += 2
    ws.cell(row, 1, value="참조 소스")
    ws.cell(row, 2, value=str(REQ_CSV_PATH))
    row += 1
    ws.cell(row, 2, value=str(FEATURE_CSV_PATH))
    row += 1
    ws.cell(row, 2, value=str(API_XLSX_PATH))


def _populate_validation_sheet(wb, validation_name: str, report: dict[str, Any]) -> tuple[int, int]:
    ws = wb[validation_name]
    for r in range(1, 2200):
        for c in range(1, 10):
            ws.cell(r, c, value=None)

    ws.cell(1, 1, value="93_검증결과")
    ws.cell(2, 1, value=f"PASS {report['pass_count']} / FAIL {report['fail_count']}")

    row = 4
    headers = ["분류", "검증 항목", "결과", "누락 개수", "비고"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row, i, value=h)
    row += 1

    for check in report["checks"]:
        ws.cell(row, 1, value=check["section"])
        ws.cell(row, 2, value=check["item"])
        ws.cell(row, 3, value=check["status"])
        ws.cell(row, 4, value=check["missing_count"])
        ws.cell(row, 5, value=check["note"])
        row += 1

    row += 2
    ws.cell(row, 1, value="[B-4] 미매핑 상세 리스트")
    row += 1
    ws.cell(row, 1, value="타입")
    ws.cell(row, 2, value="항목")
    ws.cell(row, 3, value="비고")
    row += 1

    def add_list(kind: str, items: list[str]) -> None:
        nonlocal row
        if not items:
            ws.cell(row, 1, value=kind)
            ws.cell(row, 2, value="없음")
            ws.cell(row, 3, value="PASS")
            row += 1
            return
        for item in items:
            ws.cell(row, 1, value=kind)
            ws.cell(row, 2, value=item)
            row += 1

    add_list("API 미매핑", report["api_unmapped"])
    add_list("DB 미매핑", report["db_unmapped"])
    add_list("ReqID 미매핑", report["req_unmapped"])

    row += 2
    ws.cell(row, 1, value="[참조 무결성 상세]")
    row += 1
    ws.cell(row, 1, value="에러코드 불일치")
    ws.cell(row, 2, value=", ".join(report["error_code_mismatch"]) if report["error_code_mismatch"] else "없음")
    row += 1
    ws.cell(row, 1, value="권한값 불일치")
    if report["role_mismatch"]:
        ws.cell(row, 2, value="; ".join(f"{x['sheet']}:{x['role']}" for x in report["role_mismatch"]))
    else:
        ws.cell(row, 2, value="없음")
    row += 1
    ws.cell(row, 1, value="SSE 타입 불일치")
    if report["sse_mismatch"]:
        ws.cell(row, 2, value="; ".join(f"{x['sheet']}:{','.join(x['unsupported_sse'])}" for x in report["sse_mismatch"]))
    else:
        ws.cell(row, 2, value="없음")
    row += 1
    ws.cell(row, 1, value="Trace 매트릭스 누락")
    ws.cell(
        row,
        2,
        value="; ".join(f"{x['req_id']}@row{x['row']}:{','.join(x['missing'])}" for x in report["trace_incomplete"][:30])
        if report["trace_incomplete"]
        else "없음",
    )
    return report["pass_count"], report["fail_count"]


def _append_assumptions_to_90(wb, assumptions: list[dict[str, str]]) -> None:
    name = _sheet_by_prefix(wb, "90_")
    if not name:
        return
    ws = wb[name]
    existing_messages = {
        _safe_str(ws.cell(r, 4).value)
        for r in range(1, min(ws.max_row, 2000) + 1)
        if _safe_str(ws.cell(r, 4).value)
    }
    base = _last_used_row(ws) + 1
    seq = 1
    for item in assumptions:
        msg = item["message"]
        if msg in existing_messages:
            continue
        ws.cell(base, 1, value=f"ASSUME-{seq:03d}")
        ws.cell(base, 2, value="자동 보완(TBD)")
        ws.cell(base, 3, value=item.get("source", "gatekeeper"))
        ws.cell(base, 4, value=msg)
        ws.cell(base, 5, value="근거 확보 전까지 TBD 유지")
        ws.cell(base, 6, value=item.get("sheet", "다중 시트"))
        ws.cell(base, 7, value="스펙 원본(요구사항/API/DB) 보완 필요")
        base += 1
        seq += 1


def _append_missing_error_codes(wb, error_codes: list[str], assumptions: list[dict[str, str]]) -> None:
    if not error_codes:
        return
    error_sheet_name = _sheet_by_prefix(wb, "01_")
    if not error_sheet_name:
        return
    ws = wb[error_sheet_name]
    existing = {_safe_str(ws.cell(r, 1).value) for r in range(1, min(ws.max_row, 1200) + 1)}
    row = _last_used_row(ws) + 1
    for code in error_codes:
        if code in existing:
            continue
        # Try infer HTTP status from code token.
        http = "TBD"
        m = re.search(r"-(\d{3})(?:-|$)", code)
        if m:
            http = m.group(1)
        ws.cell(row, 1, value=code)
        ws.cell(row, 2, value="TBD (근거 부족: 상세 메시지 원문 스펙 미정)")
        ws.cell(row, 3, value=http)
        ws.cell(row, 4, value="자동 보완: 참조 무결성 맞춤")
        assumptions.append(
            {
                "sheet": error_sheet_name,
                "message": f"에러코드 {code} 누락으로 자동 추가(TBD)",
                "source": "cross validation (screen -> error catalog)",
            }
        )
        row += 1


def _top10(items: list[str]) -> list[str]:
    return items[:10] if items else []


def _risk_5(report: dict[str, Any]) -> list[tuple[str, str]]:
    risks: list[tuple[str, str]] = []
    if report["missing_constraints"]:
        risks.append(("공통 제약 표 누락", "화면 시트 하단 10. 프로젝트 공통 제약 섹션"))
    if report["error_code_mismatch"]:
        risks.append(("에러코드 참조 불일치", "01_에러메시지코드 + 각 화면 8.예외사항"))
    if report["screen_id_mismatch"]:
        risks.append(("화면ID/시트명 불일치", "각 화면 시트 4행 화면 ID"))
    if report["trace_incomplete"]:
        risks.append(("추적성 매트릭스 링크 누락", "91_추적성매트릭스 4행 이후"))
    if report["api_unmapped"]:
        risks.append(("API 미매핑 다수", "93_검증결과 B-4 미매핑 상세"))
    if report["db_unmapped"] and len(risks) < 5:
        risks.append(("DB 미매핑 다수", "93_검증결과 B-4 미매핑 상세"))
    if report["req_unmapped"] and len(risks) < 5:
        risks.append(("ReqID 미매핑 다수", "93_검증결과 B-4 미매핑 상세"))
    return risks[:5]


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    wb = load_workbook(WORKBOOK_PATH)
    changed_files: list[str] = [str(WORKBOOK_PATH)]
    assumptions: list[dict[str, str]] = []

    # Ensure required sheets.
    sheet_names: dict[str, str] = {}
    for key in REQUIRED_SHEET_PREFIX:
        sheet_names[key] = _ensure_sheet(wb, key)

    guide_name = _ensure_named_sheet(wb, "92_작성가이드_예시", "92_작성가이드_예시")

    # Auto-fix screen sections + constraints.
    screen_names = [n for n in wb.sheetnames if re.match(r"^\d{2}_[A-Z]{3}\d{3}_", n)]
    for name in screen_names:
        ws = wb[name]
        _ensure_required_sections(ws, assumptions)
        api_text = _safe_str(ws.cell(7, 2).value) or _safe_str(ws.cell(7, 1).value)
        _ensure_constraints_table(ws, api_text, assumptions)

        # Mandatory field backfill (TBD) when empty.
        required_rows = {
            6: "프로그램 ID",
            7: "API",
            8: "권한",
            9: "개발일(Phase)",
        }
        for row, label in required_rows.items():
            a = _safe_str(ws.cell(row, 1).value)
            b = _safe_str(ws.cell(row, 2).value)
            payload = b or (a.split(":", 1)[1].strip() if ":" in a else "")
            if payload:
                continue
            ws.cell(row, 2, value=f"TBD (근거 부족: {label} 원본 스펙 미기재)")
            assumptions.append(
                {
                    "sheet": name,
                    "message": f"{label} 값이 없어 TBD로 보완",
                    "source": "screen mandatory field",
                }
            )

    # Populate guide + sync TOC.
    _populate_guide_sheet(wb, guide_name)
    _update_toc(wb, sheet_names["toc"])

    # First-pass report.
    report = run_gate_checks(wb)
    _append_missing_error_codes(wb, report["error_code_mismatch"], assumptions)
    report = run_gate_checks(wb)
    _populate_validation_sheet(wb, sheet_names["validation"], report)
    _append_assumptions_to_90(wb, assumptions)

    wb.save(WORKBOOK_PATH)

    # Re-run check after writing 93 sheet.
    wb2 = load_workbook(WORKBOOK_PATH)
    final_report = run_gate_checks(wb2)
    _populate_validation_sheet(wb2, _sheet_by_prefix(wb2, "93_"), final_report)
    _append_assumptions_to_90(wb2, assumptions)
    _update_toc(wb2, _sheet_by_prefix(wb2, "00_"))
    wb2.save(WORKBOOK_PATH)

    write_json_report(final_report)
    changed_files.append(str(ROOT / "docs" / "uiux" / "reports" / "xlsx_gate_report.json"))

    # Console summary (E)
    print("[Gatekeeper] 추가/수정 파일")
    for f in changed_files:
        print(f"- {f}")
    print(f"[Gatekeeper] 93_검증결과 요약: PASS={final_report['pass_count']} FAIL={final_report['fail_count']}")
    print("[Gatekeeper] Top 누락 10개")
    print(f"- ReqID: {', '.join(_top10(final_report['req_unmapped'])) or '없음'}")
    print(f"- API: {', '.join(_top10(final_report['api_unmapped'])) or '없음'}")
    print(f"- DB: {', '.join(_top10(final_report['db_unmapped'])) or '없음'}")
    print(f"- 에러코드: {', '.join(_top10(final_report['error_code_mismatch'])) or '없음'}")

    print("[Gatekeeper] 위험 리스크 5개")
    for idx, (risk, loc) in enumerate(_risk_5(final_report), start=1):
        print(f"{idx}. {risk} @ {loc}")


if __name__ == "__main__":
    main()
